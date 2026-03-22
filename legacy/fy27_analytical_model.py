"""
Generate RSU 5 Financial Analysis Excel workbook.

Architecture:
  INPUT sheets (yellow tabs) -- raw data with source citations on every sheet.
  CALC  sheets (green tabs)  -- 100% Excel formulas referencing input sheets.

  Every statistic traces: Result cell -> formula -> input cell -> source citation.
  All math lives in Excel formulas; this script only creates the file.

Tab ordering (stakeholder-facing first):
  C-Summary | C-FY28Projection | C-FY27Scenarios | detailed C- sheets | I- sheets | Sources

Each sheet is built by its own function for maintainability.
Cross-sheet row references are stored on the R namespace object.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#  STYLES
# ═══════════════════════════════════════════════════════════════
BOLD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=14)
SECTION = Font(bold=True, size=12, color="1F4E79")
NOTE = Font(italic=True, size=10, color="666666")
SOURCE = Font(italic=True, size=10, color="0070C0")
RESULT_FONT = Font(bold=True, size=11, color="C00000")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL   = PatternFill("solid", fgColor="E2EFDA")
RESULT_FILL = PatternFill("solid", fgColor="FCE4D6")
SOURCE_FILL = PatternFill("solid", fgColor="DAEEF3")
PARAM_FILL  = PatternFill("solid", fgColor="FFFF00")  # bright yellow for user-adjustable
THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
USD  = '#,##0'
USD2 = '#,##0.00'
PCT  = '0.0%'
PCT2 = '0.00%'
SIGNED = '+#,##0;-#,##0;0'


# ═══════════════════════════════════════════════════════════════
#  SHEET NAME CONSTANTS (for cross-sheet formulas)
# ═══════════════════════════════════════════════════════════════
SN = {
    'ie':  "'I-Enrollment'",
    'ib':  "'I-Budget'",
    'ir':  "'I-Revenue'",
    'it':  "'I-Tax'",
    'ic':  "'I-ECCosts'",
    'ia':  "'I-Assumptions'",
    'ird': "'I-FY27Reductions'",
    'icg': "'I-CostGrowth'",
    'ift': "'I-FTE'",
    'ce':  "'C-Enrollment'",
    'cb':  "'C-Budget'",
    'cc':  "'C-Consumption'",
    'cr':  "'C-Revenue'",
    'cm':  "'C-MSConsol'",
    'cec': "'C-ECAnalysis'",
    'c27': "'C-FY27Scenarios'",
    'c28': "'C-FY28Projection'",
    'cfp': "'C-FY28Paths'",
    'crm': "'C-RiskModel'",
    'iw':  "'I-Withdrawal'",
    'ieq': "'I-Equity'",
    'ci':  "'C-Independence'",
    'ceq': "'C-Equity'",
    'cla': "'C-LegalAnalysis'",
    'ids': "'I-DOEStaffing'",
    'cds': "'C-DOEStaffing'",
}


# ═══════════════════════════════════════════════════════════════
#  CROSS-SHEET ROW REFERENCE NAMESPACE
# ═══════════════════════════════════════════════════════════════
class _Refs:
    """Stores row numbers set during sheet construction."""
    pass

R = _Refs()


# ═══════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════
def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True)


def dat(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN
    if fill:
        cell.fill = fill
    return cell


def ttl(ws, r, text):
    ws.cell(r, 1, text).font = TITLE
    return r + 1


def sec(ws, r, text):
    ws.cell(r, 1, text).font = SECTION
    return r + 1


def source_block(ws, r, lines):
    for line in lines:
        cell = ws.cell(r, 1, line)
        cell.font = SOURCE
        cell.fill = SOURCE_FILL
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    return r


def note(ws, r, text):
    ws.cell(r, 1, text).font = NOTE
    return r + 1


def put(ws, r, c, val, fmt=None, fill=INPUT_FILL, font=None):
    cell = ws.cell(r, c, val)
    if fmt:
        cell.number_format = fmt
    cell.fill = fill
    cell.border = THIN
    if font:
        cell.font = font
    return cell


def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
#  INPUT SHEETS
# ═══════════════════════════════════════════════════════════════

def build_i_enrollment(wb):
    ws = wb.create_sheet("I-Enrollment")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [32, 14, 14, 14, 14, 14, 14])

    r = ttl(ws, 1, "INPUT: Enrollment Data")
    r = source_block(ws, r + 1, [
        "SOURCE: FY27 Budget Handbook, RSU 5 Superintendent's Office, pp.3-4 (02/11/2026) [Fn 1]",
        "\"Projected Enrollment 2026-2027\" table and Teacher/Class Size table.",
        "URL: https://resources.finalsite.net/images/v1770852522/rsu5org/ytm49nnwpbrzi4oiwg4i/FY27BUDGETHANDBOOK02112026.pdf",
        "All enrollment figures are October 1 counts as reported by RSU 5.",
    ])
    r += 1

    r = sec(ws, r, "1A. Historical Enrollment by School")
    heads = ["School", "Grades", "Oct 2023", "Oct 2024", "Oct 2025", "Oct 2026P"]
    for i, h in enumerate(heads, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(heads))
    r += 1

    schools = [
        ("MSS (Morse Street)",      "PreK-2", 316, 288, 275, 274),
        ("MLS (Mast Landing)",      "3-5",    281, 266, 282, 264),
        ("PES (Pownal Elementary)", "PreK-5",  89,  97,  98, 105),
        ("DCS (Durham Community)",  "PreK-8", 473, 466, 453, 467),
        ("FMS (Freeport Middle)",   "6-8",    293, 288, 286, 306),
        ("FHS (Freeport High)",     "9-12",   632, 592, 577, 554),
    ]
    R.E_START = r
    for name, grades, *years in schools:
        put(ws, r, 1, name)
        put(ws, r, 2, grades)
        for j, y in enumerate(years, 3):
            put(ws, r, j, y, USD)
        r += 1
    R.E_END = r - 1
    R.MSS_R, R.MLS_R, R.PES_R, R.DCS_R, R.FMS_R, R.FHS_R = range(R.E_START, R.E_START + 6)
    r += 1

    r = sec(ws, r, "1B. PES Grade-Level Enrollment (for cohort estimation)")
    r = note(ws, r, "Source: same Budget Handbook, Teacher/Class Size table, p.4")
    for i, h in enumerate(["Grade", "Students", "Teachers"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1

    pes_grades = [
        ("PreK", 16, 0.5), ("K", 14, 1), ("1", 14, 1),
        ("2", 13, 1), ("3", 17, 1), ("4", 13, 1), ("5", 18, 1),
    ]
    R.PG_START = r
    for g, s, t in pes_grades:
        put(ws, r, 1, g)
        put(ws, r, 2, s, USD)
        put(ws, r, 3, t, '0.0')
        r += 1
    R.PG_END = r - 1
    R.PG_K_START = R.PG_START + 1
    r += 1

    r = sec(ws, r, "1C. DCS Middle-Grade Enrollment")
    r = note(ws, r, "Source: same Budget Handbook, Teacher/Class Size table")
    for i, h in enumerate(["Grade", "Students"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 2)
    r += 1
    R.DCS_G6_R = r
    put(ws, r, 1, "Grade 6"); put(ws, r, 2, 41, USD); r += 1
    R.DCS_G7_R = r
    put(ws, r, 1, "Grade 7"); put(ws, r, 2, 46, USD); r += 1
    R.DCS_G8_R = r
    put(ws, r, 1, "Grade 8"); put(ws, r, 2, 54, USD); r += 1
    r += 1

    r = sec(ws, r, "1D. FMS Grade-Level Enrollment")
    r = note(ws, r, "Source: same Budget Handbook. FMS 6th total includes Pownal + Freeport students.")
    for i, h in enumerate(["Grade", "Students"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 2)
    r += 1
    R.FMS_G6_R = r
    put(ws, r, 1, "Grade 6 (all towns)"); put(ws, r, 2, 116, USD); r += 1
    R.FMS_G7_R = r
    put(ws, r, 1, "Grade 7"); put(ws, r, 2, 95, USD); r += 1
    R.FMS_G8_R = r
    put(ws, r, 1, "Grade 8"); put(ws, r, 2, 95, USD); r += 1


def build_i_budget(wb):
    ws = wb.create_sheet("I-Budget")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [35, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14])

    r = ttl(ws, 1, "INPUT: FY27 Proposed Budget by Cost Center")
    r = source_block(ws, r + 1, [
        "SOURCE: FY27 Superintendent's Proposed Budget Articles, 89-page line-item detail (02/11/2026) [Fn 2]",
        "URL: https://resources.finalsite.net/images/v1770852540/rsu5org/spzuwjskm8vh3t9bjck8/"
        "2026-2027SuperintendentsRecommendedBudgetArticlesforWebsite02112026.pdf",
        "Each cell value is the total for that school x article combination as reported in the budget articles.",
        "SYS = system-wide allocation; K8 = K-8 allocation; 912 = 9-12 allocation; CTRL = central office.",
    ])
    r += 1

    heads = ["Article", "DCS", "MSS", "PES", "MLS", "FMS", "FHS", "SYS", "K8", "912", "CTRL"]
    for i, h in enumerate(heads, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(heads))
    r += 1

    articles_data = [
        ("Art 1 - Regular Instruction",
         [4611862, 2506278, 1196683, 2338114, 3201928, 5310140, 0, 420852, 287336, 0]),
        ("Art 2 - Special Education",
         [1553029, 1037964, 329115, 971107, 1062320, 1013463, 836683, 347140, 359601, 0]),
        ("Art 4 - Other Instruction",
         [90140, 7710, 13125, 12626, 277908, 769494, 0, 1536, 0, 0]),
        ("Art 5 - Student & Staff Support",
         [581386, 400046, 219819, 363844, 490321, 1061829, 1523668, 249421, 16032, 0]),
        ("Art 7 - School Administration",
         [515193, 421185, 257979, 400206, 462558, 580903, 0, 0, 0, 0]),
        ("Art 9 - Facilities & Maintenance",
         [683965, 569640, 253384, 482373, 641780, 1745609, 1564602, 0, 0, 52424]),
    ]
    R.ART_START = r
    for art_name, vals in articles_data:
        put(ws, r, 1, art_name)
        for j, v in enumerate(vals, 2):
            put(ws, r, j, v, USD)
        r += 1
    R.ART_END = r - 1
    R.ART5_ROW = R.ART_START + 3
    R.ART7_ROW = R.ART_START + 4
    R.ART9_ROW = R.ART_START + 5
    r += 1

    r = sec(ws, r, "System-Only Articles (not allocated to individual schools)")
    r = source_block(ws, r, [
        "SOURCE: Same 89-page budget articles document [Fn 2].",
        "These articles are district-wide and have no school-level breakdown.",
    ])
    for i, h in enumerate(["Article", "Amount"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 2)
    r += 1

    sys_articles = [
        ("Art 3 - CTE", 337282),
        ("Art 6 - System Administration", 1308008),
        ("Art 8 - Transportation", 2388457),
        ("Art 10 - Debt Service", 1071577),
        ("Art 11 - All Other", 69796),
    ]
    R.SYS_START = r
    for name, val in sys_articles:
        put(ws, r, 1, name)
        put(ws, r, 2, val, USD)
        r += 1
    R.SYS_END = r - 1
    r += 1

    r = sec(ws, r, "Budget Totals")
    r = source_block(ws, r, [
        "FY27 total: Revised Budget Handbook 02/11/2026. Articles 1-11 + Adult Ed = $47,357,441 [Fn 25]",
        "URL: https://www.rsu5.org/fs/resource-manager/view/78dce71e-a8e6-4435-b33f-9746e8541a3a",
        "FY26 adopted: $44,455,929, approved by voters 06/10/2025 (817 Yes, 494 No) [Fn 9]",
    ])
    for i, h in enumerate(["Item", "Amount"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 2)
    r += 1
    R.ADULT_ED_R = r
    put(ws, r, 1, "Adult Education"); put(ws, r, 2, 88000, USD); r += 1
    R.FY26_R = r
    put(ws, r, 1, "FY26 Adopted Budget"); put(ws, r, 2, 44455929, USD); r += 1


def build_i_revenue(wb):
    ws = wb.create_sheet("I-Revenue")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [38, 18, 18, 18])

    r = ttl(ws, 1, "INPUT: Revenue by Town (FY27 Proposed)")
    r = source_block(ws, r + 1, [
        "SOURCE: FY27 Budget Handbook, RSU 5, Budget Impact Summary pp.9-10 (02/11/2026) [Fn 3]",
        "URL: https://resources.finalsite.net/images/v1770852522/rsu5org/ytm49nnwpbrzi4oiwg4i/"
        "FY27BUDGETHANDBOOK02112026.pdf",
        "Revenue components as reported on the cost-sharing summary page.",
        "Shared Revenue total from same page; allocated by ALM percentage.",
    ])
    r += 1

    for i, h in enumerate(["Component", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    rev_items = [
        ("Required Local Contribution (RLC)", 2257247.34, 3910950.09, 14173183.75),
        ("Additional Local Monies (ALM)",     2190978.34, 3724663.18, 11473075.46),
        ("State Aid",                          567179.79, 5991563.92,  1463888.50),
        ("Non-Shared Debt Service",                 0.00,  117175.57,        0.00),
    ]
    R.REV_START = r
    for label, p, d, f_ in rev_items:
        put(ws, r, 1, label)
        put(ws, r, 2, p, USD2)
        put(ws, r, 3, d, USD2)
        put(ws, r, 4, f_, USD2)
        r += 1
    R.REV_END = r - 1
    R.RLC_R = R.REV_START
    R.ALM_R = R.REV_START + 1
    R.AID_R = R.REV_START + 2
    R.DEBT_R = R.REV_START + 3
    r += 1

    R.SHARED_R = r
    put(ws, r, 1, "Total Shared Revenue")
    put(ws, r, 2, 1431699.07, USD2)
    r = note(ws, r + 1, "Shared revenue is allocated by each town's ALM percentage (computed on Calc sheets).")


def build_i_tax(wb):
    ws = wb.create_sheet("I-Tax")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [42, 20, 50])

    r = ttl(ws, 1, "INPUT: Tax Rates & Property Valuations")
    r += 1

    r = sec(ws, r, "POWNAL")
    r = source_block(ws, r, [
        "SOURCE: Pownal FY26 Real Estate Tax Commitment Book, committed 07/29/2025 [Fn 4]",
        "Tax rate: $15.300/thousand. Town of Pownal Assessor, https://www.pownalmaine.org/",
        "Tax breakdown: RSU 58.4%, County 3.3%, Town 38.3% [Fn 12].",
        "State valuation: $399,866,667 [Fn 3].",
    ])
    for i, h in enumerate(["Item", "Value", "Source Detail"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1
    R.POW_MIL_R = r
    put(ws, r, 1, "Total Mil Rate"); put(ws, r, 2, 15.300, '0.000'); put(ws, r, 3, "Commitment Book [Fn 4]"); r += 1
    R.POW_RSU_PCT_R = r
    put(ws, r, 1, "RSU Share of Tax"); put(ws, r, 2, 0.584, PCT); put(ws, r, 3, "Municipal records [Fn 12]"); r += 1
    put(ws, r, 1, "County Share of Tax"); put(ws, r, 2, 0.033, PCT); put(ws, r, 3, "Municipal records [Fn 12]"); r += 1
    put(ws, r, 1, "Town Share of Tax"); put(ws, r, 2, 0.383, PCT); put(ws, r, 3, "Municipal records [Fn 12]"); r += 1
    R.POW_TAXABLE_R = r
    put(ws, r, 1, "Taxable Valuation"); put(ws, r, 2, 392398240, USD); put(ws, r, 3, "Commitment Book final total [Fn 4]"); r += 1
    R.POW_STATE_R = r
    put(ws, r, 1, "State Valuation"); put(ws, r, 2, 399866667, USD); put(ws, r, 3, "Budget Handbook p.10 [Fn 3]"); r += 1
    r += 1

    r = sec(ws, r, "FREEPORT")
    r = source_block(ws, r, [
        "SOURCE: Freeport FY26 tax rate $13.85/thousand at 100% assessment ratio [Fn 5]",
        "Freeport Assessor, committed 09/15/2025. https://www.freeportmaine.com/158/Assessor",
        "Component breakdown from FY25 Budget Presentation (FY26 detail not yet public) [Fn 6]",
        "Valuations from FY27 Budget Handbook [Fn 3].",
    ])
    for i, h in enumerate(["Item", "Value", "Source Detail"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1
    R.FRE_MIL_R = r
    put(ws, r, 1, "Total Mil Rate"); put(ws, r, 2, 13.850, '0.000'); put(ws, r, 3, "Freeport Assessor [Fn 5]"); r += 1
    put(ws, r, 1, "RSU Share (FY25 proxy)"); put(ws, r, 2, 0.716, PCT); put(ws, r, 3, "FY25 Budget Presentation [Fn 6]"); r += 1
    R.FRE_TAXABLE_R = r
    put(ws, r, 1, "Taxable Valuation"); put(ws, r, 2, 2613679115, USD); put(ws, r, 3, "Budget Handbook [Fn 3]"); r += 1
    R.FRE_STATE_R = r
    put(ws, r, 1, "State Valuation"); put(ws, r, 2, 2510750000, USD); put(ws, r, 3, "Budget Handbook [Fn 3]"); r += 1
    r += 1

    r = sec(ws, r, "DURHAM")
    r = source_block(ws, r, [
        "SOURCE: Durham FY26 tax rate $33.58/thousand, ~53% assessment ratio [Fn 7]",
        "Durham Assessor, committed 08/12/2025. https://durhammaine.gov/pages/assessing",
        "Note: computed assessment ratio (taxable/state val) = 56.2%, vs. Assessor's stated ~53%.",
        "Valuations from FY27 Budget Handbook [Fn 3].",
    ])
    for i, h in enumerate(["Item", "Value", "Source Detail"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1
    R.DUR_MIL_R = r
    put(ws, r, 1, "Total Mil Rate"); put(ws, r, 2, 33.580, '0.000'); put(ws, r, 3, "Durham Assessor [Fn 7]"); r += 1
    R.DUR_TAXABLE_R = r
    put(ws, r, 1, "Taxable Valuation"); put(ws, r, 2, 389646650, USD); put(ws, r, 3, "Budget Handbook [Fn 3]"); r += 1
    R.DUR_STATE_R = r
    put(ws, r, 1, "State Valuation"); put(ws, r, 2, 692816667, USD); put(ws, r, 3, "Budget Handbook [Fn 3]"); r += 1
    r += 1

    r = sec(ws, r, "MAINE EPS FUNDING (FY25-26)")
    r = source_block(ws, r, [
        "SOURCE: Maine DOE FY25-26 Warrant Article F: Education Subsidy Information (09/09/2025) [Fn 8]",
        "URL: https://www.maine.gov/doe/sites/maine.gov.doe/files/inline-files/"
        "School%20Finance%20-%20FY25-26%20Warrant%20Article%20F%20Education%20Subsidy%20"
        "Information%20for%20Property%20Tax%20Bill%20-%209.9.2025.pdf",
    ])
    heads = ["Town", "Total EPS Allocation", "Local Share", "EPS Mill Rate", "State Share"]
    for i, h in enumerate(heads, 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1
    R.EPS_DATA_START = r
    for town, total, local, mill, state in [
        ("Pownal",   2682828.98, 2179326.67, 6.10,  503502.31),
        ("Durham",   9123076.68, 3655120.00, 6.10, 5467956.68),
        ("Freeport",15033673.30,13799840.67, 5.95, 1233832.63),
    ]:
        put(ws, r, 1, town)
        put(ws, r, 2, total, USD2)
        put(ws, r, 3, local, USD2)
        put(ws, r, 4, mill, '0.00')
        put(ws, r, 5, state, USD2)
        r += 1
    R.EPS_POW_R = R.EPS_DATA_START
    R.EPS_DUR_R = R.EPS_DATA_START + 1
    R.EPS_FRE_R = R.EPS_DATA_START + 2


def build_i_ec_costs(wb):
    ws = wb.create_sheet("I-ECCosts")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [42, 16, 10, 16, 50])

    r = ttl(ws, 1, "INPUT: Early Childhood (CDS Transition) Costs")
    r = source_block(ws, r + 1, [
        "SOURCE: RSU 5 Service Model Options PK3, presented at Dec 18, 2025 Task Force meeting [Fn 19]",
        "URL: https://www.rsu5.org/fs/resource-manager/view/da1426f4-8cc1-4b09-bd89-477da9c623ed",
        "Task Force membership and schedule: https://www.rsu5.org/quick-links/early-childhood-planning-cds [Fn 18]",
        "EC mandate: Portland Press Herald, 10/02/2025 [Fn 17]. Deadline: July 1, 2028 statewide;",
        "RSU 5 committed to Cohort 3 (July 2027) per Superintendent's 02/11/2026 presentation [Fn 10].",
        "NOTE: EC costs are FY28 expenses (begin July 2027). They do NOT appear in the FY27 budget.",
    ])
    r += 1

    r = sec(ws, r, "Option 1: School-Based with SpEd Classroom (Task Force Recommended)")
    for i, h in enumerate(["Position", "Unit Cost", "Count", "Total"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    ec_staff = [
        ("ECSE Coordinator (Asst. Director)", 176000, 1),
        ("Office Support / Transport Coord.", 105000, 1),
        ("SpEd Teacher (282B)", 144000, 1),
        ("Ed Techs", 53000, 3),
        ("Speech-Language Pathologist", 177000, 1),
        ("Social Worker", 177000, 1),
        ("Driver Hours", 13000, 1),
        ("Contractors (Psych, TOD, TVI, etc.)", 177000, 1),
    ]
    R.EC_START = r
    for pos, cost, count in ec_staff:
        put(ws, r, 1, pos)
        put(ws, r, 2, cost, USD)
        put(ws, r, 3, count)
        put(ws, r, 4, cost * count, USD)
        r += 1
    R.EC_END = r - 1
    r += 1
    R.EC_EQUIP_R = r
    put(ws, r, 1, "Equipment (van, tech, assessments)")
    put(ws, r, 4, 61644, USD)
    r += 2

    r = sec(ws, r, "Estimated State/Federal Offsets")
    r = source_block(ws, r, [
        "CDS budget transfer: State transfers existing CDS funding per child served [Fn 17].",
        "IDEA Part B: Federal special ed dollars follow children to districts (federal guidelines).",
        "EPS PreK subsidy: Maine DOE Public Pre-K Guidebook (01/21/2025) [Fn 15].",
        "NOTE: These are estimated ranges. Actual amounts depend on enrollment and state allocations.",
    ])
    for i, h in enumerate(["Offset Source", "Low Estimate", "High Estimate"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1
    R.OFF_START = r
    for label, low, high in [
        ("CDS budget transfer to districts [Fn 17]", 200000, 400000),
        ("IDEA Part B (federal 3-5 funding)", 100000, 150000),
        ("EPS subsidy for PreK students [Fn 15]", 50000, 100000),
    ]:
        put(ws, r, 1, label)
        put(ws, r, 2, low, USD)
        put(ws, r, 3, high, USD)
        r += 1
    R.OFF_END = r - 1


def build_i_assumptions(wb):
    ws = wb.create_sheet("I-Assumptions")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [50, 18, 60])

    r = ttl(ws, 1, "INPUT: Analyst Assumptions & Estimates")
    r = source_block(ws, r + 1, [
        "These are assumptions used in scenario modeling. They are NOT from official RSU 5 documents.",
        "Each assumption is sourced where possible; otherwise marked as analyst estimate.",
        "Users should adjust these values to test sensitivity of results.",
    ])
    r += 1

    for i, h in enumerate(["Assumption", "Value", "Source / Rationale"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1

    assumptions = [
        ("Cost per FTE (salary + benefits)", 107700, "Derived from budget articles average compensation [Fn 2]"),
        ("7-8 target class size", 22, "RSU 5 current middle school practice [Fn 1]"),
        ("6th grade target class size", 20, "Upper elementary norm; lower than MS to reflect transition year"),
        ("Extracurricular consolidation savings", 40000, "Analyst estimate: consolidated athletics/activities for 7-8"),
        ("Bus route cost (standard, annual)", 75000, "Analyst estimate: driver + fuel + maintenance per route"),
        ("Freeport 6th to DCS routes needed", 2, "~101 students, ~9 miles; 2 bus routes required"),
        ("Pownal K-6 to DCS routes needed (Path B)", 3, "~104 students from rural Pownal, ~11 miles"),
        ("Freeport PreK to PES routes needed (Path B)", 3, "~64 students, specialized vehicles, ~13 miles"),
        ("Durham PreK to PES routes needed (Path B)", 2, "~48 students, specialized vehicles, ~11 miles"),
        ("PES partial conversion cost (Path B)", 425000, "K-5 to EC; bathrooms, playground, plumbing [Fn 21][Fn 24]"),
        ("PES full conversion multiplier", 1.75, "Full EC center requires most rooms converted vs. partial"),
        ("Conversion amortization period (years)", 10, "Standard capital project amortization"),
        ("DCS multi-community coordination cost", 75000, "Analyst estimate: admin time for Pownal family integration"),
        ("Support services efficiency (PES closure)", 0.50, "50% of PES Art 5 saved; remainder transfers to receiving schools"),
        ("Additional efficiencies (Path A, low)", 325000, "Transport optimization, facilities phasing, attrition"),
        ("Additional efficiencies (Path A, high)", 650000, "Same levers, optimistic scenario"),
        ("Population: Pownal [Fn 13]", 1590, "Census/ACS 2023 estimate"),
        ("Population: Durham [Fn 13]", 4339, "Census/ACS 2023 estimate"),
        ("Population: Freeport [Fn 13]", 8771, "Census/ACS 2023 estimate"),
        ("Birth rate per 1,000 [Fn 14]", 9.5, "Cumberland/Androscoggin County avg, CDC WONDER"),
    ]
    R.ASM_START = r
    for label, val, src in assumptions:
        put(ws, r, 1, label)
        if isinstance(val, float) and val < 1:
            put(ws, r, 2, val, PCT)
        elif isinstance(val, float):
            put(ws, r, 2, val, '0.0')
        elif val >= 1000:
            put(ws, r, 2, val, USD)
        else:
            put(ws, r, 2, val)
        put(ws, r, 3, src)
        r += 1
    R.ASM_END = r - 1

    R.A_COST_FTE    = R.ASM_START
    R.A_CS_78       = R.ASM_START + 1
    R.A_CS_6        = R.ASM_START + 2
    R.A_EXTRACURR   = R.ASM_START + 3
    R.A_BUS_COST    = R.ASM_START + 4
    R.A_FRE6_ROUTES = R.ASM_START + 5
    R.A_POW_K6_ROUTES = R.ASM_START + 6
    R.A_FRE_PK_ROUTES = R.ASM_START + 7
    R.A_DUR_PK_ROUTES = R.ASM_START + 8
    R.A_PES_PARTIAL = R.ASM_START + 9
    R.A_PES_MULT    = R.ASM_START + 10
    R.A_AMORT       = R.ASM_START + 11
    R.A_DCS_COORD   = R.ASM_START + 12
    R.A_SUPPORT_EFF = R.ASM_START + 13
    R.A_EFF_LOW     = R.ASM_START + 14
    R.A_EFF_HIGH    = R.ASM_START + 15


def build_i_fy27_reductions(wb):
    """NEW: All T1/T2 cuts already baked into the FY27 proposed budget."""
    ws = wb.create_sheet("I-FY27Reductions")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [8, 48, 12, 16, 16, 16])

    r = ttl(ws, 1, "INPUT: FY27 Proposed Budget Reductions")
    r = source_block(ws, r + 1, [
        "SOURCE: FY27 Budget Handbook, pp.6-7 Expenditure/Reductions Summary (02/11/2026) [Fn 25]",
        "URL: https://resources.finalsite.net/images/v1770852522/rsu5org/ytm49nnwpbrzi4oiwg4i/FY27BUDGETHANDBOOK02112026.pdf",
        "These reductions are ALREADY INCLUDED in the $47,357,441 proposed budget.",
        "The initial request was $49,148,089 (10.55% increase). Reductions bring it to 6.53%.",
        "Rating: M=Maintenance, N=Necessary, R=Recommended, T1=Tier 1 Cut, T2=Tier 2 Cut.",
    ])
    r += 1

    r = sec(ws, r, "Budget Timeline")
    r = note(ws, r, "Board Adopts FY27 Budget: March 25, 2026")
    r = note(ws, r, "Annual Budget Meeting (voter vote): May 13, 2026")
    r = note(ws, r, "Budget Validation Referendum: June 19, 2026")
    r = note(ws, r, "EC Cohort 3 deadline: July 1, 2027 (FY28, not FY27)")
    r += 1

    r = sec(ws, r, "Initial Request vs. Proposed")
    for i, h in enumerate(["Item", "Amount"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 2)
    r += 1
    R.RD_INIT_INCR = r
    put(ws, r, 1, "Initial request increase (10.55%)")
    put(ws, r, 2, 4692160, USD); r += 1
    R.RD_TOTAL_REDUCTIONS = r
    put(ws, r, 1, "Total reductions applied")
    put(ws, r, 2, -1715649, SIGNED); r += 1
    R.RD_RESERVES = r
    put(ws, r, 1, "Redirected to reserves")
    put(ws, r, 2, -75000, SIGNED); r += 1
    R.RD_NET_INCR = r
    put(ws, r, 1, "NET proposed increase (6.53%)")
    put(ws, r, 2, 2901512, USD); r += 1
    r += 1

    r = sec(ws, r, "Tier 1 Reductions (Staff Positions)")
    for i, h in enumerate(["Tier", "Description", "School", "Initial", "Reduction", "Proposed"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 6)
    r += 1

    t1_items = [
        ("T1", "1 FTE Grade 1 Educator",            "DCS", 0, -107700, -107700),
        ("T1", "1 FTE Grade 2 Educator",            "MSS", 0, -107700, -107700),
        ("T1", "1 FTE Grade 3 Educator",            "MLS", 0, -107700, -107700),
        ("T1", ".25 FTE Math RTI Educator",          "FMS", 0,  -26925,  -26925),
        ("T1", "16.25hr Office Support",             "FMS", 0,  -26756,  -26756),
        ("T1", ".5 FTE ELA Educator",                "FMS", 0,  -53850,  -53850),
        ("T1", ".25 FTE ELA Educator",               "FMS", 0,  -26925,  -26925),
        ("T1", "1 FTE French Educator",              "FMS", 0, -107700, -107700),
        ("T1", "1 FTE Math Educator",                "FHS", 0, -107700, -107700),
        ("T1", "Printing Clerk",                     "FHS", 0,  -21738,  -21738),
        ("T1", "1 FTE Latin Educator",               "FHS", 0, -107700, -107700),
        ("T1", ".5 FTE Community Outreach Coord.",   "FHS", 0,  -57059,  -57059),
        ("T1", "1 FTE ESOL Educator",                "DW",  0, -107700, -107700),
    ]
    R.T1_START = r
    for tier, desc, sch, init, red, prop in t1_items:
        put(ws, r, 1, tier); put(ws, r, 2, desc); put(ws, r, 3, sch)
        put(ws, r, 4, init, USD); put(ws, r, 5, red, SIGNED); put(ws, r, 6, prop, SIGNED)
        r += 1
    R.T1_END = r - 1
    r += 1

    r = sec(ws, r, "Tier 2 Reductions (Support Positions)")
    for i, h in enumerate(["Tier", "Description", "School", "Initial", "Reduction", "Proposed"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 6)
    r += 1

    t2_items = [
        ("T2", "32.5hr Ed Tech (Bench Sub)",        "DCS", 0, -52600, -52600),
        ("T2", "32.5hr Ed Tech (RTI)",               "MSS", 0, -52600, -52600),
        ("T2", "32.5hr Ed Tech (RTI)",               "MLS", 0, -52600, -52600),
        ("T2", "32.5hr Ed Tech (General Ed)",        "FMS", 0, -52600, -52600),
        ("T2", "32.5hr Ed Tech (General Ed)",        "FMS", 0, -52600, -52600),
        ("T2", "1 Instructional Strategist (K-5)",   "DW",  0, -107700, -107700),
        ("T2", "School Nutrition",                    "DW",  0, -69796, -69796),
    ]
    R.T2_START = r
    for tier, desc, sch, init, red, prop in t2_items:
        put(ws, r, 1, tier); put(ws, r, 2, desc); put(ws, r, 3, sch)
        put(ws, r, 4, init, USD); put(ws, r, 5, red, SIGNED); put(ws, r, 6, prop, SIGNED)
        r += 1
    R.T2_END = r - 1
    r += 1

    r = sec(ws, r, "Other Reductions")
    for i, h in enumerate(["Tier", "Description", "School", "Initial", "Reduction", "Proposed"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 6)
    r += 1
    R.OTH_START = r
    other_items = [
        ("R/T1/T2", "Technology (software, devices, AV)", "DW", 336758, -250000, 86758),
        ("R/T1",    "Facilities Maintenance (paint/carpet)", "DW", 116000, -58000, 58000),
        ("R",       "Facilities Study",                    "DW", 75000, -75000, 0),
    ]
    for tier, desc, sch, init, red, prop in other_items:
        put(ws, r, 1, tier); put(ws, r, 2, desc); put(ws, r, 3, sch)
        put(ws, r, 4, init, USD); put(ws, r, 5, red, SIGNED); put(ws, r, 6, prop, SIGNED)
        r += 1
    R.OTH_END = r - 1


def build_i_cost_growth(wb):
    """NEW: 10-year adopted budget history for FY28 projection."""
    ws = wb.create_sheet("I-CostGrowth")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [12, 18, 18, 14])

    r = ttl(ws, 1, "INPUT: 10-Year Adopted Budget History")
    r = source_block(ws, r + 1, [
        "SOURCE: FY27 Budget Handbook, p.5 '10 Year Adopted Budget History' (02/11/2026) [Fn 25]",
        "URL: https://resources.finalsite.net/images/v1770852522/rsu5org/ytm49nnwpbrzi4oiwg4i/FY27BUDGETHANDBOOK02112026.pdf",
        "Used to project FY28 cost growth from FY27 adopted baseline.",
    ])
    r += 1

    for i, h in enumerate(["FY", "Adopted Budget", "$ Increase", "% Increase"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    history = [
        ("FY17", 31008672, 1518469, 0.0515),
        ("FY18", 32310685, 1302013, 0.0420),
        ("FY19", 33058025,  747340, 0.0231),
        ("FY20", 34192295, 1134270, 0.0343),
        ("FY21", 34984700,  792405, 0.0232),
        ("FY22", 35714863,  730163, 0.0209),
        ("FY23", 37223151, 1508288, 0.0422),
        ("FY24", 39080569, 1857418, 0.0499),
        ("FY25", 41612460, 2531891, 0.0648),
        ("FY26", 44455929, 2843469, 0.0683),
    ]
    R.CG_START = r
    for fy, adopted, inc, pct in history:
        put(ws, r, 1, fy)
        put(ws, r, 2, adopted, USD)
        put(ws, r, 3, inc, USD)
        put(ws, r, 4, pct, PCT2)
        r += 1
    R.CG_END = r - 1


def build_i_fte(wb):
    """NEW: Teacher FTE by school and grade band for staffing model traceability."""
    ws = wb.create_sheet("I-FTE")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [32, 14, 14, 14, 50])

    r = ttl(ws, 1, "INPUT: Teacher FTE by School & Grade Band")
    r = source_block(ws, r + 1, [
        "SOURCE: FY27 Budget Handbook, p.4 Teacher/Class Size table (02/11/2026) [Fn 1]",
        "NOTE: These are PROPOSED FY27 staffing levels. T1/T2 cuts are already reflected.",
        "FTE figures include classroom teachers only (not aides, specialists, or admin).",
        "Grade-band detail for FMS and DCS drives the middle school consolidation model.",
    ])
    r += 1

    r = sec(ws, r, "School-Level Teacher FTE (Proposed FY27)")
    for i, h in enumerate(["School", "Grades", "Enrollment", "Teacher FTE", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1

    schools_fte = [
        ("PES (Pownal Elementary)",  "PreK-5", 105,  6.5,  "Verified from grade-level data (I-Enrollment 1B)"),
        ("DCS (Durham Community)",   "PreK-8", 467, 26.5,  "Includes PreK through 8; T1: -1 Grade 1 educator"),
        ("MSS (Morse Street)",       "PreK-2", 274, 14.0,  "T1: -1 Grade 2 educator; T2: -1 Ed Tech"),
        ("MLS (Mast Landing)",       "3-5",    264, 14.0,  "T1: -1 Grade 3 educator; T2: -1 Ed Tech"),
        ("FMS (Freeport Middle)",    "6-8",    306, 15.48, "T1: -2.0 FTE (Math, ELA, French); T2: -2 Ed Techs"),
        ("FHS (Freeport High)",      "9-12",   554, 38.0,  "Estimated from Art 1 budget; T1: -2.5 FTE"),
    ]
    R.FT_SCH_START = r
    for name, grades, enroll, fte, notes in schools_fte:
        put(ws, r, 1, name)
        put(ws, r, 2, grades)
        put(ws, r, 3, enroll, USD)
        put(ws, r, 4, fte, '0.00')
        put(ws, r, 5, notes)
        r += 1
    R.FT_SCH_END = r - 1
    R.FT_PES = R.FT_SCH_START
    R.FT_DCS = R.FT_SCH_START + 1
    R.FT_MSS = R.FT_SCH_START + 2
    R.FT_MLS = R.FT_SCH_START + 3
    R.FT_FMS = R.FT_SCH_START + 4
    R.FT_FHS = R.FT_SCH_START + 5
    r += 1

    r = sec(ws, r, "Grade-Band Detail: DCS (for consolidation model)")
    for i, h in enumerate(["Grade Band", "Students", "Teacher FTE", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    R.FT_DCS_PK = r
    put(ws, r, 1, "DCS PreK");      put(ws, r, 2, 48, USD); put(ws, r, 3, 2.5, '0.0')
    put(ws, r, 4, "3 sections, ~16/class"); r += 1
    R.FT_DCS_K5 = r
    put(ws, r, 1, "DCS K-5");       put(ws, r, 2, 278, USD); put(ws, r, 3, 16.0, '0.0')
    put(ws, r, 4, "Post T1 cut: 1 Grade 1 educator removed"); r += 1
    R.FT_DCS_6 = r
    put(ws, r, 1, "DCS Grade 6");   put(ws, r, 2, 41, USD); put(ws, r, 3, 2.0, '0.0')
    put(ws, r, 4, "2 sections, ~20.5/class"); r += 1
    R.FT_DCS_78 = r
    put(ws, r, 1, "DCS Grades 7-8"); put(ws, r, 2, 100, USD); put(ws, r, 3, 6.0, '0.0')
    put(ws, r, 4, "Grades 7 (46) + 8 (54)"); r += 1
    R.FT_DCS_CHK = r
    put(ws, r, 1, "DCS Check Total", fill=CALC_FILL)
    ws.cell(r, 3).value = f"=C{R.FT_DCS_PK}+C{R.FT_DCS_K5}+C{R.FT_DCS_6}+C{R.FT_DCS_78}"
    dat(ws, r, 3, CALC_FILL).number_format = '0.0'
    put(ws, r, 4, "Should equal 26.5", fill=CALC_FILL)
    r += 2

    r = sec(ws, r, "Grade-Band Detail: FMS (for consolidation model)")
    for i, h in enumerate(["Grade Band", "Students", "Teacher FTE", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    R.FT_FMS_6 = r
    put(ws, r, 1, "FMS Grade 6");    put(ws, r, 2, 116, USD); put(ws, r, 3, 5.16, '0.00')
    put(ws, r, 4, "Pownal + Freeport 6th graders"); r += 1
    R.FT_FMS_78 = r
    put(ws, r, 1, "FMS Grades 7-8"); put(ws, r, 2, 190, USD); put(ws, r, 3, 10.32, '0.00')
    put(ws, r, 4, "Post T1: -0.25 Math RTI, -0.75 ELA, -1.0 French"); r += 1
    R.FT_FMS_CHK = r
    put(ws, r, 1, "FMS Check Total", fill=CALC_FILL)
    ws.cell(r, 3).value = f"=C{R.FT_FMS_6}+C{R.FT_FMS_78}"
    dat(ws, r, 3, CALC_FILL).number_format = '0.00'
    put(ws, r, 4, "Should be close to 15.5", fill=CALC_FILL)
    r += 2

    r = sec(ws, r, "Scenario 2 (Path B) FTE Model")
    r = note(ws, r, "Source: District-wide restructuring plan from Superintendent's 02/11/2026 presentation [Fn 10].")
    r = note(ws, r, "Shows how each school's role and FTE would change under full Scenario 2.")
    for i, h in enumerate(["School", "Current FTE", "Scenario 2 FTE", "Change", "Scenario 2 Role"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1

    sc2_data = [
        ("PES", 6.5,  5.0, "EC center (district PreK)"),
        ("DCS", 26.5, 23.0, "K-6 Durham + Pownal"),
        ("MSS", 14.0, 16.0, "K-3 Freeport"),
        ("MLS", 14.0, 15.0, "4-6 Freeport"),
        ("FMS", 15.5, 13.0, "7-8 district-wide"),
    ]
    R.FT_SC2_START = r
    for name, cur, sc2, role in sc2_data:
        put(ws, r, 1, name)
        put(ws, r, 2, cur, '0.0')
        put(ws, r, 3, sc2, '0.0')
        ws.cell(r, 4).value = f"=C{r}-B{r}"
        dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
        put(ws, r, 5, role)
        r += 1
    R.FT_SC2_END = r - 1
    r += 1

    R.FT_SC2_CUR_TOT = r
    put(ws, r, 1, "TOTAL", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.FT_SC2_START}:B{R.FT_SC2_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 3).value = f"=SUM(C{R.FT_SC2_START}:C{R.FT_SC2_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 4).value = f"=C{r}-B{r}"
    dat(ws, r, 4, RESULT_FILL).number_format = '+0.0;-0.0;0.0'
    ws.cell(r, 4).font = RESULT_FONT
    r += 1
    R.FT_SC2_NET = r
    put(ws, r, 1, "Net FTE Saved (Scenario 2)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-D{R.FT_SC2_CUR_TOT}"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 2).font = RESULT_FONT


def build_i_doe_staffing(wb):
    """Maine DOE historical staffing data by school and position category, 2016-2026.

    Data spans 2016-2026. The 2016 data point uses a different staffing
    classification model; baseline comparisons in C-DOEStaffing use 2017.
    """
    ws = wb.create_sheet("I-DOEStaffing")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [32, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10])

    r = ttl(ws, 1, "INPUT: Maine DOE Historical Staffing (FTE)")
    r = source_block(ws, r + 1, [
        "SOURCE: Maine DOE NEO Staff Historical Dec 1 Headcount Data [Fn 45]",
        "Dataset: [Fed_Reporting].Reports.StaffHistoricalDec1, SAUOrgId=1449 (RSU 5)",
        "Requested by Mark Dohle (03/04/2026); fulfilled by Maine DOE Data Team (03/09/2026).",
        "DOE Issue #66636. Raw data file: 'MDohle RSU 5 Staff by FTE.xlsx'",
        "FTE = total sum of time staff is employed, as reported Dec 1 each year.",
        "NOTE: 2016 uses a different staffing classification model that produces",
        "systematically lower FTE counts across all schools. Baseline comparisons",
        "use 2017 as the first year of consistent methodology.",
    ])
    r += 1

    YEARS = list(range(2016, 2027))
    hdrs = ["School"] + [str(y) for y in YEARS]

    # ── Section 1: Total FTE by School ──
    r = sec(ws, r, "1. Total FTE by School (All Position Categories)")
    for i, h in enumerate(hdrs, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(hdrs))
    r += 1

    total_fte = [
        ("PES (Pownal Elementary)",  [26.3, 33.8, 29.6, 28.0, 27.7, 27.3, 24.5, 22.5, 22.9, 21.5, 24.2]),
        ("DCS (Durham Community)",   [71.6, 88.0, 84.8, 92.1, 94.2, 93.0, 87.6, 83.5, 78.4, 88.8, 85.7]),
        ("MSS (Morse Street)",       [42.1, 42.7, 50.5, 57.1, 55.7, 55.5, 60.8, 58.5, 59.8, 64.7, 55.0]),
        ("MLS (Mast Landing)",       [37.0, 48.2, 46.4, 50.2, 55.2, 46.4, 50.5, 49.1, 45.7, 53.8, 50.5]),
        ("FMS (Freeport Middle)",    [63.9, 71.4, 75.7, 74.6, 71.7, 58.9, 67.8, 69.7, 62.8, 64.9, 66.4]),
        ("FHS (Freeport High)",      [102.0, 112.5, 106.0, 124.1, 123.1, 112.4, 100.4, 100.0, 97.4, 108.0, 98.5]),
        ("District-wide",            [103.9, 123.4, 111.4, 125.4, 115.4, 118.9, 119.8, 129.9, 116.2, 144.3, 139.8]),
    ]
    R.DS_TOT_START = r
    for name, vals in total_fte:
        put(ws, r, 1, name)
        for j, v in enumerate(vals, 2):
            put(ws, r, j, v, '0.0')
        r += 1
    R.DS_TOT_END = r - 1
    R.DS_TOT_PES = R.DS_TOT_START
    R.DS_TOT_DCS = R.DS_TOT_START + 1
    R.DS_TOT_MSS = R.DS_TOT_START + 2
    R.DS_TOT_MLS = R.DS_TOT_START + 3
    R.DS_TOT_FMS = R.DS_TOT_START + 4
    R.DS_TOT_FHS = R.DS_TOT_START + 5
    R.DS_TOT_DW  = R.DS_TOT_START + 6

    R.DS_GRAND = r
    put(ws, r, 1, "GRAND TOTAL", fill=CALC_FILL, font=BOLD)
    for c in range(2, 13):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"=SUM({cl}{R.DS_TOT_START}:{cl}{R.DS_TOT_END})"
        dat(ws, r, c, CALC_FILL).number_format = '0.0'
    r += 2

    # ── Section 2: Classroom Teacher FTE ──
    r = sec(ws, r, "2. Classroom Teacher FTE by School")
    r = note(ws, r, "DOE 'Classroom Teacher' includes all certified teaching positions (core + specialists).")
    for i, h in enumerate(hdrs, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(hdrs))
    r += 1

    teacher_fte = [
        ("PES",  [10.4, 10.2, 6.3, 8.3, 8.7, 10.2, 9.2, 10.8, 9.8, 10.2, 9.3]),
        ("DCS",  [26.0, 31.3, 25.5, 30.0, 31.0, 34.0, 33.0, 33.5, 34.5, 34.5, 35.5]),
        ("MSS",  [16.0, 16.2, 16.3, 18.5, 19.8, 20.8, 21.2, 20.6, 21.6, 21.2, 19.2]),
        ("MLS",  [14.6, 17.6, 16.5, 17.7, 16.7, 17.6, 16.6, 17.6, 16.6, 19.6, 19.6]),
        ("FMS",  [25.5, 26.2, 24.5, 24.5, 23.9, 22.1, 23.5, 23.5, 24.0, 23.5, 23.3]),
        ("FHS",  [37.0, 39.1, 34.5, 38.6, 40.7, 42.2, 42.2, 42.7, 42.7, 42.6, 41.7]),
    ]
    R.DS_TCH_START = r
    for name, vals in teacher_fte:
        put(ws, r, 1, name)
        for j, v in enumerate(vals, 2):
            put(ws, r, j, v, '0.0')
        r += 1
    R.DS_TCH_END = r - 1
    R.DS_TCH_PES = R.DS_TCH_START
    R.DS_TCH_DCS = R.DS_TCH_START + 1
    r += 1

    # ── Section 3: Special Education Teacher FTE ──
    r = sec(ws, r, "3. Special Education Teacher FTE by School")
    r = note(ws, r, "Includes Special Education Teacher + Special Education Consultant positions.")
    for i, h in enumerate(hdrs, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(hdrs))
    r += 1

    sped_fte = [
        ("PES",  [1.8, 1.8, 2.0, 0.5, 1.0, 1.0, 1.0, 0.0, 0.0, 1.0, 1.5]),
        ("DCS",  [4.0, 5.0, 5.0, 4.5, 4.0, 4.0, 4.0, 4.0, 4.0, 5.0, 6.1]),
        ("MSS",  [2.0, 2.0, 2.0, 3.0, 3.0, 3.0, 3.0, 4.1, 4.1, 6.0, 4.9]),
        ("MLS",  [2.0, 2.0, 3.0, 2.0, 3.0, 3.0, 3.0, 3.0, 4.0, 4.0, 4.5]),
        ("FMS",  [4.0, 4.0, 4.0, 4.0, 3.0, 4.0, 4.0, 4.0, 4.0, 4.0, 4.5]),
        ("FHS",  [5.0, 6.0, 5.0, 5.0, 5.0, 5.0, 4.0, 4.0, 4.0, 5.0, 4.5]),
    ]
    R.DS_SPED_START = r
    for name, vals in sped_fte:
        put(ws, r, 1, name)
        for j, v in enumerate(vals, 2):
            put(ws, r, j, v, '0.0')
        r += 1
    R.DS_SPED_END = r - 1
    R.DS_SPED_PES = R.DS_SPED_START
    r += 1

    # ── Section 4: Ed Tech FTE (I + II + III + Library/Media) ──
    r = sec(ws, r, "4. Ed Tech FTE (I + II + III + Library/Media) by School")
    r = note(ws, r, "Includes Ed Tech I, Ed Tech II, Ed Tech II - Library/Media, Ed Tech III.")
    for i, h in enumerate(hdrs, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(hdrs))
    r += 1

    edtech_fte = [
        ("PES",  [6.4, 7.4, 5.7, 4.7, 4.2, 4.5, 3.1, 2.1, 5.1, 3.1, 4.0]),
        ("DCS",  [10.0, 11.0, 15.0, 13.0, 14.0, 14.5, 16.0, 14.0, 12.0, 13.0, 13.5]),
        ("MSS",  [8.0, 8.0, 13.4, 12.4, 12.5, 14.0, 15.0, 14.6, 16.6, 15.0, 12.0]),
        ("MLS",  [6.1, 5.6, 4.5, 3.0, 7.0, 6.0, 8.0, 6.0, 7.0, 11.0, 7.0]),
        ("FMS",  [8.5, 10.5, 13.5, 12.5, 11.5, 9.1, 11.2, 16.2, 14.2, 10.2, 11.1]),
        ("FHS",  [10.0, 11.3, 9.0, 8.0, 12.0, 10.0, 10.0, 8.0, 5.5, 9.0, 6.0]),
    ]
    R.DS_ET_START = r
    for name, vals in edtech_fte:
        put(ws, r, 1, name)
        for j, v in enumerate(vals, 2):
            put(ws, r, j, v, '0.0')
        r += 1
    R.DS_ET_END = r - 1
    R.DS_ET_PES = R.DS_ET_START


# ═══════════════════════════════════════════════════════════════
#  CALCULATION SHEETS
# ═══════════════════════════════════════════════════════════════

def build_c_enrollment(wb):
    ws = wb.create_sheet("C-Enrollment")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [35, 16, 16, 16, 16])
    IE = SN['ie']

    r = ttl(ws, 1, "CALC: Enrollment & Town-of-Origin Analysis")
    r = note(ws, r + 1, "All values are formulas referencing I-Enrollment.")
    r += 1

    r = sec(ws, r, "Cohort Averages")
    R.C1_PES_AVG = r
    put(ws, r, 1, "PES avg K-5 grade size", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE({IE}!B{R.PG_K_START}:{IE}!B{R.PG_END})"
    dat(ws, r, 2, CALC_FILL).number_format = '0.0'
    r += 1

    R.C1_DCS_AVG = r
    put(ws, r, 1, "DCS avg 7-8 grade size", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE({IE}!B{R.DCS_G7_R},{IE}!B{R.DCS_G8_R})"
    dat(ws, r, 2, CALC_FILL).number_format = '0.0'
    r += 2

    r = sec(ws, r, "Town-of-Origin at Shared Schools")
    for i, h in enumerate(["School", "Pownal (est.)", "Durham (est.)", "Freeport (est.)"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    R.C1_FMS_R = r
    put(ws, r, 1, "At FMS", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=ROUND(B{R.C1_PES_AVG},0)*3"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = 0; dat(ws, r, 3, CALC_FILL)
    ws.cell(r, 4).value = f"={IE}!F{R.FMS_R}-B{r}"
    dat(ws, r, 4, CALC_FILL).number_format = USD
    r += 1

    R.C1_FHS_R = r
    put(ws, r, 1, "At FHS", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=ROUND(B{R.C1_PES_AVG},0)*4"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=ROUND(B{R.C1_DCS_AVG},0)*4"
    dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"={IE}!F{R.FHS_R}-B{r}-C{r}"
    dat(ws, r, 4, CALC_FILL).number_format = USD
    r += 2

    r = sec(ws, r, "Total Students by Town")
    for i, h in enumerate(["Town", "Own Schools", "At FMS", "At FHS", "TOTAL"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1

    R.C1_POW = r
    put(ws, r, 1, "Pownal", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IE}!F{R.PES_R}"; dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=B{R.C1_FMS_R}"; dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"=B{R.C1_FHS_R}"; dat(ws, r, 4, CALC_FILL).number_format = USD
    ws.cell(r, 5).value = f"=SUM(B{r}:D{r})"; dat(ws, r, 5, RESULT_FILL).number_format = USD
    r += 1

    R.C1_DUR = r
    put(ws, r, 1, "Durham", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IE}!F{R.DCS_R}"; dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = 0; dat(ws, r, 3, CALC_FILL)
    ws.cell(r, 4).value = f"=C{R.C1_FHS_R}"; dat(ws, r, 4, CALC_FILL).number_format = USD
    ws.cell(r, 5).value = f"=SUM(B{r}:D{r})"; dat(ws, r, 5, RESULT_FILL).number_format = USD
    r += 1

    R.C1_FRE = r
    put(ws, r, 1, "Freeport", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IE}!F{R.MSS_R}+{IE}!F{R.MLS_R}"; dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=D{R.C1_FMS_R}"; dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"=D{R.C1_FHS_R}"; dat(ws, r, 4, CALC_FILL).number_format = USD
    ws.cell(r, 5).value = f"=SUM(B{r}:D{r})"; dat(ws, r, 5, RESULT_FILL).number_format = USD
    r += 1

    R.C1_DIST = r
    put(ws, r, 1, "DISTRICT TOTAL", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 5).value = f"=E{R.C1_POW}+E{R.C1_DUR}+E{R.C1_FRE}"
    dat(ws, r, 5, RESULT_FILL).number_format = USD
    ws.cell(r, 5).font = BOLD


def build_c_budget(wb):
    ws = wb.create_sheet("C-Budget")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [35, 16, 16, 16, 16, 16, 16, 16])
    IB = SN['ib']
    IE = SN['ie']

    r = ttl(ws, 1, "CALC: Budget Summary & School Direct Costs")
    r = note(ws, r + 1, "All values are SUM formulas referencing I-Budget input sheet.")
    r += 1

    r = sec(ws, r, "School Direct Costs (sum of all articles per school)")
    school_names = ["DCS", "MSS", "PES", "MLS", "FMS", "FHS"]
    for i, h in enumerate(["School"] + school_names, 1):
        ws.cell(r, i, h)
    hdr(ws, r, 7)
    r += 1

    R.C2_DIRECT = r
    for j, sch in enumerate(school_names, 2):
        col_l = get_column_letter(j)
        ws.cell(r, j).value = f"=SUM({IB}!{col_l}{R.ART_START}:{col_l}{R.ART_END})"
        dat(ws, r, j, CALC_FILL).number_format = USD
    put(ws, r, 1, "Direct Cost", fill=CALC_FILL, font=BOLD)
    r += 1

    R.C2_PERSTUD = r
    enroll_rows = [R.DCS_R, R.MSS_R, R.PES_R, R.MLS_R, R.FMS_R, R.FHS_R]
    for j, er in enumerate(enroll_rows, 2):
        col_l = get_column_letter(j)
        ws.cell(r, j).value = f"={col_l}{R.C2_DIRECT}/{IE}!F{er}"
        dat(ws, r, j, CALC_FILL).number_format = USD
    put(ws, r, 1, "Per Student", fill=CALC_FILL)
    r += 2

    r = sec(ws, r, "System-Wide Pool")
    R.C2_SYS_ARTS = r
    put(ws, r, 1, "System-only articles total", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM({IB}!B{R.SYS_START}:{IB}!B{R.SYS_END})"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    r += 1

    R.C2_SYS_ALLOC = r
    put(ws, r, 1, "SYS/K8/912/CTRL from school articles", fill=CALC_FILL)
    ws.cell(r, 2).value = (
        f"=SUM({IB}!H{R.ART_START}:{IB}!H{R.ART_END})"
        f"+SUM({IB}!I{R.ART_START}:{IB}!I{R.ART_END})"
        f"+SUM({IB}!J{R.ART_START}:{IB}!J{R.ART_END})"
        f"+SUM({IB}!K{R.ART_START}:{IB}!K{R.ART_END})"
    )
    dat(ws, r, 2, CALC_FILL).number_format = USD
    r += 1

    R.C2_SYSPOOL = r
    put(ws, r, 1, "TOTAL SYSTEM-WIDE POOL", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C2_SYS_ARTS}+B{R.C2_SYS_ALLOC}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = BOLD
    r += 2

    R.C2_GRAND = r
    put(ws, r, 1, "GRAND TOTAL (Articles 1-11)", fill=RESULT_FILL, font=BOLD)
    direct_parts = "+".join(f"{get_column_letter(j)}{R.C2_DIRECT}" for j in range(2, 8))
    ws.cell(r, 2).value = f"={direct_parts}+B{R.C2_SYSPOOL}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = BOLD
    r += 1

    R.C2_FY27 = r
    put(ws, r, 1, "FY27 TOTAL OPERATING (+ Adult Ed)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C2_GRAND}+{IB}!B{R.ADULT_ED_R}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = BOLD
    r += 1

    R.C2_INCR = r
    put(ws, r, 1, "FY27 Increase ($) from FY26", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C2_FY27}-{IB}!B{R.FY26_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    r += 1

    R.C2_INCR_PCT = r
    put(ws, r, 1, "FY27 Increase (%)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C2_INCR}/{IB}!B{R.FY26_R}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2


def build_c_consumption(wb):
    ws = wb.create_sheet("C-Consumption")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [35, 18, 14, 18])
    CB = SN['cb']; CE = SN['ce']; IE = SN['ie']

    r = ttl(ws, 1, "CALC: Budget Consumption by Town of Origin")
    r = note(ws, r + 1, "Direct costs -> 100% to town. FMS/FHS -> proportional. System -> by student count.")
    r += 1

    def sch_direct(col_idx):
        return f"{CB}!{get_column_letter(col_idx)}{R.C2_DIRECT}"

    DCS_D, MSS_D, PES_D, MLS_D, FMS_D, FHS_D = (sch_direct(j) for j in range(2, 8))
    SYS_D = f"{CB}!B{R.C2_SYSPOOL}"
    POW_T = f"{CE}!E{R.C1_POW}"
    DUR_T = f"{CE}!E{R.C1_DUR}"
    FRE_T = f"{CE}!E{R.C1_FRE}"
    DIST_T = f"{CE}!E{R.C1_DIST}"
    FMS_ENR = f"{IE}!F{R.FMS_R}"
    FHS_ENR = f"{IE}!F{R.FHS_R}"

    def cons_block(ws, r, town_label, components):
        r = sec(ws, r, f"{town_label} Consumption")
        for i, h in enumerate(["Component", "School Total", "Share %", "Town's Share"], 1):
            ws.cell(r, i, h)
        hdr(ws, r, 4)
        r += 1
        start = r
        for label, total_ref, share_formula in components:
            put(ws, r, 1, label, fill=CALC_FILL)
            ws.cell(r, 2).value = f"={total_ref}"; dat(ws, r, 2, CALC_FILL).number_format = USD
            ws.cell(r, 3).value = f"={share_formula}"; dat(ws, r, 3, CALC_FILL).number_format = PCT
            ws.cell(r, 4).value = f"=B{r}*C{r}"; dat(ws, r, 4, CALC_FILL).number_format = USD
            r += 1
        end = r - 1
        put(ws, r, 1, f"{town_label} TOTAL", fill=RESULT_FILL, font=BOLD)
        ws.cell(r, 4).value = f"=SUM(D{start}:D{end})"
        dat(ws, r, 4, RESULT_FILL).number_format = USD
        ws.cell(r, 4).font = BOLD
        total_row = r
        r += 1
        put(ws, r, 1, f"{town_label} Per Student", fill=CALC_FILL)
        r += 1
        return r + 1, total_row

    r, R.C3_POW_TOTAL = cons_block(ws, r, "Pownal", [
        ("PES (100%)", PES_D, "1"),
        ("FMS (Pownal share)", FMS_D, f"{CE}!B{R.C1_FMS_R}/{FMS_ENR}"),
        ("FHS (Pownal share)", FHS_D, f"{CE}!B{R.C1_FHS_R}/{FHS_ENR}"),
        ("System (Pownal share)", SYS_D, f"{POW_T}/{DIST_T}"),
    ])
    ws.cell(R.C3_POW_TOTAL + 1, 4).value = f"=D{R.C3_POW_TOTAL}/{POW_T}"
    dat(ws, R.C3_POW_TOTAL + 1, 4, CALC_FILL).number_format = USD

    r, R.C3_DUR_TOTAL = cons_block(ws, r, "Durham", [
        ("DCS (100%)", DCS_D, "1"),
        ("FHS (Durham share)", FHS_D, f"{CE}!C{R.C1_FHS_R}/{FHS_ENR}"),
        ("System (Durham share)", SYS_D, f"{DUR_T}/{DIST_T}"),
    ])
    ws.cell(R.C3_DUR_TOTAL + 1, 4).value = f"=D{R.C3_DUR_TOTAL}/{DUR_T}"
    dat(ws, R.C3_DUR_TOTAL + 1, 4, CALC_FILL).number_format = USD

    r, R.C3_FRE_TOTAL = cons_block(ws, r, "Freeport", [
        ("MSS (100%)", MSS_D, "1"),
        ("MLS (100%)", MLS_D, "1"),
        ("FMS (Freeport share)", FMS_D, f"{CE}!D{R.C1_FMS_R}/{FMS_ENR}"),
        ("FHS (Freeport share)", FHS_D, f"{CE}!D{R.C1_FHS_R}/{FHS_ENR}"),
        ("System (Freeport share)", SYS_D, f"{FRE_T}/{DIST_T}"),
    ])
    ws.cell(R.C3_FRE_TOTAL + 1, 4).value = f"=D{R.C3_FRE_TOTAL}/{FRE_T}"
    dat(ws, R.C3_FRE_TOTAL + 1, 4, CALC_FILL).number_format = USD

    r += 1
    r = sec(ws, r, "Integrity Check")
    R.C3_CHECK = r
    put(ws, r, 1, "Sum of Town Consumption", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 4).value = f"=D{R.C3_POW_TOTAL}+D{R.C3_DUR_TOTAL}+D{R.C3_FRE_TOTAL}"
    dat(ws, r, 4, RESULT_FILL).number_format = USD
    ws.cell(r, 4).font = BOLD
    r += 1
    put(ws, r, 1, "Budget Total (should match)", fill=RESULT_FILL)
    ws.cell(r, 4).value = f"={CB}!B{R.C2_GRAND}"; dat(ws, r, 4, RESULT_FILL).number_format = USD
    r += 1
    put(ws, r, 1, "Variance", fill=RESULT_FILL)
    ws.cell(r, 4).value = f"=D{R.C3_CHECK}-D{R.C3_CHECK+1}"; dat(ws, r, 4, RESULT_FILL).number_format = USD


def build_c_revenue(wb):
    ws = wb.create_sheet("C-Revenue")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [35, 18, 18, 18])
    IR = SN['ir']; CE = SN['ce']; CC = SN['cc']

    r = ttl(ws, 1, "CALC: Revenue & Net Fiscal Position")
    r = note(ws, r + 1, "All values reference I-Revenue and C-Enrollment.")
    r += 1

    r = sec(ws, r, "ALM Percentage (computed)")
    alm_sum = f"({IR}!B{R.ALM_R}+{IR}!C{R.ALM_R}+{IR}!D{R.ALM_R})"
    R.C4_ALM = r
    put(ws, r, 1, "Pownal ALM %", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!B{R.ALM_R}/{alm_sum}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 1
    R.C4_ALM_D = r
    put(ws, r, 1, "Durham ALM %", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!C{R.ALM_R}/{alm_sum}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 1
    R.C4_ALM_F = r
    put(ws, r, 1, "Freeport ALM %", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!D{R.ALM_R}/{alm_sum}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 2

    r = sec(ws, r, "Total Revenue by Town")
    for i, h in enumerate(["Component", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.C4_CONTRIB = r
    put(ws, r, 1, "Total Contribution", fill=CALC_FILL)
    for c, tc in [(2, "B"), (3, "C"), (4, "D")]:
        ws.cell(r, c).value = f"=SUM({IR}!{tc}{R.REV_START}:{IR}!{tc}{R.REV_END})"
        dat(ws, r, c, CALC_FILL).number_format = USD
    r += 1

    R.C4_SHARED = r
    put(ws, r, 1, "Shared Revenue", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!B{R.SHARED_R}*B{R.C4_ALM}"; dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"={IR}!B{R.SHARED_R}*B{R.C4_ALM_D}"; dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"={IR}!B{R.SHARED_R}*B{R.C4_ALM_F}"; dat(ws, r, 4, CALC_FILL).number_format = USD
    r += 1

    R.C4_TOTAL_REV = r
    put(ws, r, 1, "TOTAL REVENUE", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.C4_CONTRIB}+{cl}{R.C4_SHARED}"
        dat(ws, r, c, RESULT_FILL).number_format = USD
        ws.cell(r, c).font = BOLD
    r += 2

    r = sec(ws, r, "Net Fiscal Position")
    POW_T = f"{CE}!E{R.C1_POW}"; DUR_T = f"{CE}!E{R.C1_DUR}"; FRE_T = f"{CE}!E{R.C1_FRE}"
    for i, h in enumerate(["Town", "Revenue", "Consumption", "Net Position", "Students", "Rev/Stu", "Cost/Stu"], 1):
        ws.cell(r, i, h)
        if i > 4:
            ws.column_dimensions[get_column_letter(i)].width = 14
    hdr(ws, r, 7); r += 1

    for town, rev_f, cons_f, stu_f in [
        ("Pownal",   f"=B{R.C4_TOTAL_REV}", f"={CC}!D{R.C3_POW_TOTAL}", POW_T),
        ("Durham",   f"=C{R.C4_TOTAL_REV}", f"={CC}!D{R.C3_DUR_TOTAL}", DUR_T),
        ("Freeport", f"=D{R.C4_TOTAL_REV}", f"={CC}!D{R.C3_FRE_TOTAL}", FRE_T),
    ]:
        put(ws, r, 1, town, fill=CALC_FILL)
        ws.cell(r, 2).value = rev_f; dat(ws, r, 2, CALC_FILL).number_format = USD
        ws.cell(r, 3).value = cons_f; dat(ws, r, 3, CALC_FILL).number_format = USD
        ws.cell(r, 4).value = f"=B{r}-C{r}"; dat(ws, r, 4, RESULT_FILL).number_format = SIGNED
        ws.cell(r, 4).font = RESULT_FONT
        ws.cell(r, 5).value = f"={stu_f}"; dat(ws, r, 5, CALC_FILL).number_format = USD
        ws.cell(r, 6).value = f"=B{r}/E{r}"; dat(ws, r, 6, CALC_FILL).number_format = USD
        ws.cell(r, 7).value = f"=C{r}/E{r}"; dat(ws, r, 7, CALC_FILL).number_format = USD
        r += 1


def build_c_tax(wb):
    ws = wb.create_sheet("C-Tax")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [35, 18, 18, 18, 18, 18])
    IR = SN['ir']; IT = SN['it']; CE = SN['ce']

    r = ttl(ws, 1, "CALC: Tax Analysis")
    r = note(ws, r + 1, "All values reference I-Tax and I-Revenue input sheets.")
    r += 1

    r = sec(ws, r, "RSU Net Tax Impact (RLC + ALM + Non-Shared Debt)")
    R.C5_POW_TAX = r
    put(ws, r, 1, "Pownal", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!B{R.RLC_R}+{IR}!B{R.ALM_R}+{IR}!B{R.DEBT_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1
    R.C5_DUR_TAX = r
    put(ws, r, 1, "Durham", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!C{R.RLC_R}+{IR}!C{R.ALM_R}+{IR}!C{R.DEBT_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1
    R.C5_FRE_TAX = r
    put(ws, r, 1, "Freeport", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IR}!D{R.RLC_R}+{IR}!D{R.ALM_R}+{IR}!D{R.DEBT_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 2

    r = sec(ws, r, "Equalized School Mil Rates")
    for i, h in enumerate(["Town", "RSU Net Tax", "State Val", "Eq School Mil", "Local Mil", "Assessment Ratio"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 6); r += 1
    for town, tax_row, sv_r, tv_r in [
        ("Pownal",   R.C5_POW_TAX, R.POW_STATE_R, R.POW_TAXABLE_R),
        ("Durham",   R.C5_DUR_TAX, R.DUR_STATE_R, R.DUR_TAXABLE_R),
        ("Freeport", R.C5_FRE_TAX, R.FRE_STATE_R, R.FRE_TAXABLE_R),
    ]:
        put(ws, r, 1, town, fill=CALC_FILL)
        ws.cell(r, 2).value = f"=B{tax_row}"; dat(ws, r, 2, CALC_FILL).number_format = USD
        ws.cell(r, 3).value = f"={IT}!B{sv_r}"; dat(ws, r, 3, CALC_FILL).number_format = USD
        ws.cell(r, 4).value = f"=B{r}/C{r}*1000"; dat(ws, r, 4, CALC_FILL).number_format = '0.00'
        ws.cell(r, 5).value = f"=B{r}/{IT}!B{tv_r}*1000"; dat(ws, r, 5, CALC_FILL).number_format = '0.00'
        ws.cell(r, 6).value = f"={IT}!B{tv_r}/{IT}!B{sv_r}"; dat(ws, r, 6, CALC_FILL).number_format = PCT
        r += 1

    r += 1
    r = sec(ws, r, "Per-Student Tax Burden")
    for i, h in enumerate(["Town", "RSU Net Tax", "Students", "Tax/Student"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1
    POW_T = f"{CE}!E{R.C1_POW}"; DUR_T = f"{CE}!E{R.C1_DUR}"; FRE_T = f"{CE}!E{R.C1_FRE}"
    for town, tax_row, stu_ref in [
        ("Pownal", R.C5_POW_TAX, POW_T), ("Durham", R.C5_DUR_TAX, DUR_T), ("Freeport", R.C5_FRE_TAX, FRE_T),
    ]:
        put(ws, r, 1, town, fill=CALC_FILL)
        ws.cell(r, 2).value = f"=B{tax_row}"; dat(ws, r, 2, CALC_FILL).number_format = USD
        ws.cell(r, 3).value = f"={stu_ref}"; dat(ws, r, 3, CALC_FILL).number_format = USD
        ws.cell(r, 4).value = f"=B{r}/C{r}"; dat(ws, r, 4, RESULT_FILL).number_format = USD
        ws.cell(r, 4).font = RESULT_FONT
        r += 1


def build_c_cost_premium(wb):
    ws = wb.create_sheet("C-CostPremium")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [35, 14, 14, 14, 14, 14, 16])
    IB = SN['ib']; IE = SN['ie']

    r = ttl(ws, 1, "CALC: PES Per-Student Cost Premium")
    r = note(ws, r + 1, "Compares PES per-student cost vs average of DCS, MSS, MLS from I-Budget.")
    r += 1

    for i, h in enumerate(["Article", "PES $/Stu", "DCS $/Stu", "MSS $/Stu", "MLS $/Stu", "Other Avg", "PES Premium"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 7); r += 1

    art_labels = ["Art 1 - Regular Instr.", "Art 2 - Special Ed", "Art 4 - Other Instr.",
                  "Art 5 - Student Support", "Art 7 - Administration", "Art 9 - Facilities"]
    R.CP_START = r
    for idx, art_name in enumerate(art_labels):
        art_r = R.ART_START + idx
        put(ws, r, 1, art_name, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={IB}!D{art_r}/{IE}!F{R.PES_R}"; dat(ws, r, 2, CALC_FILL).number_format = USD
        ws.cell(r, 3).value = f"={IB}!B{art_r}/{IE}!F{R.DCS_R}"; dat(ws, r, 3, CALC_FILL).number_format = USD
        ws.cell(r, 4).value = f"={IB}!C{art_r}/{IE}!F{R.MSS_R}"; dat(ws, r, 4, CALC_FILL).number_format = USD
        ws.cell(r, 5).value = f"={IB}!E{art_r}/{IE}!F{R.MLS_R}"; dat(ws, r, 5, CALC_FILL).number_format = USD
        ws.cell(r, 6).value = f"=AVERAGE(C{r}:E{r})"; dat(ws, r, 6, CALC_FILL).number_format = USD
        ws.cell(r, 7).value = f"=B{r}-F{r}"; dat(ws, r, 7, CALC_FILL).number_format = SIGNED
        r += 1
    R.CP_END = r - 1

    put(ws, r, 1, "TOTAL", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.CP_START}:B{R.CP_END})"; dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 6).value = f"=SUM(F{R.CP_START}:F{R.CP_END})"; dat(ws, r, 6, RESULT_FILL).number_format = USD
    ws.cell(r, 7).value = f"=B{r}-F{r}"; dat(ws, r, 7, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 7).font = RESULT_FONT
    R.CP_TOTAL = r
    r += 2
    put(ws, r, 1, "Total Gap (premium x PES enrollment)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=G{R.CP_TOTAL}*{IE}!F{R.PES_R}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT


def build_c_ms_consolidation(wb):
    ws = wb.create_sheet("C-MSConsol")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [42, 16, 16, 16])
    CE = SN['ce']; IE = SN['ie']; IA = SN['ia']; IFT = SN['ift']

    r = ttl(ws, 1, "CALC: Middle School Consolidation")
    r = note(ws, r + 1, "References I-FTE for current staffing, I-Enrollment for student counts, I-Assumptions for targets.")
    r += 1

    r = sec(ws, r, "Student Movement")
    R.C7_POW6 = r
    put(ws, r, 1, "Pownal 6th graders (est.)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=ROUND({CE}!B{R.C1_PES_AVG},0)"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C7_FRE6 = r
    put(ws, r, 1, "Freeport 6th graders (est.)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IE}!B{R.FMS_G6_R}-B{R.C7_POW6}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C7_NEW_FMS = r
    put(ws, r, 1, "New FMS total (district 7-8)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IE}!B{R.FMS_G7_R}+{IE}!B{R.FMS_G8_R}+{IE}!B{R.DCS_G7_R}+{IE}!B{R.DCS_G8_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C7_NEW_DCS6 = r
    put(ws, r, 1, "New DCS 6th total (district-wide)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IE}!B{R.DCS_G6_R}+{IE}!B{R.FMS_G6_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 2

    r = sec(ws, r, "Staffing Impact")
    for i, h in enumerate(["Grade Level", "Current FTE", "Proposed FTE", "Change"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.C7_78 = r
    put(ws, r, 1, "Grades 7-8", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IFT}!C{R.FT_DCS_78}+{IFT}!C{R.FT_FMS_78}"
    dat(ws, r, 2, CALC_FILL).number_format = '0.0'
    ws.cell(r, 3).value = f"=B{R.C7_NEW_FMS}/{IA}!B{R.A_CS_78}"
    dat(ws, r, 3, CALC_FILL).number_format = '0.0'
    ws.cell(r, 4).value = f"=C{r}-B{r}"; dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
    r += 1

    R.C7_6 = r
    put(ws, r, 1, "Grade 6", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IFT}!C{R.FT_DCS_6}+{IFT}!C{R.FT_FMS_6}"
    dat(ws, r, 2, CALC_FILL).number_format = '0.0'
    ws.cell(r, 3).value = f"=B{R.C7_NEW_DCS6}/{IA}!B{R.A_CS_6}"
    dat(ws, r, 3, CALC_FILL).number_format = '0.0'
    ws.cell(r, 4).value = f"=C{r}-B{r}"; dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
    r += 1

    R.C7_NET = r
    put(ws, r, 1, "NET", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C7_78}+B{R.C7_6}"; dat(ws, r, 2, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 3).value = f"=C{R.C7_78}+C{R.C7_6}"; dat(ws, r, 3, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 4).value = f"=C{r}-B{r}"; dat(ws, r, 4, RESULT_FILL).number_format = '+0.0;-0.0;0.0'
    r += 2

    r = sec(ws, r, "Financial Summary")
    R.C7_SAVED = r
    put(ws, r, 1, "Net FTE Saved", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-D{R.C7_NET}"; dat(ws, r, 2, CALC_FILL).number_format = '0.0'; r += 1

    R.C7_TEACH_SAV = r
    put(ws, r, 1, "Teacher Savings", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C7_SAVED}*{IA}!B{R.A_COST_FTE}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C7_EXTRA = r
    put(ws, r, 1, "Extracurricular Savings", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IA}!B{R.A_EXTRACURR}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C7_TOTAL = r
    put(ws, r, 1, "TOTAL MS CONSOLIDATION SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C7_TEACH_SAV}+B{R.C7_EXTRA}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT


def build_c_ec_analysis(wb):
    ws = wb.create_sheet("C-ECAnalysis")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [42, 18, 18, 18])
    IC = SN['ic']

    r = ttl(ws, 1, "CALC: EC Net Cost Analysis")
    r = note(ws, r + 1, "References I-ECCosts for staff costs and offsets.")
    r = note(ws, r, "NOTE: EC costs are FY28 expenses (begin July 2027), not FY27.")
    r += 1

    R.C8_GROSS = r
    put(ws, r, 1, "Option 1 Gross Cost", fill=CALC_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM({IC}!D{R.EC_START}:{IC}!D{R.EC_END})+{IC}!D{R.EC_EQUIP_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C8_OFF_LOW = r
    put(ws, r, 1, "Offsets (high -> low net cost)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM({IC}!C{R.OFF_START}:{IC}!C{R.OFF_END})"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C8_OFF_HIGH = r
    put(ws, r, 1, "Offsets (low -> high net cost)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM({IC}!B{R.OFF_START}:{IC}!B{R.OFF_END})"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C8_OFF_MID = r
    put(ws, r, 1, "Offsets (midpoint)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE(B{R.C8_OFF_LOW},B{R.C8_OFF_HIGH})"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 2

    put(ws, r, 1, "", fill=CALC_FILL)
    ws.cell(r, 2, "Low"); ws.cell(r, 3, "High"); ws.cell(r, 4, "Midpoint")
    hdr(ws, r, 4); r += 1

    R.C8_NET = r
    put(ws, r, 1, "NET EC COST (annual, FY28+)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C8_GROSS}-B{R.C8_OFF_LOW}"; dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"=B{R.C8_GROSS}-B{R.C8_OFF_HIGH}"; dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 4).value = f"=B{R.C8_GROSS}-B{R.C8_OFF_MID}"; dat(ws, r, 4, RESULT_FILL).number_format = USD
    ws.cell(r, 4).font = RESULT_FONT


def build_c_fy27_scenarios(wb):
    """NEW: FY27 adoption scenarios -- what goes to voters this spring."""
    ws = wb.create_sheet("C-FY27Scenarios")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [48, 20, 20, 20])
    IB = SN['ib']; IRD = SN['ird']; CB = SN['cb']

    r = ttl(ws, 1, "CALC: FY27 Budget Adoption Scenarios")
    r = note(ws, r + 1, "The FY27 budget goes to voters May/June 2026. Board adopts March 25.")
    r = note(ws, r, "EC mandate costs and structural changes are NOT in FY27. They are FY28 decisions.")
    r += 1

    r = sec(ws, r, "FY27 Starting Point")
    R.C27_FY26 = r
    put(ws, r, 1, "FY26 Adopted Budget", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IB}!B{R.FY26_R}"; dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C27_INIT = r
    put(ws, r, 1, "Initial request increase (pre-cut)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IRD}!B{R.RD_INIT_INCR}"; dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C27_MAINT = r
    put(ws, r, 1, "\"Maintenance of effort\" total (before cuts)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C27_FY26}+B{R.C27_INIT}"; dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 1

    R.C27_MAINT_PCT = r
    put(ws, r, 1, "Maint. of effort % increase", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C27_INIT}/B{R.C27_FY26}"; dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 2

    r = sec(ws, r, "Reductions Already in Proposed Budget")
    R.C27_T1 = r
    put(ws, r, 1, "Tier 1 reductions (staff positions)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM({IRD}!F{R.T1_START}:{IRD}!F{R.T1_END})"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.C27_T2 = r
    put(ws, r, 1, "Tier 2 reductions (support positions)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM({IRD}!F{R.T2_START}:{IRD}!F{R.T2_END})"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.C27_OTH = r
    put(ws, r, 1, "Other reductions (tech, facilities)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM({IRD}!E{R.OTH_START}:{IRD}!E{R.OTH_END})"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.C27_TOTAL_RED = r
    put(ws, r, 1, "TOTAL REDUCTIONS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C27_T1}+B{R.C27_T2}+B{R.C27_OTH}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # Three scenarios
    r = sec(ws, r, "Three Adoption Scenarios")
    for i, h in enumerate(["Metric", "Conservative", "Proposed", "Higher"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.C27_SC_BUD = r
    put(ws, r, 1, "FY27 Budget Total", fill=RESULT_FILL, font=BOLD)
    # Conservative: Board targets ~5% increase
    ws.cell(r, 2).value = f"=B{R.C27_FY26}*1.05"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    # Proposed: as-is from C-Budget
    ws.cell(r, 3).value = f"={CB}!B{R.C2_FY27}"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    # Higher: restore T2 positions
    ws.cell(r, 4).value = f"={CB}!B{R.C2_FY27}-B{R.C27_T2}"
    dat(ws, r, 4, RESULT_FILL).number_format = USD
    r += 1

    R.C27_SC_PCT = r
    put(ws, r, 1, "% Increase from FY26", fill=RESULT_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"=({cl}{R.C27_SC_BUD}-B{R.C27_FY26})/B{R.C27_FY26}"
        dat(ws, r, c, RESULT_FILL).number_format = PCT2
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    R.C27_SC_ADDL = r
    put(ws, r, 1, "Additional cuts beyond proposed", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C27_SC_BUD}-C{R.C27_SC_BUD}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = 0; dat(ws, r, 3, CALC_FILL)
    ws.cell(r, 4).value = f"=D{R.C27_SC_BUD}-C{R.C27_SC_BUD}"
    dat(ws, r, 4, CALC_FILL).number_format = SIGNED
    r += 1

    put(ws, r, 1, "Positions cut (T1)", fill=CALC_FILL)
    ws.cell(r, 2).value = "13+ (T1+T2+more)"; dat(ws, r, 2, CALC_FILL)
    ws.cell(r, 3).value = "13 (T1 only)"; dat(ws, r, 3, CALC_FILL)
    ws.cell(r, 4).value = "13 (T1; T2 restored)"; dat(ws, r, 4, CALC_FILL)
    r += 1
    put(ws, r, 1, "T2 support positions", fill=CALC_FILL)
    ws.cell(r, 2).value = "Cut"; dat(ws, r, 2, CALC_FILL)
    ws.cell(r, 3).value = "Cut"; dat(ws, r, 3, CALC_FILL)
    ws.cell(r, 4).value = "Restored"; dat(ws, r, 4, CALC_FILL)
    r += 2

    r = note(ws, r, "The 'Proposed' column matches the Superintendent's recommendation ($47,357,441).")
    r = note(ws, r, "The 'Conservative' scenario targets 5.0% increase. Board may choose differently.")
    r = note(ws, r, "The 'Higher' scenario restores T2 support positions (ed techs, instructional strategist).")


def build_c_fy28_projection(wb):
    """NEW: FY28 projection -- the multi-year chain from FY27 to FY28."""
    ws = wb.create_sheet("C-FY28Projection")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [48, 20, 20, 20, 20])
    C27 = SN['c27']; ICG = SN['icg']; CEC = SN['cec']; CFP = SN['cfp']

    r = ttl(ws, 1, "CALC: FY28 Budget Projection")
    r = note(ws, r + 1, "Projects from FY27 adopted -> FY28 baseline -> FY28 with structural changes.")
    r = note(ws, r, "FY28 is when EC mandate and any restructuring take effect (2027-2028 school year).")
    r += 1

    # FY27 Adopted (user-selectable)
    r = sec(ws, r, "Step 1: FY27 Adopted Budget (select scenario)")
    R.C28_FY27_SEL = r
    put(ws, r, 1, "SELECTED FY27 ADOPTED >>", fill=PARAM_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={C27}!C{R.C27_SC_BUD}"  # default to Proposed
    dat(ws, r, 2, PARAM_FILL).number_format = USD
    ws.cell(r, 2).font = BOLD
    r += 1
    r = note(ws, r, "Change cell B above to test different FY27 outcomes. Default = Proposed.")
    r = note(ws, r, f"Conservative: ={C27}!B{R.C27_SC_BUD}  |  Higher: ={C27}!D{R.C27_SC_BUD}")
    r += 1

    # Growth rates from history
    r = sec(ws, r, "Step 2: Cost Growth Rate (from I-CostGrowth)")
    R.C28_GR_3 = r
    put(ws, r, 1, "3-year average growth (FY24-FY26)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE({ICG}!D{R.CG_END-2}:{ICG}!D{R.CG_END})"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 1

    R.C28_GR_5 = r
    put(ws, r, 1, "5-year average growth (FY22-FY26)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE({ICG}!D{R.CG_END-4}:{ICG}!D{R.CG_END})"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 1

    R.C28_GR_10 = r
    put(ws, r, 1, "10-year average growth (FY17-FY26)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE({ICG}!D{R.CG_START}:{ICG}!D{R.CG_END})"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 1

    R.C28_GR_SEL = r
    put(ws, r, 1, "SELECTED GROWTH RATE >>", fill=PARAM_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C28_GR_3}"  # default to 3-year
    dat(ws, r, 2, PARAM_FILL).number_format = PCT2
    ws.cell(r, 2).font = BOLD
    r += 1
    r = note(ws, r, "Change cell B above to test different growth assumptions. Default = 3-year avg.")
    r += 1

    # EC net cost
    r = sec(ws, r, "Step 3: EC Mandate (FY28 new cost)")
    R.C28_EC = r
    put(ws, r, 1, "EC net cost (midpoint)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CEC}!D{R.C8_NET}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 2

    # FY28 baseline
    r = sec(ws, r, "Step 4: FY28 Baseline (no structural changes)")
    R.C28_OPER = r
    put(ws, r, 1, "FY28 operating (FY27 + cost growth)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.C28_FY27_SEL}*(1+B{R.C28_GR_SEL})"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C28_BASE = r
    put(ws, r, 1, "FY28 BASELINE (operating + EC)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C28_OPER}+B{R.C28_EC}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.C28_BASE_PCT = r
    put(ws, r, 1, "FY28 increase from FY27", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=(B{R.C28_BASE}-B{R.C28_FY27_SEL})/B{R.C28_FY27_SEL}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 2

    # Sensitivity matrix
    r = sec(ws, r, "Sensitivity: FY28 Baseline by FY27 Outcome x Growth Rate")
    for i, h in enumerate(["FY27 Scenario", "3-yr Growth", "5-yr Growth", "10-yr Growth"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.C28_MATRIX = r
    for label, fy27_ref in [
        ("Conservative", f"={C27}!B{R.C27_SC_BUD}"),
        ("Proposed",     f"={C27}!C{R.C27_SC_BUD}"),
        ("Higher",       f"={C27}!D{R.C27_SC_BUD}"),
    ]:
        put(ws, r, 1, label, fill=CALC_FILL)
        for c, gr_row in [(2, R.C28_GR_3), (3, R.C28_GR_5), (4, R.C28_GR_10)]:
            # FY28 = FY27_scenario * (1 + growth) + EC
            # Need to inline the FY27 ref since it's not a single cell for all rows
            # Use a helper column approach: put FY27 value in col 5 (hidden)
            pass
        r += 1
    # Rewrite the matrix with a cleaner approach
    r = R.C28_MATRIX
    fy27_refs = [
        f"{C27}!B{R.C27_SC_BUD}",
        f"{C27}!C{R.C27_SC_BUD}",
        f"{C27}!D{R.C27_SC_BUD}",
    ]
    for i, (label, ref) in enumerate(zip(["Conservative", "Proposed", "Higher"], fy27_refs)):
        put(ws, r, 1, label, fill=CALC_FILL)
        for c, gr_row in [(2, R.C28_GR_3), (3, R.C28_GR_5), (4, R.C28_GR_10)]:
            ws.cell(r, c).value = f"={ref}*(1+B{gr_row})+B{R.C28_EC}"
            dat(ws, r, c, CALC_FILL).number_format = USD
        r += 1
    r += 1

    # FY28 with structural changes
    r = sec(ws, r, "FY28 with Structural Changes (savings from C-FY28Paths)")
    R.C28_PA_SAV = r
    put(ws, r, 1, "Path A total savings", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{0}"  # placeholder -- set after C-FY28Paths is built
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.C28_PB_SAV = r
    put(ws, r, 1, "Path B gross savings", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{0}"  # placeholder
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 2

    R.C28_PA_BUD = r
    put(ws, r, 1, "FY28 with Path A", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C28_BASE}-B{R.C28_PA_SAV}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 1

    R.C28_PB_BUD = r
    put(ws, r, 1, "FY28 with Path B (gross)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.C28_BASE}-B{R.C28_PB_SAV}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 1

    for label, bud_row in [("Path A increase from FY27", R.C28_PA_BUD), ("Path B increase from FY27", R.C28_PB_BUD)]:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = f"=(B{bud_row}-B{R.C28_FY27_SEL})/B{R.C28_FY27_SEL}"
        dat(ws, r, 2, CALC_FILL).number_format = PCT2
        r += 1


def build_c_fy28_paths(wb):
    """Detailed Path A / Path B analysis. Replaces the old C-Scenarios sheet with proper FY28 labeling."""
    ws = wb.create_sheet("C-FY28Paths")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [48, 20, 50])
    IB = SN['ib']; IA = SN['ia']; CM = SN['cm']; CB = SN['cb']
    C28 = SN['c28']; CEC = SN['cec']

    r = ttl(ws, 1, "CALC: FY28 Structural Scenario Comparison")
    r = note(ws, r + 1, "Path A and Path B are FY28 (2027-2028) decisions, not FY27.")
    r = note(ws, r, "Savings are independent of the FY28 baseline; they come from restructuring.")
    r += 1

    # ── Path A ──
    r = sec(ws, r, "PATH A: Preserve & Strengthen (FY28)")
    for i, h in enumerate(["Item", "Budget Impact", "Source / Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    pa_items = [
        ("MS consolidation teacher savings",
         f"=-{CM}!B{R.C7_TOTAL}",
         "See C-MSConsol sheet"),
        ("DCS admin change (multi-community K-6)",
         "0",
         "Revised to $0: complexity increases (I-Assumptions)"),
        ("Freeport 6th transportation",
         f"={IA}!B{R.A_FRE6_ROUTES}*{IA}!B{R.A_BUS_COST}",
         "Routes x cost/route from I-Assumptions"),
        ("DCS portable classrooms",
         "0",
         "Sections freed by 7-8 offset 6th additions"),
        ("Additional efficiencies (midpoint)",
         f"=-AVERAGE({IA}!B{R.A_EFF_LOW},{IA}!B{R.A_EFF_HIGH})",
         "Midpoint of low/high range from I-Assumptions"),
    ]
    R.PA_START = r
    for label, formula, note_text in pa_items:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={formula}" if not formula.startswith("=") else formula
        dat(ws, r, 2, CALC_FILL).number_format = SIGNED
        put(ws, r, 3, note_text, fill=CALC_FILL)
        r += 1
    R.PA_END = r - 1

    R.FP_PA_SAV = r
    put(ws, r, 1, "TOTAL PATH A SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-SUM(B{R.PA_START}:B{R.PA_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ── Path B (Gross lens) ──
    r = sec(ws, r, "PATH B: Scenario 2 -- Gross Lens (FY28)")
    r = note(ws, r, "Treats full PES instruction budget as eliminated (traditional closure analysis)")
    for i, h in enumerate(["Item", "Budget Impact", "Source / Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.FP_PB_INSTR = r
    put(ws, r, 1, "PES instruction eliminated (gross)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-({CB}!D{R.C2_DIRECT}-{IB}!D{R.ART9_ROW})"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    put(ws, r, 3, "PES direct cost minus Art 9 facility (stays regardless)", fill=CALC_FILL)
    r += 1

    pb_cost_items = [
        ("Transport: Pownal K-6 to DCS",
         f"={IA}!B{R.A_POW_K6_ROUTES}*{IA}!B{R.A_BUS_COST}",
         "Routes x cost/route from I-Assumptions"),
        ("Transport: Freeport PreK to PES",
         f"={IA}!B{R.A_FRE_PK_ROUTES}*{IA}!B{R.A_BUS_COST}",
         "Specialized EC routes from I-Assumptions"),
        ("Transport: Durham PreK to PES",
         f"={IA}!B{R.A_DUR_PK_ROUTES}*{IA}!B{R.A_BUS_COST}",
         "Specialized EC routes from I-Assumptions"),
        ("DCS multi-community coordination",
         f"={IA}!B{R.A_DCS_COORD}",
         "Analyst estimate from I-Assumptions"),
        ("PES full EC conversion (amortized)",
         f"={IA}!B{R.A_PES_PARTIAL}*{IA}!B{R.A_PES_MULT}/{IA}!B{R.A_AMORT}",
         "Partial cost x multiplier / years from I-Assumptions"),
    ]
    R.PB_COST_START = r
    for label, formula, note_text in pb_cost_items:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={formula}" if not formula.startswith("=") else formula
        dat(ws, r, 2, CALC_FILL).number_format = SIGNED
        put(ws, r, 3, note_text, fill=CALC_FILL)
        r += 1
    R.PB_COST_END = r - 1

    R.FP_PB_SAV = r
    put(ws, r, 1, "PATH B GROSS SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-B{R.FP_PB_INSTR}-SUM(B{R.PB_COST_START}:B{R.PB_COST_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ── Path B (True efficiency lens) ──
    r = sec(ws, r, "PATH B: True Efficiency Lens -- Absorption-Adjusted (FY28)")
    r = note(ws, r, "Accounts for the fact that Pownal students still need teachers at receiving schools.")
    r += 1

    IFT = SN['ift']
    pb2_items = [
        ("Marginal teacher savings (Scenario 2 total - Path A)",
         f"=({IFT}!B{R.FT_SC2_NET}-{CM}!B{R.C7_SAVED})*{IA}!B{R.A_COST_FTE}",
         "I-FTE Scenario 2 net FTE - C-MSConsol Path A FTE = marginal"),
        ("PES administration eliminated",
         f"={IB}!D{R.ART7_ROW}",
         "Full Art 7 PES from I-Budget"),
        ("DCS coordination cost",
         f"=-{IA}!B{R.A_DCS_COORD}",
         "Absorbing Pownal families from I-Assumptions"),
        ("PES support services (partial efficiency)",
         f"={IB}!D{R.ART5_ROW}*{IA}!B{R.A_SUPPORT_EFF}",
         "Art 5 x efficiency % from I-Assumptions"),
    ]
    R.PB2_START = r
    for label, formula, note_text in pb2_items:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={formula}" if not formula.startswith("=") else formula
        dat(ws, r, 2, CALC_FILL).number_format = USD
        put(ws, r, 3, note_text, fill=CALC_FILL)
        r += 1
    R.PB2_END = r - 1

    R.FP_TRUE_EFF = r
    put(ws, r, 1, "True Efficiency Savings", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.PB2_START}:B{R.PB2_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Minus transport (Path B)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-SUM(B{R.PB_COST_START}:B{R.PB_COST_START+2})"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    R.FP_TRANS = r; r += 1

    put(ws, r, 1, "Minus PES conversion (amortized)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-B{R.PB_COST_END}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    R.FP_CONV = r; r += 1

    R.FP_TRUE_NET = r
    put(ws, r, 1, "TRUE NET SAVINGS FROM CLOSING PES", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.FP_TRUE_EFF}+B{R.FP_TRANS}+B{R.FP_CONV}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    r += 2

    # ── Final side-by-side (C28 refs are placeholders; fixed in main block) ──
    r = sec(ws, r, "FINAL COMPARISON (FY28)")
    for i, h in enumerate(["Metric", "No Action (FY28)", "Path A (FY28)", "Path B Gross (FY28)"], 1):
        ws.cell(r, i, h)
        ws.column_dimensions[get_column_letter(i)].width = max(ws.column_dimensions[get_column_letter(i)].width, 20)
    hdr(ws, r, 4); r += 1

    R.FP_FINAL_BUD = r
    put(ws, r, 1, "FY28 Budget", fill=RESULT_FILL, font=BOLD)
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    dat(ws, r, 4, RESULT_FILL).number_format = USD
    r += 1

    R.FP_FINAL_PCT = r
    put(ws, r, 1, "Increase from FY27 adopted", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        dat(ws, r, c, RESULT_FILL).number_format = PCT
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "Savings vs. FY28 Baseline", fill=RESULT_FILL)
    put(ws, r, 2, "--", fill=RESULT_FILL)
    ws.cell(r, 3).value = f"=B{R.FP_PA_SAV}"; dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 4).value = f"=B{R.FP_PB_SAV}"; dat(ws, r, 4, RESULT_FILL).number_format = USD
    r += 1

    put(ws, r, 1, "True Net (absorption-adjusted)", fill=RESULT_FILL)
    put(ws, r, 2, "N/A", fill=RESULT_FILL)
    put(ws, r, 3, "N/A", fill=RESULT_FILL)
    ws.cell(r, 4).value = f"=B{R.FP_TRUE_NET}"; dat(ws, r, 4, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 4).font = RESULT_FONT


def build_c_risk_model(wb):
    """Quantified risk analysis for Path B (closing PES) with cited sources."""
    ws = wb.create_sheet("C-RiskModel")
    ws.sheet_properties.tabColor = "FF4444"
    col_widths(ws, [48, 20, 20, 20, 44])
    IT = SN['it']; IR = SN['ir']; IFT = SN['ift']; IA = SN['ia']
    CFP = SN['cfp']; CEC = SN['cec']; IC = SN['ic']; CM = SN['cm']

    r = ttl(ws, 1, "CALC: Quantified Risk Analysis")
    r = note(ws, r + 1, "Risks that are NOT captured in the C-FY28Paths savings model.")
    r = note(ws, r, "All values are formula-driven or sourced estimates. Citations in brackets.")
    r += 1

    # ═══ Section 1: Property Value Impact ═══
    r = sec(ws, r, "1. PROPERTY VALUE IMPACT (Pownal) [Fn 20][Fn 27]")
    r = note(ws, r, "Research: 5-15% decline in communities losing their school (Duncombe & Yinger 2010;")
    r = note(ws, r, "EdWorkingPapers #22-530, 2022). Applied to Pownal's taxable valuation.")
    r += 1

    put(ws, r, 1, "Pownal taxable valuation", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IT}!B{R.POW_TAXABLE_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_POW_VAL = r; r += 1

    put(ws, r, 1, "Estimated housing units (ACS 2023) [Fn 27]", fill=CALC_FILL)
    put(ws, r, 2, 636, USD)
    R.RM_HOUSES = r; r += 1

    put(ws, r, 1, "Median home value (ACS 2023) [Fn 27]", fill=CALC_FILL)
    put(ws, r, 2, 340000, USD)
    R.RM_MED_HOME = r; r += 1

    put(ws, r, 1, "Total Pownal mil rate", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IT}!B{R.POW_MIL_R}"
    dat(ws, r, 2, CALC_FILL).number_format = '0.000'
    R.RM_MIL = r; r += 1

    put(ws, r, 1, "Pownal RSU share of tax", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IT}!B{R.POW_RSU_PCT_R}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT
    R.RM_RSU_SH = r; r += 1

    for i, h in enumerate(["Scenario", "Decline %", "Total Value Lost",
                            "Per-Household Loss", "Annual Tax Revenue Lost"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.RM_PV_START = r
    for label, pct in [("Low (5%)", 0.05), ("Mid (10%)", 0.10), ("High (15%)", 0.15)]:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, pct, PCT)
        ws.cell(r, 3).value = f"=B{R.RM_POW_VAL}*B{r}"
        dat(ws, r, 3, CALC_FILL).number_format = USD
        ws.cell(r, 4).value = f"=B{R.RM_MED_HOME}*B{r}"
        dat(ws, r, 4, CALC_FILL).number_format = USD
        ws.cell(r, 5).value = f"=C{r}*B{R.RM_MIL}/1000"
        dat(ws, r, 5, CALC_FILL).number_format = USD
        r += 1
    R.RM_PV_END = r - 1

    put(ws, r, 1, "RSU 5 share of mid-scenario tax loss", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=E{R.RM_PV_START+1}*B{R.RM_RSU_SH}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    R.RM_PV_RSU_LOSS = r; r += 1

    r = note(ws, r, "Note: Tax revenue loss occurs gradually (reassessment cycles). Wealth loss is immediate.")
    r = note(ws, r, "At 10% decline, median Pownal home loses ~$34,000 in value.")
    r += 1

    # ═══ Section 2: Enrollment Attrition ═══
    r = sec(ws, r, "2. ENROLLMENT ATTRITION RISK [Fn 20][Fn 28]")
    r = note(ws, r, "If PES closes, some families may leave RSU 5 (move, homeschool, private).")
    r = note(ws, r, "State aid follows the student: each departure costs RSU 5 ongoing revenue.")
    r += 1

    put(ws, r, 1, "Current PES enrollment", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IFT}!C{R.FT_PES}"
    dat(ws, r, 2, CALC_FILL).number_format = '0'
    R.RM_PES_ENROLL = r; r += 1

    put(ws, r, 1, "Pownal state aid (FY25-26)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IT}!E{R.EPS_POW_R}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_POW_AID = r; r += 1

    put(ws, r, 1, "State aid per PES student", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.RM_POW_AID}/B{R.RM_PES_ENROLL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_AID_PP = r; r += 1

    for i, h in enumerate(["Scenario", "Attrition %", "Students Lost",
                            "Annual State Revenue Lost"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.RM_EA_START = r
    for label, pct in [("Low (5%)", 0.05), ("Mid (10%)", 0.10), ("High (20%)", 0.20)]:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, pct, PCT)
        ws.cell(r, 3).value = f"=ROUND(B{R.RM_PES_ENROLL}*B{r},0)"
        dat(ws, r, 3, CALC_FILL).number_format = '0'
        ws.cell(r, 4).value = f"=C{r}*B{R.RM_AID_PP}"
        dat(ws, r, 4, CALC_FILL).number_format = USD
        r += 1
    R.RM_EA_END = r - 1

    r = note(ws, r, "Cost side does NOT shrink proportionally: fixed costs remain, class sizes are already small.")
    r = note(ws, r, "At 10% attrition: 10-11 students leave, ~$50K/yr in state revenue lost to RSU 5.")
    r += 1

    # ═══ Section 3: EC Opt-Out ═══
    r = sec(ws, r, "3. EC OPT-OUT RISK (Path B Only)")
    r = note(ws, r, "Under Path B, all PreK goes to PES in Pownal. Freeport/Durham families")
    r = note(ws, r, "may refuse 22+ mile bus ride for 3-year-olds and choose private childcare instead.")
    r = note(ws, r, "District PreK is voluntary -- families cannot be compelled to enroll.")
    r += 1

    put(ws, r, 1, "Total district PreK enrollment (all schools)", fill=CALC_FILL)
    put(ws, r, 2, 128, '0')
    put(ws, r, 3, "MSS 64 + DCS 48 + PES 16", fill=CALC_FILL)
    R.RM_EC_TOTAL = r; r += 1

    put(ws, r, 1, "EC gross annual cost", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CEC}!B{R.C8_GROSS}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_EC_GROSS = r; r += 1

    put(ws, r, 1, "EC cost per student (baseline)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.RM_EC_GROSS}/B{R.RM_EC_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_EC_PP = r; r += 1

    for i, h in enumerate(["Scenario", "Opt-Out %", "Students Lost",
                            "Cost/Remaining Student", "SpEd LRE Risk"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.RM_OPT_START = r
    for label, pct, lre in [("Low (10%)", 0.10, "Moderate"),
                             ("Mid (20%)", 0.20, "Significant"),
                             ("High (30%)", 0.30, "Severe")]:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, pct, PCT)
        ws.cell(r, 3).value = f"=ROUND(B{R.RM_EC_TOTAL}*B{r},0)"
        dat(ws, r, 3, CALC_FILL).number_format = '0'
        ws.cell(r, 4).value = f"=B{R.RM_EC_GROSS}/(B{R.RM_EC_TOTAL}-C{r})"
        dat(ws, r, 4, CALC_FILL).number_format = USD
        put(ws, r, 5, lre, fill=CALC_FILL)
        r += 1
    R.RM_OPT_END = r - 1

    r = note(ws, r, "LRE = Least Restrictive Environment. SpEd requires peers for inclusion classrooms.")
    r = note(ws, r, "Fewer typical peers = harder to deliver compliant services = potential legal exposure.")
    r = note(ws, r, "EC offsets (CDS transfer, IDEA Part B, EPS PreK subsidy) drop with enrollment.")
    r += 1

    # ═══ Section 4: One-Time Transition Costs ═══
    r = sec(ws, r, "4. ONE-TIME TRANSITION COSTS [Fn 28]")
    r = note(ws, r, "Costs incurred during implementation year. Not captured in recurring savings model.")
    r += 1

    for i, h in enumerate(["Item", "Low Estimate", "High Estimate", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    onetime = [
        ("Staff transition (unemployment, reassignment)",
         60000, 150000, "4-6 positions affected; depends on attrition vs. layoff"),
        ("PES building conversion to EC center",
         425000, 743750, "From I-Assumptions; Ch.124 compliance [Fn 21][Fn 24]"),
        ("Legal/process costs (hearings, consultants)",
         50000, 200000, "Maine 20-A MRSA requires public process [Fn 28]"),
        ("DCS portable classrooms (if needed)",
         0, 300000, "Depends on final grade configuration at DCS"),
        ("Curriculum/technology transfer",
         25000, 75000, "Materials, equipment relocation, IT setup"),
        ("Community engagement/transition support",
         15000, 50000, "Notification, town meetings, family support"),
    ]
    R.RM_OT_START = r
    for label, low, high, notes in onetime:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, low, USD)
        put(ws, r, 3, high, USD)
        put(ws, r, 4, notes, fill=CALC_FILL)
        r += 1
    R.RM_OT_END = r - 1

    put(ws, r, 1, "TOTAL ONE-TIME COSTS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.RM_OT_START}:B{R.RM_OT_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"=SUM(C{R.RM_OT_START}:C{R.RM_OT_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; ws.cell(r, 3).font = RESULT_FONT
    R.RM_OT_TOTAL = r; r += 2

    # ═══ Section 5: Risk-Adjusted Net Savings ═══
    r = sec(ws, r, "5. RISK-ADJUSTED NET SAVINGS (Path B vs. Path A)")
    r = note(ws, r, "Combines C-FY28Paths true-efficiency savings with quantified annual risk costs.")
    r += 1

    put(ws, r, 1, "Path A annual savings (from C-FY28Paths)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.FP_PA_SAV}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_PA_SAV = r; r += 1

    put(ws, r, 1, "Path B true net savings (from C-FY28Paths)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.FP_TRUE_NET}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    R.RM_PB_NET = r; r += 1

    put(ws, r, 1, "Path B MARGINAL savings beyond Path A", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.RM_PB_NET}-B{R.RM_PA_SAV}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    R.RM_MARGINAL = r; r += 2

    for i, h in enumerate(["Annual Risk Factor", "Low", "Mid", "High"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.RM_RISK_START = r
    put(ws, r, 1, "Property value tax revenue loss", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=E{R.RM_PV_START}*B{R.RM_RSU_SH}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=E{R.RM_PV_START+1}*B{R.RM_RSU_SH}"
    dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"=E{R.RM_PV_END}*B{R.RM_RSU_SH}"
    dat(ws, r, 4, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Enrollment attrition revenue loss", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=D{R.RM_EA_START}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=D{R.RM_EA_START+1}"
    dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"=D{R.RM_EA_END}"
    dat(ws, r, 4, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "One-time costs (amortized 5 years)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.RM_OT_TOTAL}/5"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=AVERAGE(B{R.RM_OT_TOTAL},C{R.RM_OT_TOTAL})/5"
    dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"=C{R.RM_OT_TOTAL}/5"
    dat(ws, r, 4, CALC_FILL).number_format = USD; r += 1
    R.RM_RISK_END = r - 1

    put(ws, r, 1, "TOTAL ANNUAL RISK COST", fill=RESULT_FILL, font=BOLD)
    for c, cl in [(2, 'B'), (3, 'C'), (4, 'D')]:
        ws.cell(r, c).value = f"=SUM({cl}{R.RM_RISK_START}:{cl}{R.RM_RISK_END})"
        dat(ws, r, c, RESULT_FILL).number_format = USD
    ws.cell(r, 3).font = RESULT_FONT
    R.RM_TOTAL_RISK = r; r += 2

    put(ws, r, 1, "PATH B RISK-ADJUSTED NET (mid scenario)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.RM_PB_NET}-C{R.RM_TOTAL_RISK}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    R.RM_PB_RISK_ADJ = r; r += 1

    put(ws, r, 1, "Path A savings (unchanged -- no closure risks)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.RM_PA_SAV}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    r = sec(ws, r, "6. BREAK-EVEN ANALYSIS")
    r = note(ws, r, "How long until Path B's one-time costs are recovered by marginal savings (if any)?")
    r += 1
    put(ws, r, 1, "Path B marginal annual benefit (true net - Path A)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.RM_MARGINAL}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    R.RM_BE_MARGIN = r; r += 1

    put(ws, r, 1, "One-time costs (midpoint)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE(B{R.RM_OT_TOTAL},C{R.RM_OT_TOTAL})"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    R.RM_BE_ONETIME = r; r += 1

    put(ws, r, 1, "Break-even period (years)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f'=IF(B{R.RM_BE_MARGIN}>0,B{R.RM_BE_ONETIME}/B{R.RM_BE_MARGIN},"NEVER")'
    dat(ws, r, 2, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 2).font = RESULT_FONT
    R.RM_BREAKEVEN = r; r += 1

    r = note(ws, r, "If marginal annual savings are negative, one-time costs are NEVER recovered.")
    r = note(ws, r, "This means Path B is strictly worse than Path A from year one onward.")
    r += 2

    # ═══ Section 7: Path A Risks ═══
    r = sec(ws, r, "7. PATH A RISKS (for balanced analysis)")
    r = note(ws, r, "Path A also carries implementation risks, though substantially lower:")
    r += 1
    pa_risks = [
        ("MS consolidation: transport cost for Freeport 6th to DCS",
         f"={IA}!B{R.A_FRE6_ROUTES}*{IA}!B{R.A_BUS_COST}", "Already modeled in Path A budget"),
        ("Community identity: DCS becomes multi-town 6th grade",
         "0", "No direct $ cost, but political sensitivity"),
        ("Additional efficiencies depend on Board action",
         "0", "$325K-$650K range is aspirational, not guaranteed"),
        ("DCS scheduling complexity with Pownal 6th graders",
         "0", "Manageable with existing admin"),
    ]
    for label, formula, notes in pa_risks:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={formula}" if not formula.startswith("=") else formula
        dat(ws, r, 2, CALC_FILL).number_format = SIGNED
        put(ws, r, 3, notes, fill=CALC_FILL)
        r += 1
    r = note(ws, r, "Path A risks are primarily political/logistical, NOT financial or irreversible.")
    r = note(ws, r, "Path A is reversible: if MS consolidation doesn't work, grades can be redistributed.")


def build_i_withdrawal(wb):
    """Input data for Pownal independence / RSU 5 withdrawal modeling."""
    ws = wb.create_sheet("I-Withdrawal")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [52, 20, 20, 50])

    r = ttl(ws, 1, "INPUT: RSU 5 Withdrawal & Independence Data")
    r = source_block(ws, r + 1, [
        "Data for modeling Pownal's options if RSU 5 proceeds with PES conversion.",
        "Includes current RSU assessment, independent operating estimates, tuition rates,",
        "EPS allocations, withdrawal process costs, and legal challenge estimates.",
        "Sources cited per item. See Sources sheet for full citations.",
    ])
    r += 1

    # ── Section 1: Pownal's Current RSU 5 Financial Position ──
    r = sec(ws, r, "1. POWNAL'S CURRENT RSU 5 POSITION [Fn 3][Fn 8]")
    for i, h in enumerate(["Item", "Value", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_RSU_START = r
    put(ws, r, 1, "Pownal Required Local Contribution (RLC)")
    put(ws, r, 2, 2257247, USD); put(ws, r, 3, "FY27 Budget Handbook [Fn 3]"); r += 1
    put(ws, r, 1, "Pownal Additional Local Monies (ALM)")
    put(ws, r, 2, 2190978, USD); put(ws, r, 3, "FY27 Budget Handbook [Fn 3]"); r += 1
    R.WD_POW_LOCAL = r
    put(ws, r, 1, "Pownal LOCAL Tax (RLC + ALM)")
    ws.cell(r, 2).value = f"=B{R.WD_RSU_START}+B{R.WD_RSU_START+1}"
    dat(ws, r, 2).number_format = USD
    put(ws, r, 3, "What Pownal taxpayers actually pay"); r += 1
    R.WD_POW_STATE_AID = r
    put(ws, r, 1, "Pownal State Aid")
    put(ws, r, 2, 567180, USD); put(ws, r, 3, "FY27 Budget Handbook [Fn 3]"); r += 1
    R.WD_POW_TOTAL = r
    put(ws, r, 1, "Pownal Total Contribution (local + state)")
    put(ws, r, 2, 5015405, USD); put(ws, r, 3, "RLC + ALM + State Aid [Fn 3]"); r += 1
    put(ws, r, 1, "Pownal Share of RSU 5 Budget")
    put(ws, r, 2, 0.106, PCT); put(ws, r, 3, "~10.6% of $47.36M total [Fn 25]"); r += 1
    R.WD_RSU_END = r - 1
    r += 1

    # ── Section 2: PES Current Operating Cost ──
    r = sec(ws, r, "2. PES CURRENT OPERATING COST [Fn 2]")
    r = source_block(ws, r, [
        "PES budget detail from I-Budget sheet. These are the costs RSU 5 currently spends on PES.",
    ])
    for i, h in enumerate(["Budget Article", "PES Amount", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_PES_START = r
    pes_articles = [
        ("Art 1 - Regular Instruction", 1196683, "I-Budget [Fn 2]"),
        ("Art 2 - Special Education", 329115, "I-Budget [Fn 2]"),
        ("Art 4 - Other Instruction", 13125, "I-Budget [Fn 2]"),
        ("Art 5 - Student & Staff Support", 219819, "I-Budget [Fn 2]"),
        ("Art 7 - School Administration", 257979, "I-Budget [Fn 2]"),
        ("Art 9 - Facilities & Maintenance", 253384, "I-Budget [Fn 2]"),
    ]
    for label, val, src in pes_articles:
        put(ws, r, 1, label); put(ws, r, 2, val, USD); put(ws, r, 3, src); r += 1
    R.WD_PES_END = r - 1

    R.WD_PES_SUBTOTAL = r
    put(ws, r, 1, "PES School-Level Subtotal", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_PES_START}:B{R.WD_PES_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1
    R.WD_PES_ENROLL = r
    put(ws, r, 1, "PES Enrollment (FY27)"); put(ws, r, 2, 89, '0')
    put(ws, r, 3, "Budget Handbook [Fn 1]"); r += 1
    R.WD_PES_PP = r
    put(ws, r, 1, "PES Cost Per Student", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.WD_PES_SUBTOTAL}/B{R.WD_PES_ENROLL}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 2

    # ── Section 3: Independent Operation Cost Estimates ──
    r = sec(ws, r, "3. INDEPENDENT OPERATION: ADDITIONAL COSTS [Fn 37][Fn 35]")
    r = source_block(ws, r, [
        "Costs Pownal would incur beyond PES school-level budget if operating independently.",
        "Estimates derived from comparable Maine independent school units and state averages.",
        "Low/high ranges reflect uncertainty in cost estimates for a new SAU.",
        "SOURCE: Maine Monitor withdrawal explainer [Fn 37]; DOE tuition rates [Fn 36].",
    ])
    for i, h in enumerate(["Cost Category", "Low Estimate", "High Estimate", "Basis / Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.WD_IND_START = r
    ind_items = [
        ("Superintendent (shared or part-time)", 95000, 95000,
         "50% of ~$190K salary+benefits; comparable AOS [Fn 37]"),
        ("Business Manager / Finance", 75000, 75000,
         "Part-time or shared; required for SAU reporting"),
        ("SpEd Director (shared or contracted)", 65000, 65000,
         "50% of ~$130K; share via AOS or contract"),
        ("Central office support (HR, payroll, records)", 55000, 55000,
         "1 FTE clerical + systems; comparable small-town SAUs"),
        ("Transportation (K-5 PES routes)", 225000, 225000,
         "3 routes at $75K/route; current RSU 5 Pownal alloc."),
        ("Transportation (6-12 secondary routes)", 100000, 100000,
         "1-2 routes Pownal->Freeport, ~12mi, $50-75K/route"),
    ]
    for label, low, high, src in ind_items:
        put(ws, r, 1, label); put(ws, r, 2, low, USD); put(ws, r, 3, high, USD)
        put(ws, r, 4, src); r += 1
    R.WD_SEC_TRANSPORT = R.WD_IND_START + 5

    R.WD_TUITION_68 = r
    put(ws, r, 1, "Tuition: grades 6-8 (elementary, §5804)", fill=CALC_FILL)
    put(ws, r, 4, "= grade 6-8 students x §5804 rate [Fn 36]"); r += 1
    R.WD_TUITION_912 = r
    put(ws, r, 1, "Tuition: grades 9-12 (secondary, §5805)", fill=CALC_FILL)
    put(ws, r, 4, "= grade 9-12 students x §5805 capped rate [Fn 36]"); r += 1
    R.WD_TUITION_LINE = R.WD_TUITION_68

    R.WD_SPED_TUITION = r
    put(ws, r, 1, "SpEd for tuitioned students (§7302)")
    put(ws, r, 2, 100000, USD); put(ws, r, 3, 200000, USD)
    put(ws, r, 4, "Est. 15-20 IEP students at $5K-$20K each"); r += 1

    ind_other = [
        ("Facilities/maintenance supplement", 50000, 50000,
         "Beyond Art 9 in PES budget; grounds, minor repairs"),
        ("Curriculum and assessment", 15000, 15000,
         "Standardized testing, instructional materials"),
        ("IT infrastructure (SIS, network, cybersecurity)", 30000, 50000,
         "Student info system, network mgmt, email, security"),
        ("Food service (net after USDA reimbursement)", 20000, 60000,
         "NSLP-compliant lunch program; §1466(4)(A)(12)"),
        ("Annual audit (state requirement)", 10000, 15000,
         "Maine law requires annual audit of every SAU"),
        ("District-level insurance (liability, E&O, WC)", 15000, 30000,
         "Beyond building-level insurance in PES Art 9"),
        ("Capital reserve and bus replacement fund", 15000, 35000,
         "Building reserve + bus amort. (~$100K/bus/10-15yr)"),
        ("Board/governance costs", 20000, 30000,
         "Legal counsel retainer, meeting costs, elections"),
        ("Contingency (5% of above)", 0, 0,
         "Calculated as formula"),
    ]
    for label, low, high, src in ind_other:
        put(ws, r, 1, label)
        if low > 0:
            put(ws, r, 2, low, USD)
        if high > 0:
            put(ws, r, 3, high, USD)
        put(ws, r, 4, src)
        r += 1
    R.WD_IND_END = r - 1
    R.WD_CONTINGENCY = R.WD_IND_END
    ws.cell(R.WD_CONTINGENCY, 2).value = f"=SUM(B{R.WD_IND_START}:B{R.WD_IND_END-1})*0.05"
    dat(ws, R.WD_CONTINGENCY, 2, INPUT_FILL).number_format = USD
    ws.cell(R.WD_CONTINGENCY, 3).value = f"=SUM(C{R.WD_IND_START}:C{R.WD_IND_END-1})*0.05"
    dat(ws, R.WD_CONTINGENCY, 3, INPUT_FILL).number_format = USD

    R.WD_IND_TOTAL = r
    put(ws, r, 1, "TOTAL ADDITIONAL COSTS (independence)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_IND_START}:B{R.WD_IND_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).value = f"=SUM(C{R.WD_IND_START}:C{R.WD_IND_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 3).font = RESULT_FONT; r += 2

    # ── Section 4: AOS Shared Services Savings ──
    r = sec(ws, r, "4. AOS SHARED SERVICES SAVINGS ESTIMATE")
    r = source_block(ws, r, [
        "If Pownal joins an Alternative Organizational Structure (AOS), admin costs are shared.",
        "AOS #94 model: 6 communities share superintendent, SpEd, business office.",
        "Bristol/South Bristol formed 2-town ESC after leaving AOS 93 (2025). [Fn 37]",
    ])
    for i, h in enumerate(["Shared Service", "Estimated Savings", "Rationale"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_AOS_START = r
    aos_items = [
        ("Superintendent cost reduction", 47500, "Share with 1-2 other towns -> 50% of $95K"),
        ("SpEd director cost reduction", 32500, "Share with partner town(s) -> 50% of $65K"),
        ("Business office cost reduction", 25000, "Share finance/HR -> ~33% of $75K"),
        ("Transportation coordination", 15000, "Shared routing/dispatch with partner"),
    ]
    for label, val, src in aos_items:
        put(ws, r, 1, label); put(ws, r, 2, val, USD); put(ws, r, 3, src); r += 1
    R.WD_AOS_END = r - 1

    R.WD_AOS_TOTAL = r
    put(ws, r, 1, "TOTAL AOS SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_AOS_START}:B{R.WD_AOS_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ── Section 5: Tuition Rates ──
    r = sec(ws, r, "5. MAINE TUITION RATES (FY26) [Fn 36]")
    r = source_block(ws, r, [
        "SOURCE: Maine DOE Public School Tuition Rates (FY26) [Fn 36]",
        "URL: https://www.maine.gov/doe/funding/reports/tuition",
        "Pownal would tuition secondary students (6-12) to Freeport or other districts.",
        "Year 1 (§1466(4)(A)(1)): secondary tuition is uncapped (actual cost).",
        "Year 2+: secondary tuition capped at state average per §5806(2).",
    ])
    for i, h in enumerate(["Item", "Rate", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_TUIT_START = r
    R.WD_TUIT_ELEM = r
    put(ws, r, 1, "Elementary tuition rate (§5804, grades 6-8)")
    put(ws, r, 2, 15417, USD); put(ws, r, 3, "Receiving district per-pupil cost, FY26 [Fn 36]"); r += 1
    R.WD_TUIT_SEC = r
    put(ws, r, 1, "Secondary tuition rate (§5805, grades 9-12)")
    put(ws, r, 2, 15055, USD); put(ws, r, 3, "Capped at state avg, FY26 DOE [Fn 36]"); r += 1
    R.WD_TUIT_RATE = R.WD_TUIT_SEC
    R.WD_TUIT_UNCAPPED = r
    put(ws, r, 1, "Year 1 uncapped secondary rate (RSU 5 actual)")
    put(ws, r, 2, 17307, USD); put(ws, r, 3, "RSU 5 computed per-pupil cost, FY26 [Fn 36]"); r += 1
    R.WD_TUIT_PENALTY = r
    put(ws, r, 1, "Year 1 tuition penalty (uncapped - capped, per student)")
    ws.cell(r, 2).value = f"=B{R.WD_TUIT_UNCAPPED}-B{R.WD_TUIT_SEC}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "Additional cost per student in withdrawal Year 1"); r += 1
    R.WD_TUIT_END = r - 1
    r += 1

    # ── Section 6: Secondary Student Count ──
    r = sec(ws, r, "6. POWNAL SECONDARY STUDENTS [Fn 1]")
    r = source_block(ws, r, [
        "Estimated from PES cohort size. PES has ~15 students/grade historically.",
        "Current 6-12 students entered PES when enrollment was lower (~89-97).",
        "As PES grows (now 105), future steady-state secondary count will be higher.",
        "Matches I-Equity attribution methodology for consistency.",
    ])
    for i, h in enumerate(["Grade Band", "Count", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_SEC_START = r
    put(ws, r, 1, "Grades 6-8 (Pownal residents at FMS)")
    put(ws, r, 2, 45, '0'); put(ws, r, 3, "PES avg ~15/grade x 3 grades [Fn 1]; matches I-Equity"); r += 1
    put(ws, r, 1, "Grades 9-12 (Pownal residents at FHS)")
    put(ws, r, 2, 55, '0'); put(ws, r, 3, "PES avg ~14/grade x 4 grades [Fn 1]; matches I-Equity"); r += 1
    R.WD_SEC_END = r - 1
    R.WD_SEC_TOTAL = r
    put(ws, r, 1, "Total secondary students (current)", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_SEC_START}:B{R.WD_SEC_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = '0'; r += 1

    R.WD_SEC_STEADY = r
    put(ws, r, 1, "Steady-state estimate (at PES 105 enrollment)", fill=CALC_FILL)
    put(ws, r, 2, 123, '0')
    put(ws, r, 3, "105/6 grades ~ 18/grade; 18x3=54 MS + 18x4=72 HS. Higher than current."); r += 1

    r = note(ws, r, "SENSITIVITY: If actual count is 100-123, tuition cost = $1.41M-$1.73M.")
    r = note(ws, r, "This is the largest variable in the independence budget.")
    r += 1

    # Backfill split tuition formulas in Section 3
    R.WD_SEC_68 = R.WD_SEC_START
    R.WD_SEC_912 = R.WD_SEC_START + 1
    ws.cell(R.WD_TUITION_68, 2).value = f"=B{R.WD_SEC_68}*B{R.WD_TUIT_ELEM}"
    dat(ws, R.WD_TUITION_68, 2, CALC_FILL).number_format = USD
    ws.cell(R.WD_TUITION_68, 3).value = f"=B{R.WD_SEC_68}*B{R.WD_TUIT_ELEM}"
    dat(ws, R.WD_TUITION_68, 3, CALC_FILL).number_format = USD
    ws.cell(R.WD_TUITION_912, 2).value = f"=B{R.WD_SEC_912}*B{R.WD_TUIT_SEC}"
    dat(ws, R.WD_TUITION_912, 2, CALC_FILL).number_format = USD
    ws.cell(R.WD_TUITION_912, 3).value = f"=B{R.WD_SEC_912}*B{R.WD_TUIT_SEC}"
    dat(ws, R.WD_TUITION_912, 3, CALC_FILL).number_format = USD

    # ── Section 7: Withdrawal Process Costs ──
    r = sec(ws, r, "7. WITHDRAWAL PROCESS COSTS [Fn 32][Fn 37]")
    r = source_block(ws, r, [
        "One-time costs to execute the 22-step withdrawal process under Title 20-A §1466.",
        "Embden (RSU 74) authorized $50K; Phillips authorized $50K. [Fn 37]",
        "Freeport's 2014 RSU 5 withdrawal committee incurred similar costs. [Fn 33]",
    ])
    for i, h in enumerate(["Item", "Estimate", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_PROC_START = r
    proc_items = [
        ("Legal counsel (withdrawal committee)", 35000, "Comparable to Embden/Phillips [Fn 37]"),
        ("Financial analysis and consulting", 10000, "Budget modeling, EPS projections"),
        ("Public hearings and elections", 5000, "2-3 votes required by statute [Fn 32]"),
        ("Negotiations with RSU 5 Board", 5000, "Committee time, meeting costs"),
        ("Administrative setup (new SAU)", 15000, "Payroll systems, reporting, accounts"),
    ]
    for label, val, src in proc_items:
        put(ws, r, 1, label); put(ws, r, 2, val, USD); put(ws, r, 3, src); r += 1
    R.WD_PROC_END = r - 1

    R.WD_PROC_TOTAL = r
    put(ws, r, 1, "TOTAL PROCESS COSTS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_PROC_START}:B{R.WD_PROC_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ── Section 7A: Additional Startup Costs (One-Time) ──
    r = sec(ws, r, "7A. ADDITIONAL STARTUP COSTS (ONE-TIME)")
    r = source_block(ws, r, [
        "One-time costs in Year 1 beyond the withdrawal process costs in Section 7.",
        "These are operational startup costs for establishing the new SAU.",
    ])
    for i, h in enumerate(["Item", "Low", "High"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_STARTUP_START = r
    startup_items = [
        ("IT system setup (SIS, payroll, email migration)", 15000, 30000),
        ("Insurance policy setup and first premium delta", 5000, 10000),
        ("CBA transition legal review", 10000, 20000),
        ("Superintendent contract obligation share", 10000, 30000),
    ]
    for label, low, high in startup_items:
        put(ws, r, 1, label); put(ws, r, 2, low, USD); put(ws, r, 3, high, USD); r += 1
    R.WD_STARTUP_END = r - 1

    R.WD_STARTUP_TOTAL = r
    put(ws, r, 1, "TOTAL STARTUP COSTS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_STARTUP_START}:B{R.WD_STARTUP_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"=SUM(C{R.WD_STARTUP_START}:C{R.WD_STARTUP_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).font = RESULT_FONT; r += 2

    # ── Section 8: Legal Challenge Costs ──
    r = sec(ws, r, "8. LEGAL CHALLENGE COST ESTIMATES [Fn 30][Fn 39]")
    r = source_block(ws, r, [
        "If Pownal challenges the 'reconfiguration vs. closure' framing under §1512.",
        "Legal costs for petition to Commissioner or court challenge.",
    ])
    for i, h in enumerate(["Item", "Low", "High"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_LEGAL_START = r
    legal_items = [
        ("Legal counsel for §1512 challenge", 15000, 50000),
        ("Expert witness / education law consultant", 5000, 20000),
        ("Court filing and process costs", 2000, 10000),
        ("Commissioner petition preparation", 3000, 8000),
    ]
    for label, low, high in legal_items:
        put(ws, r, 1, label); put(ws, r, 2, low, USD); put(ws, r, 3, high, USD); r += 1
    R.WD_LEGAL_END = r - 1

    R.WD_LEGAL_TOTAL = r
    put(ws, r, 1, "TOTAL LEGAL CHALLENGE COSTS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.WD_LEGAL_START}:B{R.WD_LEGAL_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"=SUM(C{R.WD_LEGAL_START}:C{R.WD_LEGAL_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).font = RESULT_FONT; r += 2

    # ── Section 9: EPS Independent Allocation Estimate ──
    r = sec(ws, r, "9. EPS ALLOCATION IF INDEPENDENT [Fn 8]")
    r = source_block(ws, r, [
        "Pownal's EPS allocation would follow the student. Current RSU 5 allocation shown.",
        "Isolated small school adjustment may apply (<200 students). [Fn 8]",
        "State share depends on property valuation vs. enrollment weighting.",
    ])
    for i, h in enumerate(["Item", "Value", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_EPS_START = r
    put(ws, r, 1, "Pownal total EPS allocation (FY25-26)")
    put(ws, r, 2, 2682829, USD); put(ws, r, 3, "Warrant Article F [Fn 8]"); r += 1
    put(ws, r, 1, "Pownal EPS state share (FY25-26)")
    put(ws, r, 2, 503502, USD); put(ws, r, 3, "Warrant Article F [Fn 8]"); r += 1
    R.WD_EPS_STATE = R.WD_EPS_START + 1
    put(ws, r, 1, "Isolated small school adjustment (est.)")
    put(ws, r, 2, 75000, USD); put(ws, r, 3, "Analyst est.; DOE formula for <200 students [Fn 8]"); r += 1
    R.WD_EPS_SMALL = r - 1
    put(ws, r, 1, "Pownal taxable valuation")
    put(ws, r, 2, 392398240, USD); put(ws, r, 3, "Commitment Book [Fn 4]"); r += 1
    R.WD_VALUATION = r - 1
    put(ws, r, 1, "Housing units (ACS 2023)")
    put(ws, r, 2, 636, '0'); put(ws, r, 3, "Census ACS [Fn 27]"); r += 1
    R.WD_HOUSING = r - 1
    put(ws, r, 1, "Current total mil rate")
    put(ws, r, 2, 15.300, '0.000'); put(ws, r, 3, "Commitment Book [Fn 4]"); r += 1
    R.WD_MIL = r - 1
    put(ws, r, 1, "Current RSU share of mil rate")
    put(ws, r, 2, 8.935, '0.000'); put(ws, r, 3, "15.300 x 58.4% [Fn 4][Fn 12]"); r += 1
    R.WD_RSU_MIL = r - 1
    R.WD_EPS_END = r - 1
    r += 1

    # ── Section 9A: Fund Balance Return (One-Time Credit) ──
    r = sec(ws, r, "9A. FUND BALANCE RETURN (ONE-TIME CREDIT)")
    r = source_block(ws, r, [
        "Under §1466(4)(A)(10), Pownal receives its share of RSU 5 undesignated fund balance.",
        "Pownal's share is approximately 12.6% based on assessment ratio.",
        "RSU 5 undesignated fund balance estimated at $1.5M-$2.5M.",
    ])
    for i, h in enumerate(["Item", "Low", "High"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_FUND_START = r
    put(ws, r, 1, "Pownal's 12.6% share of RSU 5 fund balance")
    put(ws, r, 2, 190000, USD); put(ws, r, 3, 315000, USD); r += 1
    R.WD_FUND_END = r - 1

    R.WD_FUND_TOTAL = r
    put(ws, r, 1, "TOTAL FUND BALANCE RETURN", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.WD_FUND_START}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"=C{R.WD_FUND_START}"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).font = RESULT_FONT; r += 2

    # ── Section 10: Precedent Data ──
    r = sec(ws, r, "10. WITHDRAWAL PRECEDENTS [Fn 33][Fn 34][Fn 35]")
    for i, h in enumerate(["Town / District", "Year", "Enrollment", "Mil Rate Change",
                            "Budget Impact"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.WD_PREC_START = r
    precedents = [
        ("Freeport / RSU 5 (FAILED)", "2014", "N/A", "N/A",
         "Vote failed 2152-2228 (76 margin) [Fn 33]"),
        ("Andover / SAD 44", "2015", "~30 K-5", "15.6 -> 19.6",
         "Budget $700K -> $1.5M; full independence"),
        ("Dayton / RSU 23", "2014", "~200 K-5", "18.47 -> 20.85",
         "Budget to $2.9M; gained full control"),
        ("Athens / SAD 59", "2013", "small", "N/A",
         "Successfully withdrew"),
        ("Embden / RSU 74", "2025", "~50", "TBD",
         "Process ongoing; authorized $50K [Fn 37]"),
        ("Buckfield+Hartford+Sumner / RSU 10", "2025", "~150 combined", "TBD",
         "3-town withdrawal in progress [Fn 37]"),
    ]
    for town, yr, enroll, mil, impact in precedents:
        put(ws, r, 1, town); put(ws, r, 2, yr)
        put(ws, r, 3, enroll); put(ws, r, 4, mil)
        put(ws, r, 5, impact); r += 1
    R.WD_PREC_END = r - 1
    r += 1

    # ── Section 11: K-8 PES Expansion Option ──
    r = sec(ws, r, "11. K-8 PES EXPANSION OPTION (Scenario E inputs)")
    r = source_block(ws, r, [
        "If PES expands from K-5 to K-8, Pownal only tuitions grades 9-12.",
        "PES capacity: designed for ~180 students; K-8 at ~150-160 would fit.",
        "Requires DOE approval, middle school content teachers, possible facility upgrades.",
        "Tradeoff: higher operating cost vs. lower tuition spend and more local control.",
    ])
    for i, h in enumerate(["Item", "Value", "Source / Rationale"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    R.WD_K8_START = r
    k8_items = [
        ("Grades 6-8 Pownal students (current)", None,
         "From Section 6 above"),
        ("Additional teaching FTE for grades 6-8", 5.0,
         "Core: ELA, Math, Science, Social Studies + 1 specials FTE"),
        ("Cost per FTE (salary + benefits)", 108000,
         "RSU 5 average compensation [Fn 2]"),
        ("Annual staffing cost (6-8 teachers)", None,
         "= FTE x cost per FTE (formula)"),
        ("Curriculum / materials (one-time startup)", 60000,
         "Middle school curriculum, textbooks, lab supplies, technology"),
        ("Curriculum / materials (ongoing annual)", 15000,
         "Annual replenishment, testing, software licenses"),
        ("Facility modifications (one-time)", 75000,
         "Science lab upgrade, lockers, minor reconfiguration for MS"),
        ("Additional extracurricular / athletics", 25000,
         "MS sports, clubs; limited by small cohort size (~45 students)"),
    ]
    for label, val, src in k8_items:
        put(ws, r, 1, label)
        put(ws, r, 3, src)
        if val is not None:
            if isinstance(val, float):
                put(ws, r, 2, val, '0.0')
            else:
                put(ws, r, 2, val, USD)
        r += 1

    R.WD_K8_STUDENTS = R.WD_K8_START
    ws.cell(R.WD_K8_STUDENTS, 2).value = f"=B{R.WD_SEC_START}"
    dat(ws, R.WD_K8_STUDENTS, 2, CALC_FILL).number_format = '0'

    R.WD_K8_FTE = R.WD_K8_START + 1
    R.WD_K8_COST_FTE = R.WD_K8_START + 2
    R.WD_K8_STAFFING = R.WD_K8_START + 3
    ws.cell(R.WD_K8_STAFFING, 2).value = f"=B{R.WD_K8_FTE}*B{R.WD_K8_COST_FTE}"
    dat(ws, R.WD_K8_STAFFING, 2, CALC_FILL).number_format = USD

    R.WD_K8_STARTUP = R.WD_K8_START + 4
    R.WD_K8_ONGOING = R.WD_K8_START + 5
    R.WD_K8_FACILITY = R.WD_K8_START + 6
    R.WD_K8_EXTRA = R.WD_K8_START + 7

    R.WD_K8_ANNUAL = r
    put(ws, r, 1, "TOTAL ANNUAL ADDED COST (K-8 vs K-5)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.WD_K8_STAFFING}+B{R.WD_K8_ONGOING}+B{R.WD_K8_EXTRA}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.WD_K8_HS_STUDENTS = r
    put(ws, r, 1, "Remaining HS-only students (9-12)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.WD_SEC_END}"
    dat(ws, r, 2, CALC_FILL).number_format = '0'; r += 1

    R.WD_K8_HS_TUITION = r
    put(ws, r, 1, "HS tuition cost (9-12 only)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.WD_K8_HS_STUDENTS}*B{R.WD_TUIT_SEC}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.WD_K8_TUITION_SAVED = r
    put(ws, r, 1, "MS tuition saved (6-8 no longer tuitioned)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.WD_K8_STUDENTS}*B{R.WD_TUIT_ELEM}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    r = note(ws, r, "K-8 tradeoffs: More local control through 8th grade; limited MS course variety")
    r = note(ws, r, "(no French, Latin, band at scale); small cohort for MS sports/activities;")
    r = note(ws, r, "PES facility has capacity but may need science lab and reconfiguration.")
    r += 1


def build_i_equity(wb):
    """Input data for per-town equity analysis: enrollment by town, assessments, cost-sharing."""
    ws = wb.create_sheet("I-Equity")
    ws.sheet_properties.tabColor = "FFC000"
    col_widths(ws, [48, 18, 18, 18, 18, 40])

    r = ttl(ws, 1, "INPUT: Per-Town Equity & Cost-Sharing Data")
    r = source_block(ws, r + 1, [
        "Data for analyzing fairness of RSU 5 cost allocation across Pownal, Durham, and Freeport.",
        "Combines enrollment attribution, assessment history, EPS allocation, and cost-sharing formula.",
        "NOTE: FMS and FHS serve multiple towns; per-town estimates use cohort-based attribution.",
        "Sources cited per item. See Sources sheet for full citations.",
    ])
    r += 1

    # ── Section 1: Per-Town Enrollment Attribution ──
    r = sec(ws, r, "1. ENROLLMENT BY TOWN (FY27 Projected) [Fn 1]")
    r = source_block(ws, r, [
        "SOURCE: FY27 Budget Handbook enrollment table [Fn 1], PES grade data [Fn 1].",
        "PES and DCS are single-town schools. MSS and MLS serve Freeport only.",
        "FMS (6-8) serves Freeport + Pownal. FHS (9-12) serves all three towns.",
        "Secondary attribution is ESTIMATED from cohort sizes at each town's elementary school.",
    ])
    for i, h in enumerate(["School / Segment", "Pownal", "Durham", "Freeport", "Total", "Basis"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 6); r += 1

    R.EQ_ENROLL_START = r
    enroll_rows = [
        ("PES (PreK-5)", 105, 0, 0, 105, "PES enrollment, 100% Pownal [Fn 1]"),
        ("DCS (PreK-8)", 0, 467, 0, 467, "DCS enrollment, 100% Durham [Fn 1]"),
        ("MSS (PreK-2)", 0, 0, 274, 274, "MSS enrollment, 100% Freeport [Fn 1]"),
        ("MLS (3-5)", 0, 0, 264, 264, "MLS enrollment, 100% Freeport [Fn 1]"),
        ("FMS (6-8) - Pownal residents", 45, 0, 0, 45, "Est: PES avg 15/grade × 3 grades [Fn 1]"),
        ("FMS (6-8) - Freeport residents", 0, 0, 261, 261, "306 total FMS minus 45 Pownal est."),
        ("FHS (9-12) - Pownal residents", 55, 0, 0, 55, "Est: PES avg cohort × 4 grades [Fn 1]"),
        ("FHS (9-12) - Durham residents", 0, 225, 0, 225, "Est: DCS avg ~56/grade × 4 grades [Fn 1]"),
        ("FHS (9-12) - Freeport residents", 0, 0, 274, 274, "554 total FHS minus 55 POW minus 225 DUR"),
    ]
    for label, p, d, f_, tot, basis in enroll_rows:
        put(ws, r, 1, label)
        put(ws, r, 2, p, '0')
        put(ws, r, 3, d, '0')
        put(ws, r, 4, f_, '0')
        put(ws, r, 5, tot, '0')
        put(ws, r, 6, basis, fill=SOURCE_FILL, font=SOURCE)
        r += 1
    R.EQ_ENROLL_END = r - 1
    r += 1

    R.EQ_TOT_POW = r
    put(ws, r, 1, "TOTAL STUDENTS - Pownal", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.EQ_ENROLL_START}:B{R.EQ_ENROLL_END})"
    dat(ws, r, 2, RESULT_FILL).number_format = '0'
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.EQ_TOT_DUR = r
    put(ws, r, 1, "TOTAL STUDENTS - Durham", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 3).value = f"=SUM(C{R.EQ_ENROLL_START}:C{R.EQ_ENROLL_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = '0'
    ws.cell(r, 3).font = RESULT_FONT; r += 1

    R.EQ_TOT_FRE = r
    put(ws, r, 1, "TOTAL STUDENTS - Freeport", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 4).value = f"=SUM(D{R.EQ_ENROLL_START}:D{R.EQ_ENROLL_END})"
    dat(ws, r, 4, RESULT_FILL).number_format = '0'
    ws.cell(r, 4).font = RESULT_FONT; r += 1

    R.EQ_TOT_ALL = r
    put(ws, r, 1, "TOTAL STUDENTS - All Towns", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 5).value = f"=SUM(E{R.EQ_ENROLL_START}:E{R.EQ_ENROLL_END})"
    dat(ws, r, 5, RESULT_FILL).number_format = '0'
    ws.cell(r, 5).font = RESULT_FONT; r += 2

    # ── Section 2: FY27 Revenue / Assessment by Town ──
    r = sec(ws, r, "2. FY27 PROPOSED ASSESSMENT BY TOWN [Fn 3]")
    r = source_block(ws, r, [
        "SOURCE: FY27 Budget Handbook, pp.9-10, Budget Impact Summary (02/11/2026) [Fn 3]",
        "RLC = Required Local Contribution (set by Maine DOE via EPS formula).",
        "ALM = Additional Local Monies (allocated by RSU 5 cost-sharing formula).",
        "State Aid = EPS state subsidy (flows directly to RSU, not through town).",
    ])
    for i, h in enumerate(["Component", "Pownal", "Durham", "Freeport", "Total"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.EQ_FY27_START = r
    fy27_items = [
        ("Required Local Contribution (RLC)", 2257247, 3910950, 14173184),
        ("Additional Local Monies (ALM)", 2190978, 3724663, 11473075),
        ("State Aid", 567180, 5991564, 1463889),
    ]
    for label, p, d, f_ in fy27_items:
        put(ws, r, 1, label)
        put(ws, r, 2, p, USD); put(ws, r, 3, d, USD); put(ws, r, 4, f_, USD)
        ws.cell(r, 5).value = f"=B{r}+C{r}+D{r}"
        dat(ws, r, 5).number_format = USD
        r += 1
    R.EQ_RLC_R = R.EQ_FY27_START
    R.EQ_ALM_R = R.EQ_FY27_START + 1
    R.EQ_AID_R = R.EQ_FY27_START + 2
    r += 1

    R.EQ_LOCAL = r
    put(ws, r, 1, "LOCAL TAX ASSESSMENT (RLC + ALM)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 6):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.EQ_RLC_R}+{cl}{R.EQ_ALM_R}"
        dat(ws, r, c, RESULT_FILL).number_format = USD
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    R.EQ_TOTAL_FUNDING = r
    put(ws, r, 1, "TOTAL FUNDING (Local + State Aid)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 6):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.EQ_LOCAL}+{cl}{R.EQ_AID_R}"
        dat(ws, r, c, RESULT_FILL).number_format = USD
        ws.cell(r, c).font = RESULT_FONT
    r += 2

    # ── Section 3: EPS Allocation (from Warrant Article F) ──
    r = sec(ws, r, "3. EPS ALLOCATION BY TOWN (FY25-26) [Fn 8]")
    r = source_block(ws, r, [
        "SOURCE: Maine DOE Warrant Article F, FY25-26 Education Subsidy (09/09/2025) [Fn 8]",
        "EPS allocation = the state's calculated cost of educating each town's students.",
        "Local contribution = minimum local share required to receive state subsidy.",
        "State share = EPS allocation minus local contribution.",
    ])
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport", "RSU Total"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.EQ_EPS_START = r
    eps_data = [
        ("EPS Total Allocation", 2682829, 9123077, 15033673),
        ("EPS Local Contribution", 2179327, 3655120, 13799841),
        ("EPS State Share", 503502, 5467957, 1233833),
        ("State Share %", None, None, None),
    ]
    for label, p, d, f_ in eps_data:
        put(ws, r, 1, label)
        if p is not None:
            put(ws, r, 2, p, USD); put(ws, r, 3, d, USD); put(ws, r, 4, f_, USD)
            ws.cell(r, 5).value = f"=B{r}+C{r}+D{r}"
            dat(ws, r, 5).number_format = USD
        else:
            for c in range(2, 6):
                cl = get_column_letter(c)
                ws.cell(r, c).value = f"={cl}{r-1}/{cl}{r-3}"
                dat(ws, r, c).number_format = PCT2
        r += 1
    R.EQ_EPS_ALLOC = R.EQ_EPS_START
    R.EQ_EPS_LOCAL = R.EQ_EPS_START + 1
    R.EQ_EPS_STATE = R.EQ_EPS_START + 2
    R.EQ_EPS_PCT = R.EQ_EPS_START + 3
    r += 1
    r = note(ws, r, "Durham receives 59.9% state share vs. Freeport 8.2% and Pownal 18.8%.")
    r = note(ws, r, "This is because Durham's property values are LOW relative to its student count,")
    r = note(ws, r, "so the state subsidizes more. Freeport is 'property-rich' so gets minimal state aid.")
    r += 1

    # ── Section 4: State Valuations ──
    r = sec(ws, r, "4. STATE VALUATIONS (FY27 Budget Handbook) [Fn 3]")
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport", "Total"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.EQ_SVAL_R = r
    put(ws, r, 1, "State Valuation")
    put(ws, r, 2, 399866667, USD); put(ws, r, 3, 692816667, USD); put(ws, r, 4, 2510750000, USD)
    ws.cell(r, 5).value = f"=B{r}+C{r}+D{r}"
    dat(ws, r, 5).number_format = USD; r += 1

    R.EQ_SVAL_PCT = r
    put(ws, r, 1, "% of Total State Valuation")
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.EQ_SVAL_R}/E{R.EQ_SVAL_R}"
        dat(ws, r, c).number_format = PCT2
    r += 2

    # ── Section 5: Cost-Sharing Formula ──
    r = sec(ws, r, "5. RSU 5 COST-SHARING FORMULA [Fn 41][Fn 42]")
    r = source_block(ws, r, [
        "SOURCE: Portland Press Herald, 10/29/2019: 'RSU 5 looks at new cost-sharing formula.' [Fn 41]",
        "RSU 5 Finance Committee proposed: ALM = 85% state valuation + 15% enrollment. [Fn 41]",
        "Under Title 20-A §1481-A(3), RSU reorganization plans may specify alternate formulas. [Fn 42]",
        "RLC is always allocated by state valuation per the EPS formula (Maine DOE). [Fn 8]",
        "NOTE: The 85/15 formula was proposed in 2019. Actual adopted formula should be confirmed.",
    ])
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.EQ_FORMULA_START = r
    put(ws, r, 1, "Enrollment share (students/total)")
    for c in range(2, 5):
        tc = get_column_letter(c)
        ws.cell(r, c).value = f"={tc}{R.EQ_TOT_POW + (c-2)}/{{'I-Equity'}}!E{R.EQ_TOT_ALL}" if c == 2 else \
            f"={tc}{R.EQ_TOT_DUR}/{{'I-Equity'}}!E{R.EQ_TOT_ALL}" if c == 3 else \
            f"={tc}{R.EQ_TOT_FRE}/{{'I-Equity'}}!E{R.EQ_TOT_ALL}"
    # Simplified: use direct cell refs
    ws.cell(r, 2).value = f"=B{R.EQ_TOT_POW}/E{R.EQ_TOT_ALL}"
    ws.cell(r, 3).value = f"=C{R.EQ_TOT_DUR}/E{R.EQ_TOT_ALL}"
    ws.cell(r, 4).value = f"=D{R.EQ_TOT_FRE}/E{R.EQ_TOT_ALL}"
    for c in range(2, 5):
        dat(ws, r, c).number_format = PCT2
    r += 1

    put(ws, r, 1, "State valuation share")
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.EQ_SVAL_R}/E{R.EQ_SVAL_R}"
        dat(ws, r, c).number_format = PCT2
    R.EQ_VAL_SHARE = r; r += 1

    R.EQ_WEIGHTED = r
    put(ws, r, 1, "Weighted share (85% val + 15% enroll)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"=0.85*{cl}{R.EQ_VAL_SHARE}+0.15*{cl}{R.EQ_FORMULA_START}"
        dat(ws, r, c, RESULT_FILL).number_format = PCT2
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "ACTUAL ALM share (from budget data)")
    alm_tot = f"(B{R.EQ_ALM_R}+C{R.EQ_ALM_R}+D{R.EQ_ALM_R})"
    for c, col in [(2, 'B'), (3, 'C'), (4, 'D')]:
        ws.cell(r, c).value = f"={col}{R.EQ_ALM_R}/{alm_tot}"
        dat(ws, r, c).number_format = PCT2
    R.EQ_ACTUAL_ALM = r; r += 2

    r = note(ws, r, "If weighted share ≈ actual ALM share, the 85/15 formula is confirmed.")
    r = note(ws, r, "Any discrepancy may indicate a different formula or adjustment factor.")
    r += 1

    # ── Section 6: Historical Assessment Data ──
    r = sec(ws, r, "6. HISTORICAL POWNAL EDUCATION ASSESSMENT [Fn 9][Fn 4][Fn 43]")
    r = source_block(ws, r, [
        "SOURCE: Pownal 2025 Annual Report, FY24 audit (year ending June 30, 2024) [Fn 43].",
        "FY26 corroborated: (1) 58.4% RSU × $6.0M levy [Fn 4][Fn 12], (2) consistent with",
        "  RSU Q&A total FY26 taxation of $35,600,800 at ~9.85% Pownal share [Fn 44],",
        "  (3) FY26 handbook pp.8-9 contains per-town data (PDF image, not text-extractable).",
        "FY27 from Budget Handbook pp.9-10 [Fn 3]. Direct figure from RSU 5 documents.",
    ])
    for i, h in enumerate(["Fiscal Year", "Pownal Assessment", "RSU Budget", "Pownal % of Budget", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.EQ_HIST_START = r
    hist_rows = [
        ("FY24 (2023-2024)", 3427607, 39080569, "Audited [Fn 43]"),
        ("FY25 (2024-2025)", None, 41612460, "Budget only [Fn 9]"),
        ("FY26 (2025-2026)", 3506157, 44455929, "Corroborated [Fn 4][Fn 12][Fn 44]"),
        ("FY27 (2026-2027)", 4448225, 47357441, "Budget Handbook [Fn 3]"),
    ]
    for fy, assess, budget, src in hist_rows:
        put(ws, r, 1, fy)
        if assess is not None:
            put(ws, r, 2, assess, USD)
        put(ws, r, 3, budget, USD)
        if assess is not None:
            ws.cell(r, 4).value = f"=B{r}/C{r}"
            dat(ws, r, 4).number_format = PCT2
        put(ws, r, 5, src, fill=SOURCE_FILL, font=SOURCE)
        r += 1
    R.EQ_HIST_FY24 = R.EQ_HIST_START
    R.EQ_HIST_FY26 = R.EQ_HIST_START + 2
    R.EQ_HIST_FY27 = R.EQ_HIST_START + 3
    r += 1

    put(ws, r, 1, "FY26 Total RSU Taxation (all towns)")
    put(ws, r, 2, 35600800, USD)
    put(ws, r, 5, "Q&A Budget Doc [Fn 44]", fill=SOURCE_FILL, font=SOURCE)
    R.EQ_FY26_TOT_TAX = r; r += 1

    put(ws, r, 1, "Pownal % of FY26 total taxation")
    ws.cell(r, 2).value = f"=B{R.EQ_HIST_FY26}/B{R.EQ_FY26_TOT_TAX}"
    dat(ws, r, 2).number_format = PCT2
    put(ws, r, 5, "Crosscheck: $3.51M / $35.6M = 9.85%", fill=SOURCE_FILL, font=SOURCE)
    r += 1

    R.EQ_3YR_INCR = r
    put(ws, r, 1, "3-YEAR INCREASE (FY24→FY27)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.EQ_HIST_FY27}-B{R.EQ_HIST_FY24}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 4).value = f"=B{r}/B{R.EQ_HIST_FY24}"
    dat(ws, r, 4, RESULT_FILL).number_format = PCT2
    ws.cell(r, 4).font = RESULT_FONT
    put(ws, r, 5, "Pownal assessment grew 29.8% over 3 years", fill=SOURCE_FILL, font=SOURCE)
    r += 1

    R.EQ_1YR_INCR = r
    put(ws, r, 1, "1-YEAR INCREASE (FY26→FY27)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.EQ_HIST_FY27}-B{R.EQ_HIST_FY26}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 4).value = f"=B{r}/B{R.EQ_HIST_FY26}"
    dat(ws, r, 4, RESULT_FILL).number_format = PCT2
    ws.cell(r, 4).font = RESULT_FONT
    put(ws, r, 5, "FY26 corroborated via 3 methods [Fn 4][Fn 12][Fn 44]", fill=SOURCE_FILL, font=SOURCE)
    r += 1

    r = note(ws, r, "FY26→FY27 jump driven by Pownal's revaluation feeding into the 85/15 ALM formula.")
    r = note(ws, r, "Pownal's assessed values rose ~20% via revaluation; state valuation followed,")
    r = note(ws, r, "shifting more ALM to Pownal under the valuation-weighted cost-sharing formula.")
    r += 1

    # ── Section 7: Property Tax Context ──
    r = sec(ws, r, "7. PROPERTY TAX CONTEXT [Fn 4][Fn 5][Fn 7]")
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.EQ_TAX_START = r
    put(ws, r, 1, "Total mil rate (FY26)")
    put(ws, r, 2, 15.300, '0.000'); put(ws, r, 3, 33.580, '0.000'); put(ws, r, 4, 13.850, '0.000')
    r += 1
    put(ws, r, 1, "Assessment ratio")
    put(ws, r, 2, 1.00, PCT); put(ws, r, 3, 0.53, PCT); put(ws, r, 4, 1.00, PCT)
    r += 1
    put(ws, r, 1, "Taxable valuation")
    put(ws, r, 2, 392398240, USD); put(ws, r, 3, 389646650, USD); put(ws, r, 4, 2613679115, USD)
    R.EQ_TAXABLE_R = r; r += 1
    put(ws, r, 1, "RSU share of tax bill (FY26)")
    put(ws, r, 2, 0.584, PCT); put(ws, r, 3, "N/A"); put(ws, r, 4, 0.716, PCT)
    r += 1
    put(ws, r, 1, "Total tax levy")
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.EQ_TAX_START}*{cl}{R.EQ_TAXABLE_R}/1000"
        dat(ws, r, c).number_format = USD
    R.EQ_TAX_LEVY = r; r += 1
    put(ws, r, 1, "Housing units (est.)")
    put(ws, r, 2, 636, '0'); put(ws, r, 3, 1650, '0'); put(ws, r, 4, 4100, '0')
    R.EQ_HOUSING = r; r += 2

    # ── Section 8: RLC vs ALM Explanation ──
    r = sec(ws, r, "8. RLC vs. ALM: WHAT THEY ARE [Fn 8][Fn 42]")
    explanations = [
        "RLC (Required Local Contribution):",
        "  - Set by Maine DOE under the EPS formula (Title 20-A §15688) [Fn 8]",
        "  - Based on each town's STATE VALUATION and the statewide mill rate expectation",
        "  - Represents the MINIMUM local investment required to receive state subsidies",
        "  - The RSU has NO discretion over RLC amounts; they are state-mandated",
        "",
        "ALM (Additional Local Monies):",
        "  - Everything the RSU spends ABOVE the EPS-required minimum",
        "  - Allocated by RSU 5's cost-sharing formula from the reorganization plan [Fn 42]",
        "  - RSU 5 proposed 85% state valuation / 15% enrollment (2019 review) [Fn 41]",
        "  - This means high-valuation towns (Freeport, Pownal) pay disproportionately MORE",
        "",
        "What shows on property tax bills:",
        "  - The RSU % (e.g., 58.4% for Pownal) = (RLC + ALM) / total tax levy",
        "  - State Aid does NOT flow through the town; it goes directly to the RSU",
        "  - The % changes annually based on RSU assessment, town budget, and county tax",
    ]
    for line in explanations:
        if line:
            r = note(ws, r, line)
        else:
            r += 1


def build_c_equity(wb):
    """Calculations: per-student costs, subsidy flows, cost-sharing decomposition."""
    ws = wb.create_sheet("C-Equity")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [48, 20, 20, 20, 20])
    IEQ = SN['ieq']; IB = SN['ib']

    r = ttl(ws, 1, "CALC: Per-Town Equity Analysis")
    r = note(ws, r + 1, "All values are formula-driven from I-Equity inputs. Shows who pays vs. who consumes.")
    r += 1

    # ── Section 1: Per-Student Cost ──
    r = sec(ws, r, "1. PER-STUDENT COST BY TOWN (FY27)")
    for i, h in enumerate(["Metric", "Pownal", "Durham", "Freeport", "RSU Average"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.CE_START = r
    put(ws, r, 1, "Students (attributed)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IEQ}!B{R.EQ_TOT_POW}"
    ws.cell(r, 3).value = f"={IEQ}!C{R.EQ_TOT_DUR}"
    ws.cell(r, 4).value = f"={IEQ}!D{R.EQ_TOT_FRE}"
    ws.cell(r, 5).value = f"={IEQ}!E{R.EQ_TOT_ALL}"
    for c in range(2, 6):
        dat(ws, r, c, CALC_FILL).number_format = '0'
    R.CE_STUDENTS = r; r += 1

    put(ws, r, 1, "Local tax assessment (RLC + ALM)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IEQ}!B{R.EQ_LOCAL}"
    ws.cell(r, 3).value = f"={IEQ}!C{R.EQ_LOCAL}"
    ws.cell(r, 4).value = f"={IEQ}!D{R.EQ_LOCAL}"
    ws.cell(r, 5).value = f"={IEQ}!E{R.EQ_LOCAL}"
    for c in range(2, 6):
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_LOCAL = r; r += 1

    put(ws, r, 1, "Total funding (local + state aid)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IEQ}!B{R.EQ_TOTAL_FUNDING}"
    ws.cell(r, 3).value = f"={IEQ}!C{R.EQ_TOTAL_FUNDING}"
    ws.cell(r, 4).value = f"={IEQ}!D{R.EQ_TOTAL_FUNDING}"
    ws.cell(r, 5).value = f"={IEQ}!E{R.EQ_TOTAL_FUNDING}"
    for c in range(2, 6):
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_TOTAL = r; r += 1

    R.CE_LOCAL_PP = r
    put(ws, r, 1, "LOCAL TAX PER STUDENT", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 6):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_LOCAL}/{cl}{R.CE_STUDENTS}"
        dat(ws, r, c, RESULT_FILL).number_format = USD
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    R.CE_TOTAL_PP = r
    put(ws, r, 1, "TOTAL FUNDING PER STUDENT", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 6):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_TOTAL}/{cl}{R.CE_STUDENTS}"
        dat(ws, r, c, RESULT_FILL).number_format = USD
        ws.cell(r, c).font = RESULT_FONT
    r += 2

    CC = SN['cc']

    # ── Section 2A: Actual Consumption Net Position ──
    r = sec(ws, r, "2A. NET POSITION: ACTUAL CONSUMPTION (from C-Consumption)")
    r = note(ws, r, "Uses actual school-level budgets attributed to each town's students.")
    r = note(ws, r, "PES costs are 100% Pownal. FMS/FHS split by enrollment. System pool by student count.")
    r = note(ws, r, "This is the objective cost picture -- it includes PES's small-school premium.")
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.CE_ACTUAL_CONS = r
    put(ws, r, 1, "Actual budget consumed (C-Consumption)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CC}!D{R.C3_POW_TOTAL}"
    ws.cell(r, 3).value = f"={CC}!D{R.C3_DUR_TOTAL}"
    ws.cell(r, 4).value = f"={CC}!D{R.C3_FRE_TOTAL}"
    for c in range(2, 5):
        dat(ws, r, c, CALC_FILL).number_format = USD
    r += 1

    put(ws, r, 1, "Total contribution (local + state + shared)", fill=CALC_FILL)
    CR = SN['cr']
    ws.cell(r, 2).value = f"={CR}!B{R.C4_TOTAL_REV}"
    ws.cell(r, 3).value = f"={CR}!C{R.C4_TOTAL_REV}"
    ws.cell(r, 4).value = f"={CR}!D{R.C4_TOTAL_REV}"
    for c in range(2, 5):
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_ACTUAL_REV = r; r += 1

    R.CE_ACTUAL_NET = r
    put(ws, r, 1, "NET POSITION (revenue - consumption)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_ACTUAL_REV}-{cl}{R.CE_ACTUAL_CONS}"
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "Actual cost per student", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        stu_col = cl
        ws.cell(r, c).value = f"={cl}{R.CE_ACTUAL_CONS}/{cl}{R.CE_STUDENTS}"
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_ACTUAL_PP = r; r += 2

    r = note(ws, r, "Pownal's deficit is REAL but modest: PES costs more per student due to fixed-cost dilution.")
    r = note(ws, r, "Durham's deficit is larger. Freeport is the only net contributor.")
    r += 1

    # ── Section 2B: Equalized Cost Analysis ──
    r = sec(ws, r, "2B. EQUALIZED ANALYSIS: What if every student cost the same?")
    r = note(ws, r, "Uses RSU-wide average per-student cost. Isolates the TAX BURDEN question from")
    r = note(ws, r, "the SCHOOL EFFICIENCY question. Shows whether each town pays its 'fair share'")
    r = note(ws, r, "assuming equal cost per student. Contrast with 2A above for full picture.")
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    put(ws, r, 1, "RSU 5 average cost per student", fill=CALC_FILL)
    for c in range(2, 5):
        ws.cell(r, c).value = f"=E{R.CE_TOTAL_PP}"
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_AVG = r; r += 1

    put(ws, r, 1, "Equalized consumption (students x avg cost)", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_STUDENTS}*{cl}{R.CE_AVG}"
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_CONSUMED = r; r += 1

    put(ws, r, 1, "Total contribution (local + state aid)", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_TOTAL}"
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_CONTRIB = r; r += 1

    R.CE_SUBSIDY = r
    put(ws, r, 1, "EQUALIZED NET (consumed - contributed)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_CONSUMED}-{cl}{R.CE_CONTRIB}"
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    R.CE_SUBSIDY_HH = r
    put(ws, r, 1, "Equalized net per household", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_SUBSIDY}/{IEQ}!{cl}{R.EQ_HOUSING}"
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 2

    r = note(ws, r, "COMPARE 2A vs 2B: Pownal's actual deficit (2A) is $267K. But if every student cost")
    r = note(ws, r, "the same, Pownal would be a net contributor (2B). The difference = PES's cost premium.")
    r = note(ws, r, "Both are real. Presenting only one without the other is misleading.")
    r += 1

    # ── Section 3: Cost-Sharing Formula Impact ──
    r = sec(ws, r, "3. COST-SHARING FORMULA IMPACT")
    r = note(ws, r, "Shows how the 85/15 formula distributes ALM vs. what a pure per-student split would look like.")
    for i, h in enumerate(["", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    put(ws, r, 1, "Actual ALM paid", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={IEQ}!{cl}{R.EQ_ALM_R}"
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_ALM_ACTUAL = r; r += 1

    total_alm = f"(B{R.CE_ALM_ACTUAL}+C{R.CE_ALM_ACTUAL}+D{R.CE_ALM_ACTUAL})"
    put(ws, r, 1, "ALM if split by enrollment only", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = (
            f"={total_alm}*{cl}{R.CE_STUDENTS}/"
            f"(B{R.CE_STUDENTS}+C{R.CE_STUDENTS}+D{R.CE_STUDENTS})"
        )
        dat(ws, r, c, CALC_FILL).number_format = USD
    R.CE_ALM_ENROLL = r; r += 1

    R.CE_ALM_DIFF = r
    put(ws, r, 1, "FORMULA PENALTY (+) or BENEFIT (-)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_ALM_ACTUAL}-{cl}{R.CE_ALM_ENROLL}"
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    R.CE_ALM_DIFF_HH = r
    put(ws, r, 1, "Formula penalty per household", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_ALM_DIFF}/{IEQ}!{cl}{R.EQ_HOUSING}"
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 2

    # ── Section 4: Assessment Trend & Tax Impact ──
    r = sec(ws, r, "4. POWNAL ASSESSMENT TREND & TAX IMPACT")
    for i, h in enumerate(["", "Value", "Source"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    put(ws, r, 1, "FY24 Pownal assessment (audited)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IEQ}!B{R.EQ_HIST_FY24}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "Pownal Annual Report audit [Fn 43]", fill=SOURCE_FILL, font=SOURCE)
    R.CE_FY24 = r; r += 1

    put(ws, r, 1, "FY27 Pownal assessment (proposed)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IEQ}!B{R.EQ_HIST_FY27}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "Budget Handbook [Fn 3]", fill=SOURCE_FILL, font=SOURCE)
    R.CE_FY27 = r; r += 1

    R.CE_3YR_DOLLAR = r
    put(ws, r, 1, "3-year dollar increase (FY24→FY27)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CE_FY27}-B{R.CE_FY24}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CE_3YR_PCT = r
    put(ws, r, 1, "3-year percentage increase", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CE_3YR_DOLLAR}/B{R.CE_FY24}"
    dat(ws, r, 2, RESULT_FILL).number_format = PCT2
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "RSU budget growth same period", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=({IEQ}!C{R.EQ_HIST_FY27}-{IEQ}!C{R.EQ_HIST_FY24})/{IEQ}!C{R.EQ_HIST_FY24}"
    dat(ws, r, 2, CALC_FILL).number_format = PCT2
    put(ws, r, 3, "Pownal's share grew faster than budget", fill=SOURCE_FILL, font=SOURCE)
    r += 2

    put(ws, r, 1, "FY27 assessment per household (636 homes)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CE_FY27}/{IEQ}!B{R.EQ_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "FY24 assessment per household", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CE_FY24}/{IEQ}!B{R.EQ_HOUSING}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Per-household increase (3 years)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=(B{R.CE_FY27}-B{R.CE_FY24})/{IEQ}!B{R.EQ_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ── Section 5: Comparative Fairness Metrics ──
    r = sec(ws, r, "5. COMPARATIVE FAIRNESS METRICS")
    for i, h in enumerate(["Metric", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    put(ws, r, 1, "% of RSU budget paid locally", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_LOCAL}/{IEQ}!C{R.EQ_HIST_FY27}"
        dat(ws, r, c, CALC_FILL).number_format = PCT2
    r += 1

    put(ws, r, 1, "% of RSU students", fill=CALC_FILL)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"={cl}{R.CE_STUDENTS}/E{R.CE_STUDENTS}"
        dat(ws, r, c, CALC_FILL).number_format = PCT2
    R.CE_STU_PCT = r; r += 1

    R.CE_DISPARITY = r
    put(ws, r, 1, "DISPARITY: pays ÷ gets (1.0 = fair)", fill=RESULT_FILL, font=BOLD)
    for c in range(2, 5):
        cl = get_column_letter(c)
        ws.cell(r, c).value = (
            f"=({cl}{R.CE_LOCAL}/{IEQ}!C{R.EQ_HIST_FY27})"
            f"/({cl}{R.CE_STUDENTS}/E{R.CE_STUDENTS})"
        )
        dat(ws, r, c, RESULT_FILL).number_format = '0.00'
        ws.cell(r, c).font = RESULT_FONT
    r += 1
    r = note(ws, r, ">1.0 = pays more than fair share. <1.0 = pays less than fair share.")
    r = note(ws, r, "Pownal's disparity ratio shows it pays significantly more per student than it should.")


def build_c_independence(wb):
    """Financial modeling for Pownal independence scenarios."""
    ws = wb.create_sheet("C-Independence")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [52, 22, 22, 22, 44])
    IW = SN['iw']; IT = SN['it']; CFP = SN['cfp']; CRM = SN['crm']

    r = ttl(ws, 1, "CALC: Pownal Independence Financial Analysis")
    r = note(ws, r + 1, "Four scenarios modeled: §1512 veto cost, full withdrawal, withdrawal+AOS, legal challenge.")
    r = note(ws, r, "All values are formula-driven from I-Withdrawal inputs.")
    r += 1

    # ═══ Scenario A: §1512 Municipal Veto Cost ═══
    r = sec(ws, r, "SCENARIO A: §1512 MUNICIPAL VETO (if applicable)")
    r = note(ws, r, "Under §1512, if Pownal votes to keep PES open against Board wishes,")
    r = note(ws, r, "Pownal is liable for 'the amount that would be saved if the school were closed.'")
    r = note(ws, r, "IMPORTANT: §1512 likely does NOT apply to Option 2 (reconfiguration, not closure).")
    r = note(ws, r, "This scenario models the cost ceiling IF §1512 were successfully invoked.")
    r += 1

    R.CI_VETO_START = r
    put(ws, r, 1, "TRUE NET SAVINGS from closing PES (from C-FY28Paths)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.FP_TRUE_NET}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.CI_VETO_NET = r
    put(ws, r, 1, "§1512 LIABILITY = savings RSU would gain", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f'=IF(B{R.CI_VETO_START}>0,B{R.CI_VETO_START},0)'
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_VETO_PH = r
    put(ws, r, 1, "Per-household cost (636 homes)", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_VETO_NET}/{IW}!B{R.WD_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 1

    r = note(ws, r, "The TRUE NET from closing PES is approximately -$161K (a COST, not savings).")
    r = note(ws, r, "If net savings are negative, there is NO amount for Pownal to reimburse under §1512.")
    r = note(ws, r, "Revisited annually: Board must re-vote each year to attempt closure.")
    r += 1

    # ═══ Scenario B: Full Withdrawal - Operate K-5, Tuition 6-12 ═══
    r = sec(ws, r, "SCENARIO B: FULL WITHDRAWAL (K-5 independent, tuition 6-12)")
    r = note(ws, r, "Pownal withdraws from RSU 5 under Title 20-A §1466. Operates PES K-5")
    r = note(ws, r, "independently. Tuitions secondary students (6-12) to Freeport or elsewhere.")
    r += 1

    # ── Ongoing Annual Costs (Year 2+) ──
    r = note(ws, r, "ONGOING COSTS (Year 2+): Low and High estimates from I-Withdrawal Section 3.")
    r += 1

    R.CI_WD_START = r
    put(ws, r, 1, "PES school-level operating cost", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_PES_SUBTOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"={IW}!B{R.WD_PES_SUBTOTAL}"
    dat(ws, r, 3, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Additional costs (low estimate)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_IND_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Additional costs (high estimate)", fill=CALC_FILL)
    ws.cell(r, 3).value = f"={IW}!C{R.WD_IND_TOTAL}"
    dat(ws, r, 3, CALC_FILL).number_format = USD; r += 1

    R.CI_WD_GROSS = r
    put(ws, r, 1, "GROSS INDEPENDENT BUDGET", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_WD_START}+B{R.CI_WD_START+1}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).value = f"=C{R.CI_WD_START}+C{R.CI_WD_START+2}"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 3).font = RESULT_FONT; r += 1

    put(ws, r, 1, "SpEd risk reserve (annual set-aside)", fill=CALC_FILL)
    put(ws, r, 2, 50000, USD); put(ws, r, 3, 50000, USD); r += 1
    R.CI_SPED_RESERVE = r - 1; r += 1

    put(ws, r, 1, "Less: State EPS share (current)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!B{R.WD_EPS_STATE}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"=-{IW}!B{R.WD_EPS_STATE}"
    dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1
    R.CI_EPS_ADJ = r - 1

    put(ws, r, 1, "Less: Isolated small school adjustment", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!B{R.WD_EPS_SMALL}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"=-{IW}!B{R.WD_EPS_SMALL}"
    dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    R.CI_WD_LOCAL = r
    put(ws, r, 1, "NET LOCAL COST (Pownal taxpayers, ongoing)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_WD_GROSS}+B{R.CI_SPED_RESERVE}+B{R.CI_EPS_ADJ}+B{r-1}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).value = f"=C{R.CI_WD_GROSS}+C{R.CI_SPED_RESERVE}+C{R.CI_EPS_ADJ}+C{r-1}"
    dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 3).font = RESULT_FONT; r += 2

    put(ws, r, 1, "Current Pownal LOCAL tax (RLC + ALM, for comparison)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_POW_LOCAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"={IW}!B{R.WD_POW_LOCAL}"
    dat(ws, r, 3, CALC_FILL).number_format = USD
    R.CI_CUR_ASSESS = r; r += 1

    R.CI_WD_DELTA = r
    put(ws, r, 1, "CHANGE vs. CURRENT (low cost = best case)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_WD_LOCAL}-B{R.CI_CUR_ASSESS}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_WD_DELTA_HIGH = r
    put(ws, r, 1, "CHANGE vs. CURRENT (high cost = worst case)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=C{R.CI_WD_LOCAL}-C{R.CI_CUR_ASSESS}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_WD_PCT = r
    put(ws, r, 1, "Percentage change (low/high)", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_WD_DELTA}/B{R.CI_CUR_ASSESS}"
    dat(ws, r, 2, RESULT_FILL).number_format = PCT
    ws.cell(r, 3).value = f"=B{R.CI_WD_DELTA_HIGH}/C{R.CI_CUR_ASSESS}"
    dat(ws, r, 3, RESULT_FILL).number_format = PCT; r += 2

    put(ws, r, 1, "New estimated mil rate (education only, low)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_WD_LOCAL}/{IW}!B{R.WD_VALUATION}*1000"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.000'
    ws.cell(r, 2).font = RESULT_FONT
    R.CI_WD_MIL = r; r += 1

    put(ws, r, 1, "Current RSU education mil rate", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_RSU_MIL}"
    dat(ws, r, 2, CALC_FILL).number_format = '0.000'; r += 1

    R.CI_WD_MIL_DELTA = r
    put(ws, r, 1, "Mil rate change", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_WD_MIL}-B{r-1}"
    dat(ws, r, 2, RESULT_FILL).number_format = '+0.000;-0.000;0.000'; r += 1

    R.CI_WD_PH = r
    put(ws, r, 1, "Annual cost change per household (636 homes, low)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_WD_DELTA}/{IW}!B{R.WD_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_WD_PH_HIGH = r
    put(ws, r, 1, "Annual cost change per household (636 homes, high)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_WD_DELTA_HIGH}/{IW}!B{R.WD_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ── Year 1 Adjustments ──
    r = note(ws, r, "YEAR 1 ADJUSTMENTS: One-time costs and credits that apply only in the first year.")
    r += 1

    R.CI_Y1_START = r
    put(ws, r, 1, "Year 1 uncapped tuition penalty (55 HS students)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_SEC_912}*{IW}!B{R.WD_TUIT_PENALTY}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Startup costs (low)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_STARTUP_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Startup costs (high)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!C{R.WD_STARTUP_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Withdrawal process costs", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_PROC_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Less: Fund balance return (low)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!B{R.WD_FUND_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "Less: Fund balance return (high)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!C{R.WD_FUND_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.CI_Y1_NET = r
    put(ws, r, 1, "NET YEAR 1 ADJUSTMENT (low-high range)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_Y1_START}+B{R.CI_Y1_START+1}+B{R.CI_Y1_START+3}+B{R.CI_Y1_START+4}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).value = f"=B{R.CI_Y1_START}+B{R.CI_Y1_START+2}+B{R.CI_Y1_START+3}+B{R.CI_Y1_START+5}"
    dat(ws, r, 3, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 3).font = RESULT_FONT; r += 2

    # ═══ Scenario C: Withdrawal + AOS ═══
    r = sec(ws, r, "SCENARIO C: WITHDRAWAL + AOS (shared services)")
    r = note(ws, r, "Same as Scenario B but with AOS cost sharing for admin functions.")
    r = note(ws, r, "Model: AOS #94 (6 towns, shared superintendent); Bristol/South Bristol ESC.")
    r = note(ws, r, "NOTE: No local AOS partner identified. This scenario is a future optimization.")
    r += 1

    R.CI_AOS_GROSS = r
    put(ws, r, 1, "Gross independent budget (from Scenario B, low)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_WD_GROSS}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Less: AOS shared services savings", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!B{R.WD_AOS_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "SpEd risk reserve", fill=CALC_FILL)
    put(ws, r, 2, 50000, USD); r += 1

    put(ws, r, 1, "Less: State EPS + small school adjustment", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_EPS_ADJ}+B{R.CI_EPS_ADJ+1}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.CI_AOS_LOCAL = r
    put(ws, r, 1, "NET LOCAL COST (AOS model)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.CI_AOS_GROSS}:B{r-1})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_AOS_DELTA = r
    put(ws, r, 1, "CHANGE vs. CURRENT RSU ASSESSMENT", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_AOS_LOCAL}-B{R.CI_CUR_ASSESS}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_AOS_MIL = r
    put(ws, r, 1, "New estimated mil rate (AOS model)", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_AOS_LOCAL}/{IW}!B{R.WD_VALUATION}*1000"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.000'; r += 1

    R.CI_AOS_PH = r
    put(ws, r, 1, "Annual cost change per household", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_AOS_DELTA}/{IW}!B{R.WD_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ═══ Scenario D: Legal Challenge Costs ═══
    r = sec(ws, r, "SCENARIO D: LEGAL CHALLENGE (§1512 contest)")
    r = note(ws, r, "Cost to challenge the 'reconfiguration vs. closure' framing in court or")
    r = note(ws, r, "via petition to Commissioner of Education. One-time cost, uncertain outcome.")
    r += 1

    R.CI_LEGAL_LOW = r
    put(ws, r, 1, "Legal challenge cost (low estimate)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_LEGAL_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.CI_LEGAL_HIGH = r
    put(ws, r, 1, "Legal challenge cost (high estimate)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!C{R.WD_LEGAL_TOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Probability of success (analyst assessment)", fill=PARAM_FILL)
    put(ws, r, 2, 0.35, PCT); r += 1
    R.CI_LEGAL_PROB = r - 1

    r = note(ws, r, "Assessment: statutory text favors 'reconfiguration' interpretation; MSAD #1 precedent")
    r = note(ws, r, "supports district. Counterarguments exist but are untested. [Fn 30][Fn 39]")
    r += 2

    # ═══ Scenario E: K-8 PES Expansion ═══
    r = sec(ws, r, "SCENARIO E: K-8 PES EXPANSION (tuition 9-12 only)")
    r = note(ws, r, "PES expands from K-5 to K-8. Pownal operates its own middle school.")
    r = note(ws, r, "Only grades 9-12 are tuitioned to Freeport High School (or elsewhere).")
    r = note(ws, r, "Trades higher operating cost for lower tuition spend and more control.")
    r += 1

    R.CI_K8_START = r
    put(ws, r, 1, "PES school-level operating cost (K-5, current)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_PES_SUBTOTAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Additional cost: grades 6-8 staffing + operations", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_K8_ANNUAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Admin costs (Scenario B excl. tuition, SpEd, contingency)", fill=CALC_FILL)
    ws.cell(r, 2).value = (
        f"={IW}!B{R.WD_IND_TOTAL}"
        f"-{IW}!B{R.WD_TUITION_68}"
        f"-{IW}!B{R.WD_TUITION_912}"
        f"-{IW}!B{R.WD_SPED_TUITION}"
        f"-{IW}!B{R.WD_CONTINGENCY}"
    )
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "HS tuition (9-12 only)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IW}!B{R.WD_K8_HS_TUITION}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "SpEd for HS tuitioned students (9-12 only)", fill=CALC_FILL)
    put(ws, r, 2, 75000, USD); r += 1

    put(ws, r, 1, "Contingency (5% of above)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=SUM(B{R.CI_K8_START}:B{r-1})*0.05"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.CI_K8_GROSS = r
    put(ws, r, 1, "GROSS K-8 INDEPENDENT BUDGET", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{R.CI_K8_START}:B{r-1})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    put(ws, r, 1, "Less: State EPS share", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!B{R.WD_EPS_STATE}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    R.CI_K8_EPS = r; r += 1

    put(ws, r, 1, "Less: Isolated small school adjustment", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{IW}!B{R.WD_EPS_SMALL}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    R.CI_K8_LOCAL = r
    put(ws, r, 1, "NET LOCAL COST (K-8 model)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_K8_GROSS}+B{R.CI_K8_EPS}+B{r-1}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    R.CI_K8_DELTA = r
    put(ws, r, 1, "CHANGE vs. CURRENT RSU ASSESSMENT", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_K8_LOCAL}-B{R.CI_CUR_ASSESS}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    R.CI_K8_PH = r
    put(ws, r, 1, "Annual cost change per household", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.CI_K8_DELTA}/{IW}!B{R.WD_HOUSING}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "Mil rate (K-8 model)", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_K8_LOCAL}/{IW}!B{R.WD_VALUATION}*1000"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.000'; r += 1

    R.CI_K8_VS_B = r
    put(ws, r, 1, "vs. Scenario B (K-5 + full tuition)", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_K8_LOCAL}-B{R.CI_WD_LOCAL}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED; r += 2

    r = note(ws, r, "K-8 TRADEOFFS:")
    r = note(ws, r, "  PRO: Local control through 8th grade; students stay in Pownal 3 more years.")
    r = note(ws, r, "  PRO: Reduces tuition dependency; fewer students at other districts' mercy.")
    r = note(ws, r, "  PRO: PES has physical capacity (~180 design, ~150 at K-8).")
    r = note(ws, r, "  CON: Small MS cohort (~45) limits course variety -- no French, Latin, band at scale.")
    r = note(ws, r, "  CON: MS athletics challenging with 15 students/grade.")
    r = note(ws, r, "  CON: Requires DOE approval for expanded grade range.")
    r = note(ws, r, "  CON: One-time facility upgrades (science lab, ~$75K) not in annual budget.")
    r = note(ws, r, "  NEUTRAL: Net annual cost difference vs Scenario B depends on teacher count.")
    r += 1

    # ═══ Break-Even Analysis (after all scenarios defined) ═══
    r = sec(ws, r, "BREAK-EVEN: ALL SCENARIOS vs. RSU MEMBERSHIP")
    r = note(ws, r, "Compares annual ongoing cost of each independence scenario to current RSU local tax.")
    r = note(ws, r, "Scenario B shown with low/high range; C and E use low estimates.")
    r += 1

    R.CI_BE_CUR = r
    put(ws, r, 1, "Current RSU local tax (Pownal)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_CUR_ASSESS}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Scenario B cost, low (K-5 + full 6-12 tuition)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_WD_LOCAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Scenario B cost, high", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=C{R.CI_WD_LOCAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Scenario C cost (K-5 + AOS, low)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_AOS_LOCAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Scenario E cost (K-8 + HS tuition)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{R.CI_K8_LOCAL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    R.CI_BE_VERDICT = r
    put(ws, r, 1, "Savings vs RSU: Scenario B low (best case)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-B{R.CI_WD_DELTA}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "Savings vs RSU: Scenario B high (worst case)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-B{R.CI_WD_DELTA_HIGH}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "Savings vs RSU: Scenario C (K-5 + AOS, low)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-B{R.CI_AOS_DELTA}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "Savings vs RSU: Scenario E (K-8 + HS tuition)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-B{R.CI_K8_DELTA}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    r = note(ws, r, "Positive = independence SAVES money vs. current RSU local tax.")
    r = note(ws, r, "Negative = independence costs MORE than current RSU local tax.")
    r = note(ws, r, "All scenarios include $50K/year SpEd risk reserve.")
    r += 1

    # ═══ 5-Year Projection ═══
    r = sec(ws, r, "5-YEAR COST PROJECTION (Scenario B, low estimate)")
    r = note(ws, r, "Assumes 6.1% annual cost growth for both RSU and independent operation.")
    r = note(ws, r, "Year 1 includes uncapped tuition penalty, startup costs, and fund balance return.")
    r += 1

    for i, h in enumerate(["Year", "RSU Assessment (est.)", "Independent Cost (est.)",
                            "Annual Savings", "Cumulative Savings"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    R.CI_5Y_START = r
    for yr in range(1, 6):
        put(ws, r, 1, f"Year {yr}", fill=CALC_FILL)
        if yr == 1:
            ws.cell(r, 2).value = f"=B{R.CI_CUR_ASSESS}*1.061"
            ws.cell(r, 3).value = f"=B{R.CI_WD_LOCAL}*1.061+B{R.CI_Y1_NET}"
        else:
            ws.cell(r, 2).value = f"=B{r-1}*1.061"
            ws.cell(r, 3).value = f"=B{R.CI_WD_LOCAL}*1.061^{yr}"
        dat(ws, r, 2, CALC_FILL).number_format = USD
        dat(ws, r, 3, CALC_FILL).number_format = USD
        ws.cell(r, 4).value = f"=B{r}-C{r}"
        dat(ws, r, 4, CALC_FILL).number_format = SIGNED
        if yr == 1:
            ws.cell(r, 5).value = f"=D{r}"
        else:
            ws.cell(r, 5).value = f"=E{r-1}+D{r}"
        dat(ws, r, 5, CALC_FILL).number_format = SIGNED
        r += 1
    R.CI_5Y_END = r - 1

    R.CI_5Y_TOTAL = r
    put(ws, r, 1, "5-YEAR CUMULATIVE SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=E{R.CI_5Y_END}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    r = note(ws, r, "Positive values = independence SAVES money vs. staying in RSU 5.")
    r = note(ws, r, "Year 1 includes one-time adjustments (uncapped tuition, startup, fund balance return).")
    r = note(ws, r, "Years 2-5 use ongoing costs only. Growth rate: 6.1% (matches Omnibus Section 2.8).")
    r = note(ws, r, "Does NOT include non-financial benefits: full local control, guaranteed school operation,")
    r = note(ws, r, "community identity preservation, property value protection.")


def build_c_legal(wb):
    """Legal defense assessment and decision framework."""
    ws = wb.create_sheet("C-LegalAnalysis")
    ws.sheet_properties.tabColor = "FF4444"
    col_widths(ws, [52, 20, 20, 20, 44])

    r = ttl(ws, 1, "CALC: Legal & Strategic Analysis")
    r = note(ws, r + 1, "Structured assessment of Pownal's legal options. Not formula-heavy;")
    r = note(ws, r, "provides decision framework for qualitative dimensions.")
    r += 1

    # ═══ PART 1: Reconfiguration vs. Closure ═══
    r = sec(ws, r, "PART 1: RECONFIGURATION vs. CLOSURE -- §1512 APPLICABILITY [Fn 30][Fn 39]")
    r += 1

    put(ws, r, 1, "STATUTORY TEXT (Title 20-A §1512(4)):", fill=SOURCE_FILL, font=BOLD)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    put(ws, r, 1, '"...a school closing is any action by the regional school unit board', fill=SOURCE_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
    put(ws, r, 1, 'that has the effect of providing no instruction for any students at that school."', fill=SOURCE_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 2

    for i, h in enumerate(["Argument", "For §1512 Applying", "Against §1512 Applying",
                            "Strength"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.CL_ARG_START = r
    arguments = [
        ("Plain text: 'no instruction for any students'",
         "PES K-5 students lose all instruction at PES",
         "EC/PreK students still receive instruction at the building",
         "AGAINST"),
        ("'School' vs. 'building' distinction",
         "PES the K-5 school ceases to exist; EC is a new program",
         "Statute refers to 'at that school' = the physical location",
         "MIXED"),
        ("Legislative intent of §1512",
         "Enacted to protect small communities from losing their schools",
         "Legislature could have written 'program' but chose 'instruction'",
         "FOR (spirit)"),
        ("MSAD #1 precedent (Feb 2026) [Fn 39]",
         "Different district; no court ruling; administrative practice only",
         "Pine St Elem -> PreK-Gr1 treated as reconfiguration, not closure",
         "AGAINST"),
        ("De facto closure doctrine",
         "Every K-5 student displaced; community loses its school in substance",
         "Courts have not applied this doctrine to Maine §1512 specifically",
         "MIXED"),
        ("Commissioner of Education ruling",
         "Pownal could petition for ruling; untested question = uncertainty",
         "Commissioner may defer to district's administrative discretion",
         "UNCERTAIN"),
    ]
    for label, pro, con, strength in arguments:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, pro, fill=CALC_FILL)
        put(ws, r, 3, con, fill=CALC_FILL)
        strength_fill = RESULT_FILL if "AGAINST" in strength else PARAM_FILL if "FOR" in strength else CALC_FILL
        put(ws, r, 4, strength, fill=strength_fill)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)
        ws.cell(r, 3).alignment = Alignment(wrap_text=True)
        r += 1
    R.CL_ARG_END = r - 1
    r += 1

    put(ws, r, 1, "NET ASSESSMENT:", fill=RESULT_FILL, font=BOLD)
    put(ws, r, 2, "§1512 likely does NOT block Option 2 as framed.", fill=RESULT_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2).font = RESULT_FONT; r += 1
    put(ws, r, 1, "IMPLICATION:", fill=RESULT_FILL, font=BOLD)
    put(ws, r, 2, "Withdrawal (§1466) is the more reliable recourse path.", fill=RESULT_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 2).font = RESULT_FONT; r += 2

    # ═══ PART 2: Option Comparison Matrix ═══
    r = sec(ws, r, "PART 2: RECOURSE OPTION COMPARISON MATRIX")
    r = note(ws, r, "Scale: 1 (worst) to 5 (best) for each dimension.")
    r += 1

    headers = ["Option", "Legal\nStrength", "Est. Cost", "Timeline", "Control\nGained",
               "Reversibility", "Risk", "Precedent"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers)); r += 1

    R.CL_MAT_START = r
    options = [
        ("§1512 legal challenge", 2, "$25K-$88K", "6-18 months",
         2, 5, 3, 2),
        ("Full RSU withdrawal (§1466)", 5, "$70K + taxes", "12-24 months",
         5, 3, 3, 4),
        ("Withdrawal + AOS", 5, "$70K + taxes", "12-24 months",
         4, 3, 2, 3),
        ("Withdrawal + K-8 PES (Scenario E)", 5, "$70K + taxes", "12-24 months",
         5, 2, 3, 2),
        ("Advocate ALM formula change (§1481-A)", 3, "$5K-$15K", "6-12 months",
         2, 5, 1, 3),
        ("Leverage (threaten withdrawal)", 4, "$5K-$50K", "3-6 months",
         3, 5, 1, 4),
        ("Do nothing / accept", 0, "$0", "N/A",
         0, 0, 0, 0),
    ]
    for label, legal, cost, timeline, control, rev, risk, prec in options:
        put(ws, r, 1, label, fill=CALC_FILL)
        if isinstance(legal, int) and legal > 0:
            put(ws, r, 2, legal, '0', fill=CALC_FILL)
        else:
            put(ws, r, 2, legal, fill=CALC_FILL)
        put(ws, r, 3, cost, fill=CALC_FILL)
        put(ws, r, 4, timeline, fill=CALC_FILL)
        if isinstance(control, int) and control > 0:
            put(ws, r, 5, control, '0', fill=CALC_FILL)
            put(ws, r, 6, rev, '0', fill=CALC_FILL)
            put(ws, r, 7, risk, '0', fill=CALC_FILL)
            put(ws, r, 8, prec, '0', fill=CALC_FILL)
        else:
            for c in range(5, 9):
                put(ws, r, c, "N/A", fill=CALC_FILL)
        r += 1
    R.CL_MAT_END = r - 1
    r += 1

    r = note(ws, r, "Legal Strength: 5=clear statutory right, 1=contested/weak")
    r = note(ws, r, "Control: 5=full local autonomy, 1=minimal influence")
    r = note(ws, r, "Risk: 5=low risk, 1=high risk. Reversibility: 5=fully reversible, 1=permanent")
    r = note(ws, r, "Precedent: 5=strong Maine precedent, 1=no precedent")
    r += 1

    # ═══ PART 3: Precedent Comparison ═══
    r = sec(ws, r, "PART 3: PARALLEL PRECEDENTS [Fn 33][Fn 34][Fn 38][Fn 39]")
    r += 1

    prec_headers = ["District / Case", "Action", "Outcome", "Similarity\nto Pownal",
                     "Relevance"]
    for i, h in enumerate(prec_headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(prec_headers)); r += 1

    R.CL_PREC_START = r
    prec_data = [
        ("MSAD #1 / Pine Street (2026) [Fn 39]",
         "Elem -> PreK-Gr1 (reconfiguration)",
         "Treated as admin decision, not §1512 closure",
         "HIGH: same EC mandate, same building repurpose",
         "KEY"),
        ("SAD 75 / Harpswell (2026) [Fn 38]",
         "Community opposed closure of only school",
         "District abandoned ALL closure options",
         "HIGH: small town, community school identity",
         "HIGH"),
        ("Freeport / RSU 5 (2014) [Fn 33]",
         "Withdrawal petition, full agreement drafted",
         "Vote failed by 76 votes (2152-2228)",
         "DIRECT: same RSU, same legal framework",
         "KEY"),
        ("Dayton / RSU 23 (2014) [Fn 35]",
         "Full withdrawal, independent operation",
         "Success; mil rate 18.47->20.85; full control",
         "MODERATE: larger, different region",
         "HIGH"),
        ("Andover / SAD 44 (2015) [Fn 35]",
         "Full withdrawal, independent K-5",
         "Success; mil rate 15.6->19.6; ~30 students",
         "HIGH: similar enrollment size",
         "HIGH"),
        ("Embden / RSU 74 (2025) [Fn 37]",
         "Withdrawal petition filed, $50K authorized",
         "Process ongoing",
         "MODERATE: rural, different region",
         "MODERATE"),
    ]
    for dist, action, outcome, sim, rel in prec_data:
        put(ws, r, 1, dist, fill=CALC_FILL)
        put(ws, r, 2, action, fill=CALC_FILL)
        put(ws, r, 3, outcome, fill=CALC_FILL)
        put(ws, r, 4, sim, fill=CALC_FILL)
        rel_fill = RESULT_FILL if rel == "KEY" else CALC_FILL
        put(ws, r, 5, rel, fill=rel_fill)
        for c in range(1, 6):
            ws.cell(r, c).alignment = Alignment(wrap_text=True)
        r += 1
    R.CL_PREC_END = r - 1
    r += 1

    # ═══ PART 4: Strategic Timeline ═══
    r = sec(ws, r, "PART 4: STRATEGIC TIMELINE & DECISION TREE")
    r += 1

    for i, h in enumerate(["Date / Window", "Action", "Prerequisite",
                            "Statute"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    R.CL_TIME_START = r
    timeline_data = [
        ("Immediate (Spring 2026)", "Gather 10% petition signatures (§1466)",
         "Pownal voter roll from last gubernatorial election", "20-A §1466(1)"),
        ("Spring 2026", "Town vote: authorize withdrawal committee + $50K",
         "Petition signatures verified", "20-A §1466(2)"),
        ("Spring 2026 (parallel)", "File §1512 challenge with Commissioner",
         "Legal counsel retained; Board has voted on Option 2", "20-A §1512"),
        ("Summer-Fall 2026", "Withdrawal committee negotiates agreement with RSU 5",
         "Committee formed; 90-day negotiation window", "20-A §1466(4)"),
        ("By November 30, 2026", "Final referendum vote on withdrawal agreement",
         "Agreement approved by Commissioner", "20-A §1466(9)"),
        ("July 1, 2027", "Withdrawal effective; new Pownal SAU operational",
         "Referendum passed by Nov 30, 2026", "20-A §1466"),
        ("Fall 2027", "Pownal operates PES independently; tuitions 6-12",
         "SAU established, superintendent hired, tuition agreements signed", ""),
    ]
    for date, action, prereq, statute in timeline_data:
        put(ws, r, 1, date, fill=CALC_FILL)
        put(ws, r, 2, action, fill=CALC_FILL)
        put(ws, r, 3, prereq, fill=CALC_FILL)
        put(ws, r, 4, statute, fill=SOURCE_FILL)
        for c in range(1, 5):
            ws.cell(r, c).alignment = Alignment(wrap_text=True)
        r += 1
    R.CL_TIME_END = r - 1
    r += 1

    r = note(ws, r, "CRITICAL: The Nov 30 deadline is absolute. Missing it delays withdrawal by one full year.")
    r = note(ws, r, "STRATEGY: Initiate §1512 challenge AND withdrawal petition simultaneously.")
    r = note(ws, r, "If §1512 challenge succeeds -> withdrawal becomes unnecessary leverage.")
    r = note(ws, r, "If §1512 challenge fails -> withdrawal is already underway as backup.")
    r += 1

    # ═══ PART 5: Recommended Strategy ═══
    r = sec(ws, r, "PART 5: RECOMMENDED STRATEGY")
    r += 1

    strategies = [
        "PHASE 1 (Immediate): Retain education law counsel. File petition for §1512",
        "   ruling with Commissioner. Cost: $15K-$25K. This tests the legal question",
        "   and creates public record of Pownal's objection.",
        "",
        "PHASE 2 (Concurrent): Begin §1466 withdrawal petition. Collect signatures,",
        "   hold town vote to authorize committee and $50K. This is the credible threat.",
        "   Even if Pownal doesn't ultimately withdraw, the process forces RSU 5 to",
        "   negotiate seriously. Freeport's 2014 attempt (failed by 76 votes) proves",
        "   withdrawal is a realistic possibility in this district.",
        "",
        "PHASE 3 (Conditional): If RSU 5 Board abandons Option 2 (as SAD 75 did for",
        "   Harpswell), withdrawal process can be paused or terminated at any point.",
        "   If Board proceeds, complete withdrawal agreement by Nov 30 deadline.",
        "",
        "KEY INSIGHT: The withdrawal process is fully reversible until the final",
        "   referendum vote. Starting it costs ~$50K but creates enormous leverage.",
        "   NOT starting it leaves Pownal with no reliable legal defense against",
        "   Option 2 as framed.",
    ]
    for s in strategies:
        if s == "":
            r += 1
        elif s.startswith("PHASE") or s.startswith("KEY"):
            ws.cell(r, 1, s).font = BOLD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1
        else:
            r = note(ws, r, s)


def build_c_doe_staffing(wb):
    """Analysis of DOE staffing trends -- formulas referencing I-DOEStaffing."""
    ws = wb.create_sheet("C-DOEStaffing")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [38, 14, 14, 14, 14, 50])
    IDS = SN['ids']

    r = ttl(ws, 1, "CALC: DOE Staffing Trend Analysis")
    r = note(ws, r + 1, "All values are formulas referencing I-DOEStaffing. Source: Maine DOE [Fn 45]")
    r += 1

    # ── Section 1: 9-Year Total FTE Change (2017 baseline) ──
    r = sec(ws, r, "1. Nine-Year Total FTE Change by School (2017 to 2026)")
    r = note(ws, r, "2016 excluded from baseline: different staffing classification model produces non-comparable FTE counts.")
    for i, h in enumerate(["School", "2017 FTE", "2026 FTE", "Change", "% Change"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1

    school_refs = [
        ("PES (Pownal Elementary)", R.DS_TOT_PES),
        ("DCS (Durham Community)",  R.DS_TOT_DCS),
        ("MSS (Morse Street)",      R.DS_TOT_MSS),
        ("MLS (Mast Landing)",      R.DS_TOT_MLS),
        ("FMS (Freeport Middle)",   R.DS_TOT_FMS),
        ("FHS (Freeport High)",     R.DS_TOT_FHS),
        ("District-wide (central)", R.DS_TOT_DW),
    ]
    R.CD_CHG_START = r
    for name, ref in school_refs:
        put(ws, r, 1, name, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={IDS}!C{ref}"; dat(ws, r, 2, CALC_FILL).number_format = '0.0'
        ws.cell(r, 3).value = f"={IDS}!L{ref}"; dat(ws, r, 3, CALC_FILL).number_format = '0.0'
        ws.cell(r, 4).value = f"=C{r}-B{r}"; dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
        ws.cell(r, 5).value = f"=D{r}/B{r}"; dat(ws, r, 5, CALC_FILL).number_format = '+0.0%;-0.0%;0.0%'
        r += 1
    R.CD_CHG_END = r - 1
    R.CD_CHG_PES = R.CD_CHG_START
    R.CD_CHG_DW = R.CD_CHG_END
    r += 1

    put(ws, r, 1, "GRAND TOTAL", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={IDS}!C{R.DS_GRAND}"; dat(ws, r, 2, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 3).value = f"={IDS}!L{R.DS_GRAND}"; dat(ws, r, 3, RESULT_FILL).number_format = '0.0'
    ws.cell(r, 4).value = f"=C{r}-B{r}"; dat(ws, r, 4, RESULT_FILL).number_format = '+0.0;-0.0;0.0'
    ws.cell(r, 5).value = f"=D{r}/B{r}"; dat(ws, r, 5, RESULT_FILL).number_format = '+0.0%;-0.0%;0.0%'
    R.CD_GRAND_CHG = r
    r += 2

    # ── Section 2: PES Share of District Total FTE ──
    r = sec(ws, r, "2. PES Share of District Total FTE (Trend)")
    r = note(ws, r, "Compares PES's share of total staffing to its share of enrollment (~10.5%).")
    for i, h in enumerate(["Year", "PES FTE", "Grand Total", "PES Share"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    YEARS = list(range(2016, 2027))
    R.CD_SHARE_START = r
    for idx, year in enumerate(YEARS):
        cl = get_column_letter(idx + 2)
        put(ws, r, 1, year, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={IDS}!{cl}{R.DS_TOT_PES}"; dat(ws, r, 2, CALC_FILL).number_format = '0.0'
        ws.cell(r, 3).value = f"={IDS}!{cl}{R.DS_GRAND}"; dat(ws, r, 3, CALC_FILL).number_format = '0.0'
        ws.cell(r, 4).value = f"=B{r}/C{r}"; dat(ws, r, 4, CALC_FILL).number_format = '0.0%'
        r += 1
    R.CD_SHARE_END = r - 1
    R.CD_SHARE_2017 = R.CD_SHARE_START + 1
    R.CD_SHARE_2025 = R.CD_SHARE_START + 9
    R.CD_SHARE_2026 = R.CD_SHARE_END
    r += 1

    IE = SN['ie']
    put(ws, r, 1, "PES enrollment share (FY27)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={IE}!F{R.PES_R}"
    dat(ws, r, 2, RESULT_FILL).number_format = '0'
    ws.cell(r, 3).value = f"=SUM({IE}!F{R.E_START}:{IE}!F{R.E_END})"
    dat(ws, r, 3, RESULT_FILL).number_format = '0'
    ws.cell(r, 4).value = f"=B{r}/C{r}"
    dat(ws, r, 4, RESULT_FILL).number_format = '0.0%'
    ws.cell(r, 4).font = RESULT_FONT
    R.CD_ENROLL_SHARE = r
    r += 1

    put(ws, r, 1, "STAFFING GAP: PES FTE share vs enrollment share", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=D{R.CD_SHARE_2026}-D{R.CD_ENROLL_SHARE}"
    dat(ws, r, 2, RESULT_FILL).number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(r, 2).font = RESULT_FONT
    R.CD_FTE_GAP = r
    r += 2

    # ── Section 3: Instructional Staff Analysis (2017 baseline) ──
    r = sec(ws, r, "3. Instructional Staff (Teachers + Ed Techs + SpEd): 9-Year Change")
    r = note(ws, r, "Combined Classroom Teacher + Ed Tech (I/II/III) + Special Ed Teacher FTE.")
    for i, h in enumerate(["School", "2017 Instr", "2026 Instr", "Change", "% Change"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1

    instr_schools = [
        ("PES", R.DS_TCH_START, R.DS_ET_START, R.DS_SPED_START),
        ("DCS", R.DS_TCH_START + 1, R.DS_ET_START + 1, R.DS_SPED_START + 1),
        ("MSS", R.DS_TCH_START + 2, R.DS_ET_START + 2, R.DS_SPED_START + 2),
        ("MLS", R.DS_TCH_START + 3, R.DS_ET_START + 3, R.DS_SPED_START + 3),
        ("FMS", R.DS_TCH_START + 4, R.DS_ET_START + 4, R.DS_SPED_START + 4),
        ("FHS", R.DS_TCH_START + 5, R.DS_ET_START + 5, R.DS_SPED_START + 5),
    ]
    R.CD_INSTR_START = r
    for name, tch_r, et_r, sped_r in instr_schools:
        put(ws, r, 1, name, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={IDS}!C{tch_r}+{IDS}!C{et_r}+{IDS}!C{sped_r}"
        dat(ws, r, 2, CALC_FILL).number_format = '0.0'
        ws.cell(r, 3).value = f"={IDS}!L{tch_r}+{IDS}!L{et_r}+{IDS}!L{sped_r}"
        dat(ws, r, 3, CALC_FILL).number_format = '0.0'
        ws.cell(r, 4).value = f"=C{r}-B{r}"
        dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
        ws.cell(r, 5).value = f"=D{r}/B{r}"
        dat(ws, r, 5, CALC_FILL).number_format = '+0.0%;-0.0%;0.0%'
        r += 1
    R.CD_INSTR_END = r - 1
    R.CD_INSTR_PES = R.CD_INSTR_START
    r += 2

    # ── Section 4: PES SpEd Staffing History ──
    r = sec(ws, r, "4. PES Special Education Staffing (Trend)")
    r = note(ws, r, "PES SpEd FTE hit 0.0 in school years 2022-23 and 2023-24.")
    for i, h in enumerate(["Year", "PES SpEd FTE", "District Elem Avg", "PES vs Avg"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    R.CD_SPED_START = r
    for idx, year in enumerate(YEARS):
        cl = get_column_letter(idx + 2)
        put(ws, r, 1, year, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={IDS}!{cl}{R.DS_SPED_PES}"
        dat(ws, r, 2, CALC_FILL).number_format = '0.0'
        ws.cell(r, 3).value = (
            f"=AVERAGE({IDS}!{cl}{R.DS_SPED_START+1},"
            f"{IDS}!{cl}{R.DS_SPED_START+2},"
            f"{IDS}!{cl}{R.DS_SPED_START+3})"
        )
        dat(ws, r, 3, CALC_FILL).number_format = '0.0'
        ws.cell(r, 4).value = f"=B{r}-C{r}"
        dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
        r += 1
    R.CD_SPED_END = r - 1
    r += 1

    # ── Section 5: Cross-reference to I-FTE ──
    IFT = SN['ift']
    r = sec(ws, r, "5. Cross-Reference: DOE 2026 vs. I-FTE Proposed FY27")
    r = note(ws, r, "DOE 'Classroom Teacher' includes all certified teaching staff (incl. specialists).")
    r = note(ws, r, "I-FTE counts core classroom teachers only. Difference = specialists, PreK, etc.")
    for i, h in enumerate(["School", "DOE 2026 Teacher", "I-FTE Proposed FY27", "Difference", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5)
    r += 1

    xref_schools = [
        ("PES", R.DS_TCH_PES, R.FT_PES,
         "DOE counts PreK, art, music, PE, specialists as 'Classroom Teacher'"),
        ("DCS", R.DS_TCH_DCS, R.FT_DCS,
         "DCS is PreK-8; DOE includes all certified teachers"),
    ]
    R.CD_XREF_START = r
    for name, doe_r, fte_r, notes in xref_schools:
        put(ws, r, 1, name, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={IDS}!L{doe_r}"
        dat(ws, r, 2, CALC_FILL).number_format = '0.0'
        ws.cell(r, 3).value = f"={IFT}!D{fte_r}"
        dat(ws, r, 3, CALC_FILL).number_format = '0.0'
        ws.cell(r, 4).value = f"=B{r}-C{r}"
        dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
        put(ws, r, 5, notes)
        r += 1
    R.CD_XREF_END = r - 1
    r += 2

    # ── Key Findings ──
    r = sec(ws, r, "KEY STAFFING FINDINGS (from DOE data, 2017-2026)")
    r = note(ws, r, "Baseline: 2017 (first year of consistent DOE staffing classification methodology).")
    findings = [
        "1. PES suffered the largest staffing decline of any school (-28.4%), more than double",
        "   the next largest (FHS: -12.4%). Four of six schools lost FTE; only MSS and MLS grew.",
        "2. District-wide (central office) FTE grew +13.3% over the same period (+16.4 FTE).",
        "3. PES instructional staff (teachers+ed techs+sped) declined by ~4.6 FTE while other",
        "   elementary schools each gained +5.9 to +9.9 FTE.",
        "4. PES SpEd teacher FTE hit 0.0 in both 2022-23 and 2023-24 school years.",
        "5. PES share of district FTE fell from 6.5% (2017) to 3.9% (2025), vs. ~10.5% enrollment share.",
        "6. The per-student cost premium at PES is NOT from overstaffing -- PES runs leaner than all peers.",
        "7. Resources have been systematically shifted away from PES while Pownal's tax assessment grew 29.8%.",
    ]
    for f in findings:
        r = note(ws, r, f)


def build_c_summary(wb):
    """Executive dashboard -- first tab stakeholders see. Comprehensive overview."""
    ws = wb.create_sheet("C-Summary")
    ws.sheet_properties.tabColor = "70AD47"
    col_widths(ws, [52, 22, 22, 22, 22])
    C27 = SN['c27']; C28 = SN['c28']; CFP = SN['cfp']; CB = SN['cb']
    IB = SN['ib']; CM = SN['cm']; IA = SN['ia']; CEC = SN['cec']
    IFT = SN['ift']; IRD = SN['ird']

    r = ttl(ws, 1, "RSU 5 Financial Analysis: Executive Summary")
    r = note(ws, r + 1, "All values are live formulas. To test scenarios, change yellow cells on I- and C-FY28Projection sheets.")
    r += 1

    # ── PART 1: FY27 ──
    r = sec(ws, r, "PART 1: FY27 (2026-2027) -- Budget Going to Voters")
    for i, h in enumerate(["", "Amount", "% of FY26"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    put(ws, r, 1, "FY26 Adopted Budget", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IB}!B{R.FY26_R}"; dat(ws, r, 2, CALC_FILL).number_format = USD
    R.S_FY26 = r; r += 1

    put(ws, r, 1, "FY27 Proposed (Superintendent)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CB}!B{R.C2_FY27}"; dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"={CB}!B{R.C2_INCR_PCT}"; dat(ws, r, 3, CALC_FILL).number_format = PCT2
    R.S_FY27 = r; r += 1

    put(ws, r, 1, "Dollar increase from FY26", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CB}!B{R.C2_INCR}"; dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1

    put(ws, r, 1, "Superintendent's initial request was 10.55%", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={IRD}!B{R.RD_INIT_INCR}"; dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Total staff/program reductions already applied", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C27}!B{R.C27_TOTAL_RED}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "  of which: T1 teaching positions", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C27}!B{R.C27_T1}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "  of which: T2 support positions", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C27}!B{R.C27_T2}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED; r += 1

    r = note(ws, r, "Board adopts March 25, 2026 | ABM: May 13 | Referendum: June 19")
    r = note(ws, r, "EC mandate and structural changes are NOT in FY27. They are FY28.")
    r = note(ws, r, "Historical: FY26 approved at 6.83% (817-494); FY25 at 6.48%; FY24 at 4.99%.")
    r += 1

    # ── PART 2: FY28 baseline ──
    r = sec(ws, r, "PART 2: FY28 (2027-2028) -- Baseline (No Structural Action)")
    put(ws, r, 1, "FY27 adopted (selected scenario)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C28}!B{R.C28_FY27_SEL}"; dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1
    put(ws, r, 1, "  + Annual cost growth", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C28}!B{R.C28_FY27_SEL}*{C28}!B{R.C28_GR_SEL}"
    dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1
    put(ws, r, 1, "  + EC mandate net cost (midpoint)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C28}!B{R.C28_EC}"; dat(ws, r, 2, CALC_FILL).number_format = USD; r += 1
    put(ws, r, 1, "FY28 BASELINE (no restructuring)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={C28}!B{R.C28_BASE}"; dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    R.S_BASE = r; r += 1
    put(ws, r, 1, "FY28 baseline increase from FY27", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={C28}!B{R.C28_BASE_PCT}"; dat(ws, r, 2, CALC_FILL).number_format = PCT2; r += 2

    # ── PART 3: Path A detail ──
    r = sec(ws, r, "PART 3: PATH A -- Preserve & Strengthen (FY28)")
    r = note(ws, r, "Consolidate middle grades for efficiency. Preserve all elementary schools. Distribute EC.")
    for i, h in enumerate(["Structural Change", "Annual Budget Impact", "Source Sheet"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    pa_detail = [
        ("1. Consolidate grades 7-8 at FMS (district-wide)",
         f"=-{CM}!B{R.C7_TEACH_SAV}",
         "C-MSConsol: Teacher savings from 7-8 consolidation"),
        ("2. Consolidate grade 6 at DCS (district-wide)",
         "0",
         "C-MSConsol: Included in net FTE calc above"),
        ("3. Extracurricular consolidation savings",
         f"=-{CM}!B{R.C7_EXTRA}",
         "C-MSConsol: Combined athletics/activities"),
        ("4. Freeport 6th-grade transportation to DCS",
         f"={IA}!B{R.A_FRE6_ROUTES}*{IA}!B{R.A_BUS_COST}",
         "I-Assumptions: 2 routes x $75K"),
        ("5. DCS portable classrooms needed",
         "0",
         "Net zero: gains 6th sections, loses 7-8 sections"),
        ("6. DCS admin complexity (multi-community 6th)",
         "0",
         "Complexity offsets any admin savings"),
        ("7. Additional line-item efficiencies (midpoint)",
         f"=-AVERAGE({IA}!B{R.A_EFF_LOW},{IA}!B{R.A_EFF_HIGH})",
         "I-Assumptions: transport, facilities, attrition"),
    ]
    pa_start = r
    for label, formula, source in pa_detail:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = f"={formula}" if not formula.startswith("=") else formula
        dat(ws, r, 2, CALC_FILL).number_format = SIGNED
        put(ws, r, 3, source, fill=CALC_FILL)
        r += 1
    pa_end = r - 1
    put(ws, r, 1, "TOTAL PATH A SAVINGS (cross-check)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=-SUM(B{pa_start}:B{pa_end})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT; r += 1
    put(ws, r, 1, "Path A: FY28 adjusted budget", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{R.S_BASE}+SUM(B{pa_start}:B{pa_end})"
    dat(ws, r, 2, RESULT_FILL).number_format = USD; r += 1
    put(ws, r, 1, "Path A: FY28 increase from FY27", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=(B{r-1}-{C28}!B{R.C28_FY27_SEL})/{C28}!B{R.C28_FY27_SEL}"
    dat(ws, r, 2, RESULT_FILL).number_format = PCT2
    ws.cell(r, 2).font = RESULT_FONT; r += 1
    r = note(ws, r, "What does NOT change: PES stays PreK-5. All 3 elementary schools preserved.")
    r = note(ws, r, "EC distributed: SpEd classroom at MSS (64 PreK peers for LRE). Itinerant EC at all schools.")
    r = note(ws, r, "Net FTE reduction: ~2.4 teachers (from MS consolidation only).")
    r += 1

    # ── PART 4: Path B detail ──
    r = sec(ws, r, "PART 4: PATH B -- Scenario 2 / Full Restructuring (FY28)")
    r = note(ws, r, "Close PES as K-5. Pownal K-6 to DCS. All PreK to PES (EC center). District-wide 7-8 at FMS.")
    for i, h in enumerate(["Component", "Gross Lens", "True Efficiency"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    put(ws, r, 1, "PES instruction budget eliminated", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.FP_PB_INSTR}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    put(ws, r, 3, "--", fill=CALC_FILL); r += 1

    put(ws, r, 1, "MS consolidation savings (same as Path A)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=-{CM}!B{R.C7_TOTAL}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"=-{CM}!B{R.C7_TOTAL}"; dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "Marginal teacher savings from PES closure", fill=CALC_FILL)
    put(ws, r, 2, "--", fill=CALC_FILL)
    ws.cell(r, 3).value = f"={CFP}!B{R.PB2_START}"
    dat(ws, r, 3, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "PES admin eliminated / DCS coordination", fill=CALC_FILL)
    put(ws, r, 2, "--", fill=CALC_FILL)
    ws.cell(r, 3).value = f"={CFP}!B{R.PB2_START+1}+{CFP}!B{R.PB2_START+2}"
    dat(ws, r, 3, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "PES support services efficiency", fill=CALC_FILL)
    put(ws, r, 2, "--", fill=CALC_FILL)
    ws.cell(r, 3).value = f"={CFP}!B{R.PB2_END}"; dat(ws, r, 3, CALC_FILL).number_format = USD; r += 1

    put(ws, r, 1, "Transport: Pownal K-6 to DCS (3 routes)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.PB_COST_START}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"={CFP}!B{R.PB_COST_START}"; dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "Transport: Freeport PreK to PES (3 routes)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.PB_COST_START+1}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"={CFP}!B{R.PB_COST_START+1}"; dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "Transport: Durham PreK to PES (2 routes)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.PB_COST_START+2}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"={CFP}!B{R.PB_COST_START+2}"; dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "DCS multi-community coordination", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.PB_COST_START+3}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"=-{IA}!B{R.A_DCS_COORD}"; dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "PES EC conversion (amortized over 10 yrs)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CFP}!B{R.PB_COST_END}"; dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    ws.cell(r, 3).value = f"={CFP}!B{R.PB_COST_END}"; dat(ws, r, 3, CALC_FILL).number_format = SIGNED; r += 1

    put(ws, r, 1, "NET SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CFP}!B{R.FP_PB_SAV}"; dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"={CFP}!B{R.FP_TRUE_NET}"; dat(ws, r, 3, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT; ws.cell(r, 3).font = RESULT_FONT; r += 1

    r = note(ws, r, "Gross lens: treats full PES budget as eliminated. Ignores absorption costs.")
    r = note(ws, r, "True efficiency: tracks actual FTE changes. Pownal students still need teachers at DCS.")
    r = note(ws, r, f"Total FTE saved (Scenario 2): ={SN['ift']}!B{R.FT_SC2_NET} vs. Path A: ={CM}!B{R.C7_SAVED}")
    r += 1

    # ── PART 5: Side-by-side ──
    r = sec(ws, r, "PART 5: Side-by-Side Comparison (FY28)")
    for i, h in enumerate(["Metric", "No Action", "Path A", "Path B (gross)", "Path B (true)"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 5); r += 1

    put(ws, r, 1, "FY28 Budget", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={C28}!B{R.C28_BASE}"; dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 3).value = f"={C28}!B{R.C28_PA_BUD}"; dat(ws, r, 3, RESULT_FILL).number_format = USD
    ws.cell(r, 4).value = f"={C28}!B{R.C28_PB_BUD}"; dat(ws, r, 4, RESULT_FILL).number_format = USD
    ws.cell(r, 5).value = f"={C28}!B{R.C28_BASE}-{CFP}!B{R.FP_TRUE_NET}"
    dat(ws, r, 5, RESULT_FILL).number_format = USD
    R.SUM_FY28_ROW = r; r += 1

    put(ws, r, 1, "Increase from FY27 adopted", fill=RESULT_FILL)
    for c in range(2, 6):
        cl = get_column_letter(c)
        ws.cell(r, c).value = f"=({cl}{R.SUM_FY28_ROW}-{C28}!B{R.C28_FY27_SEL})/{C28}!B{R.C28_FY27_SEL}"
        dat(ws, r, c, RESULT_FILL).number_format = PCT2
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "Structural savings", fill=CALC_FILL)
    put(ws, r, 2, "--", fill=CALC_FILL)
    ws.cell(r, 3).value = f"={CFP}!B{R.FP_PA_SAV}"; dat(ws, r, 3, CALC_FILL).number_format = USD
    ws.cell(r, 4).value = f"={CFP}!B{R.FP_PB_SAV}"; dat(ws, r, 4, CALC_FILL).number_format = USD
    ws.cell(r, 5).value = f"={CFP}!B{R.FP_TRUE_NET}"; dat(ws, r, 5, CALC_FILL).number_format = SIGNED
    r += 1

    put(ws, r, 1, "Net FTE change", fill=CALC_FILL)
    put(ws, r, 2, "0", fill=CALC_FILL)
    ws.cell(r, 3).value = f"=-{CM}!B{R.C7_SAVED}"; dat(ws, r, 3, CALC_FILL).number_format = '+0.0;-0.0;0.0'
    ws.cell(r, 4).value = f"=-{SN['ift']}!B{R.FT_SC2_NET}"
    dat(ws, r, 4, CALC_FILL).number_format = '+0.0;-0.0;0.0'
    ws.cell(r, 5).value = f"=-{SN['ift']}!B{R.FT_SC2_NET}"
    dat(ws, r, 5, CALC_FILL).number_format = '+0.0;-0.0;0.0'
    r += 1

    put(ws, r, 1, "Marginal FTE beyond Path A", fill=CALC_FILL)
    put(ws, r, 2, "--", fill=CALC_FILL)
    put(ws, r, 3, "--", fill=CALC_FILL)
    ws.cell(r, 4).value = f"={SN['ift']}!B{R.FT_SC2_NET}-{CM}!B{R.C7_SAVED}"
    dat(ws, r, 4, CALC_FILL).number_format = '0.0'
    ws.cell(r, 5).value = f"=D{r}"; dat(ws, r, 5, CALC_FILL).number_format = '0.0'; r += 1

    put(ws, r, 1, "EC mandate cost (in all scenarios)", fill=CALC_FILL)
    for c in range(2, 6):
        ws.cell(r, c).value = f"={C28}!B{R.C28_EC}"; dat(ws, r, c, CALC_FILL).number_format = USD
    r += 1

    for label in [
        "PES PreK-5 preserved?",
        "Young children in home community?",
        "EC parent opt-out risk",
        "Political feasibility",
        "Reversible?",
    ]:
        put(ws, r, 1, label, fill=CALC_FILL)
        r += 1
    # Fill in the qualitative row values
    qual_row = r - 5
    for c, v in [(2, "Yes"), (3, "Yes"), (4, "No"), (5, "No")]:
        ws.cell(qual_row, c, v); dat(ws, qual_row, c, CALC_FILL)
    qual_row += 1
    for c, v in [(2, "Yes"), (3, "Yes"), (4, "No"), (5, "No")]:
        ws.cell(qual_row, c, v); dat(ws, qual_row, c, CALC_FILL)
    qual_row += 1
    for c, v in [(2, "N/A"), (3, "Low"), (4, "High"), (5, "High")]:
        ws.cell(qual_row, c, v); dat(ws, qual_row, c, CALC_FILL)
    qual_row += 1
    for c, v in [(2, "Low"), (3, "High"), (4, "Low"), (5, "Low")]:
        ws.cell(qual_row, c, v); dat(ws, qual_row, c, CALC_FILL)
    qual_row += 1
    for c, v in [(2, "--"), (3, "Yes"), (4, "No"), (5, "No")]:
        ws.cell(qual_row, c, v); dat(ws, qual_row, c, CALC_FILL)
    r += 1

    # ── PART 6: Key assumptions ──
    r = sec(ws, r, "PART 6: Key Assumptions (adjust on I- sheets to test sensitivity)")
    for i, h in enumerate(["Assumption", "Current Value", "Used In", "If Changed..."], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    assumptions_summary = [
        ("Cost per FTE", f"={IA}!B{R.A_COST_FTE}", USD, "C-MSConsol, C-FY28Paths",
         "Higher = more savings from consolidation"),
        ("Target class size (7-8)", f"={IA}!B{R.A_CS_78}", '0', "C-MSConsol",
         "Lower target = fewer FTE saved"),
        ("Bus route cost/year", f"={IA}!B{R.A_BUS_COST}", USD, "C-FY28Paths",
         "Higher = worse for Path B"),
        ("EC net cost (midpoint)", f"={CEC}!D{R.C8_NET}", USD, "C-FY28Projection",
         "Higher = higher FY28 baseline for all paths"),
        ("Cost growth rate (selected)", f"={C28}!B{R.C28_GR_SEL}", PCT2, "C-FY28Projection",
         "Higher growth = higher FY28 baseline"),
        ("PES conversion cost", f"={IA}!B{R.A_PES_PARTIAL}", USD, "C-FY28Paths",
         "Higher = worse for Path B"),
        ("Support services efficiency %", f"={IA}!B{R.A_SUPPORT_EFF}", PCT, "C-FY28Paths Path B",
         "Higher = more savings from PES closure"),
    ]
    for label, formula, fmt, used_in, sensitivity in assumptions_summary:
        put(ws, r, 1, label, fill=CALC_FILL)
        ws.cell(r, 2).value = formula; dat(ws, r, 2, CALC_FILL).number_format = fmt
        put(ws, r, 3, used_in, fill=CALC_FILL)
        put(ws, r, 4, sensitivity, fill=CALC_FILL)
        r += 1
    r += 1

    # ── PART 7: What doesn't change ──
    r = sec(ws, r, "PART 7: What Does NOT Change Between Paths")
    constants = [
        "EC mandate cost (~$690K/yr net) is identical in all scenarios. It is a state requirement.",
        "FY27 budget is decided independently. Structural changes are FY28.",
        "MS consolidation (7-8 at FMS, 6 at DCS) is common to BOTH Path A and Path B.",
        "Per-student funding formula (EPS/ALM/RLC) is set by the state, not by restructuring.",
        "Debt service ($1.07M) and CTE ($337K) are fixed regardless of school configuration.",
        "The FY28 baseline depends on which FY27 budget passes. See C-FY28Projection.",
    ]
    for c in constants:
        r = note(ws, r, c)
    r += 1

    # ── PART 8: Quantified Risks (referencing C-RiskModel) ──
    CRM = SN['crm']
    r = sec(ws, r, "PART 8: Quantified Risk Analysis (see C-RiskModel for detail)")
    for i, h in enumerate(["Risk Factor", "Mid-Scenario Impact", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    put(ws, r, 1, "Property value decline (10%, Pownal)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CRM}!C{R.RM_PV_START+1}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "Total value lost; per-home ~$34K [Fn 20]", fill=CALC_FILL); r += 1

    put(ws, r, 1, "  RSU 5 tax revenue loss (annual)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CRM}!B{R.RM_PV_RSU_LOSS}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "58.4% of lost property tax at $15.30/thousand", fill=CALC_FILL); r += 1

    put(ws, r, 1, "Enrollment attrition (10%)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CRM}!D{R.RM_EA_START+1}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "~11 students leave RSU 5; state aid lost annually", fill=CALC_FILL); r += 1

    put(ws, r, 1, "One-time transition costs (midpoint)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=AVERAGE({CRM}!B{R.RM_OT_TOTAL},{CRM}!C{R.RM_OT_TOTAL})"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "Staff, conversion, legal, portables, tech", fill=CALC_FILL); r += 1

    put(ws, r, 1, "Total annual risk cost (mid scenario)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CRM}!C{R.RM_TOTAL_RISK}"
    dat(ws, r, 2, RESULT_FILL).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "Sum of quantified annual risks", fill=RESULT_FILL); r += 1

    put(ws, r, 1, "PATH B RISK-ADJUSTED NET", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CRM}!B{R.RM_PB_RISK_ADJ}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "True net savings minus quantified risk costs", fill=RESULT_FILL); r += 1

    put(ws, r, 1, "Break-even period (Path B vs Path A)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CRM}!B{R.RM_BREAKEVEN}"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.0 "years"'
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "If negative margin, one-time costs are NEVER recovered", fill=RESULT_FILL); r += 1

    r = note(ws, r, "IRREVERSIBILITY: Rebuilding a K-5 program after closure costs more than was saved.")
    r = note(ws, r, "EC OPT-OUT: At 20% opt-out, per-student EC cost rises; SpEd LRE risk is significant.")
    r = note(ws, r, "MAINE LAW: 20-A MRSA requires public hearings, Board vote, potential referendum. [Fn 28]")
    r = note(ws, r, "PATH A RISKS are primarily political/logistical, NOT financial or irreversible.")
    r += 1

    # ── PART 9: Pownal Recourse Options (referencing C-Independence, C-LegalAnalysis) ──
    CI = SN['ci']; CLA = SN['cla']
    r = sec(ws, r, "PART 9: Pownal Recourse Options (see C-Independence, C-LegalAnalysis)")
    r = note(ws, r, "If RSU 5 Board proceeds with Option 2, Pownal has several legal/strategic options.")
    r += 1

    for i, h in enumerate(["Option", "Key Metric", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    put(ws, r, 1, "§1512 municipal veto (CONTESTED)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CI}!B{R.CI_VETO_NET}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    put(ws, r, 3, "Annual cost IF §1512 applies (likely does NOT for reconfiguration) [Fn 30][Fn 39]", fill=CALC_FILL); r += 1

    put(ws, r, 1, "Full withdrawal + tuition (§1466)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CI}!B{R.CI_WD_DELTA}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    put(ws, r, 3, "Annual cost change vs. RSU assessment; gains full control [Fn 32]", fill=CALC_FILL); r += 1

    put(ws, r, 1, "Withdrawal + AOS shared services", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CI}!B{R.CI_AOS_DELTA}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    put(ws, r, 3, "AOS reduces admin overhead; control + shared efficiency", fill=CALC_FILL); r += 1

    put(ws, r, 1, "K-8 PES expansion + HS tuition only (Scenario E)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CI}!B{R.CI_K8_DELTA}"
    dat(ws, r, 2, CALC_FILL).number_format = SIGNED
    put(ws, r, 3, "Operate K-8; tuition only 9-12; more control, trades tuition for staff", fill=CALC_FILL); r += 1

    put(ws, r, 1, "Legal challenge cost (low-high)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CI}!B{R.CI_LEGAL_LOW}"
    dat(ws, r, 2, CALC_FILL).number_format = USD
    ws.cell(r, 3).value = f"=TEXT({CI}!B{R.CI_LEGAL_HIGH},\"$#,##0\")&\" high est.; ~35% success probability [Fn 30]\""
    dat(ws, r, 3, CALC_FILL); r += 1

    put(ws, r, 1, "5-year cumulative cost of independence", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CI}!B{R.CI_5Y_TOTAL}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "Total premium over 5 years including process costs", fill=RESULT_FILL); r += 1

    r = note(ws, r, "CRITICAL: §1512 veto likely does NOT apply to Option 2 (reconfiguration, not closure).")
    r = note(ws, r, "MSAD #1 precedent (2026): converting elem to PreK-Gr1 treated as admin action. [Fn 39]")
    r = note(ws, r, "WITHDRAWAL is the more reliable recourse. Process is reversible until final referendum.")
    r = note(ws, r, "FREEPORT tried RSU 5 withdrawal in 2014; failed by only 76 votes. Template exists. [Fn 33]")
    r = note(ws, r, "Property: PES building deeds back to Pownal upon withdrawal. [Fn 33][Fn 34]")
    r += 1

    # ── PART 10: Equity & Cost-Sharing Analysis ──
    CEQ = SN['ceq']; IEQ = SN['ieq']
    r = sec(ws, r, "PART 10: Per-Town Equity & Cost-Sharing (see C-Equity, I-Equity)")
    r = note(ws, r, "Breaks down who pays vs. who consumes education in RSU 5. All formula-driven.")
    r += 1

    for i, h in enumerate(["Metric", "Pownal", "Durham", "Freeport"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4); r += 1

    put(ws, r, 1, "Students (attributed to town)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_STUDENTS}"
    ws.cell(r, 3).value = f"={CEQ}!C{R.CE_STUDENTS}"
    ws.cell(r, 4).value = f"={CEQ}!D{R.CE_STUDENTS}"
    for c in range(2, 5): dat(ws, r, c, CALC_FILL).number_format = '0'
    r += 1

    put(ws, r, 1, "Local tax assessment (RLC + ALM)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_LOCAL}"
    ws.cell(r, 3).value = f"={CEQ}!C{R.CE_LOCAL}"
    ws.cell(r, 4).value = f"={CEQ}!D{R.CE_LOCAL}"
    for c in range(2, 5): dat(ws, r, c, CALC_FILL).number_format = USD
    r += 1

    put(ws, r, 1, "LOCAL TAX PER STUDENT", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_LOCAL_PP}"
    ws.cell(r, 3).value = f"={CEQ}!C{R.CE_LOCAL_PP}"
    ws.cell(r, 4).value = f"={CEQ}!D{R.CE_LOCAL_PP}"
    for c in range(2, 5):
        dat(ws, r, c, RESULT_FILL).number_format = USD
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "NET POSITION: ACTUAL CONSUMPTION", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_ACTUAL_NET}"
    ws.cell(r, 3).value = f"={CEQ}!C{R.CE_ACTUAL_NET}"
    ws.cell(r, 4).value = f"={CEQ}!D{R.CE_ACTUAL_NET}"
    for c in range(2, 5):
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "NET POSITION: EQUALIZED (avg cost)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_SUBSIDY}"
    ws.cell(r, 3).value = f"={CEQ}!C{R.CE_SUBSIDY}"
    ws.cell(r, 4).value = f"={CEQ}!D{R.CE_SUBSIDY}"
    for c in range(2, 5):
        dat(ws, r, c, RESULT_FILL).number_format = SIGNED
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "Fairness ratio (1.0 = equitable)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_DISPARITY}"
    ws.cell(r, 3).value = f"={CEQ}!C{R.CE_DISPARITY}"
    ws.cell(r, 4).value = f"={CEQ}!D{R.CE_DISPARITY}"
    for c in range(2, 5):
        dat(ws, r, c, RESULT_FILL).number_format = '0.00'
        ws.cell(r, c).font = RESULT_FONT
    r += 1

    put(ws, r, 1, "3-year assessment increase (Pownal)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_3YR_PCT}"
    dat(ws, r, 2, RESULT_FILL).number_format = PCT2
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "FY24 audited → FY27 proposed [Fn 43][Fn 3]", fill=CALC_FILL); r += 1

    put(ws, r, 1, "Cost-sharing formula penalty (Pownal ALM)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CEQ}!B{R.CE_ALM_DIFF}"
    dat(ws, r, 2, RESULT_FILL).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "vs. enrollment-only allocation [Fn 41]", fill=CALC_FILL); r += 1

    r = note(ws, r, "ACTUAL CONSUMPTION: Pownal has a modest deficit (~$267K) due to PES cost premium.")
    r = note(ws, r, "TAX BURDEN: Pownal pays ~2x per-student local tax vs. Durham ($21.7K vs $11K).")
    r = note(ws, r, "FORMULA IMPACT: ALM formula penalizes Pownal $382K/yr vs a per-student split.")
    r = note(ws, r, "Both lenses are presented in C-Equity (sections 2A and 2B) for full objectivity.")
    r += 1

    # ── PART 11: DOE Staffing Analysis ──
    CDS = SN['cds']
    r = sec(ws, r, "PART 11: DOE Staffing Trends (see C-DOEStaffing, I-DOEStaffing) [Fn 45]")
    r = note(ws, r, "Independent Maine DOE data (Dec 1 headcounts, 2017-2026 baseline). Requested by Mark Dohle.")
    r = note(ws, r, "2016 excluded from baseline: different staffing classification model.")
    r += 1

    for i, h in enumerate(["Metric", "Value", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3); r += 1

    put(ws, r, 1, "PES total FTE (2017)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CDS}!B{R.CD_CHG_PES}"
    dat(ws, r, 2, CALC_FILL).number_format = '0.0'
    put(ws, r, 3, "DOE Dec 1 headcount, all position categories", fill=CALC_FILL); r += 1

    put(ws, r, 1, "PES total FTE (2026)", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CDS}!C{R.CD_CHG_PES}"
    dat(ws, r, 2, CALC_FILL).number_format = '0.0'
    put(ws, r, 3, "Most recent DOE data available", fill=CALC_FILL); r += 1

    put(ws, r, 1, "PES 9-year FTE change", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CDS}!E{R.CD_CHG_PES}"
    dat(ws, r, 2, RESULT_FILL).number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "Largest decline of any school; 4 of 6 schools lost FTE", fill=RESULT_FILL); r += 1

    put(ws, r, 1, "District-wide (central) 9-year FTE change", fill=CALC_FILL)
    ws.cell(r, 2).value = f"={CDS}!E{R.CD_CHG_DW}"
    dat(ws, r, 2, CALC_FILL).number_format = '+0.0%;-0.0%;0.0%'
    put(ws, r, 3, "Central office grew 123.4 to 139.8 FTE (+16.4 positions)", fill=CALC_FILL); r += 1

    put(ws, r, 1, "PES share of district FTE (2026)", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CDS}!D{R.CD_SHARE_2026}"
    dat(ws, r, 2, RESULT_FILL).number_format = '0.0%'
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "PES enrollment ~10.5%; FTE share less than half of enrollment share", fill=RESULT_FILL); r += 1

    put(ws, r, 1, "PES instructional staff 9-year change", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CDS}!D{R.CD_INSTR_PES}"
    dat(ws, r, 2, RESULT_FILL).number_format = '+0.0;-0.0;0.0'
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "Teachers+EdTechs+SpEd; other elem schools gained +5.9 to +9.9", fill=RESULT_FILL); r += 1

    put(ws, r, 1, "PES staffing gap vs enrollment share", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"={CDS}!B{R.CD_FTE_GAP}"
    dat(ws, r, 2, RESULT_FILL).number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(r, 2).font = RESULT_FONT
    put(ws, r, 3, "Negative = PES receives less staff than its enrollment share would warrant", fill=RESULT_FILL); r += 1

    r = note(ws, r, "CONCLUSION: PES per-student cost premium is NOT from overstaffing. PES runs leaner")
    r = note(ws, r, "than every other school and suffered the largest staffing decline of any school. The cost")
    r = note(ws, r, "premium exists purely because fixed costs (admin, facilities) are spread over fewer students.")
    r += 1

    # ── Key findings ──
    r = sec(ws, r, "KEY FINDINGS")
    findings = [
        "1. FY27 is a standalone budget vote. It does not include EC or restructuring costs.",
        "2. The proposed 6.53% FY27 increase is within the range of recent voter approvals.",
        "3. The FY27 budget already includes $1.72M in staff/program cuts (13 T1 + 7 T2 positions).",
        "4. EC mandate costs (~$690K/yr net) arrive in FY28, regardless of any restructuring decision.",
        "5. Path A and Path B share the same MS consolidation (grades 7-8 at FMS, 6 at DCS).",
        "6. The MARGINAL savings from closing PES (beyond what Path A achieves) are ~2.1 FTE.",
        "7. After transportation ($600K) and conversion ($74K/yr), Path B TRUE NET is approximately -$161K.",
        "8. The gross-lens analysis overstates Path B savings by ~$1.3M because it ignores absorption.",
        "9. RISK-ADJUSTED: Property value losses (~$39M at 10%), enrollment attrition, and transition",
        "   costs make Path B significantly worse. See C-RiskModel for full quantification.",
        "10. The FY28 baseline depends on which FY27 budget voters approve. See sensitivity matrix.",
        "11. Both paths require EC investment. The question is distributed (Path A) vs. centralized (Path B).",
        "12. Break-even analysis: Path B's one-time costs are NEVER recovered if marginal savings are negative.",
        "13. §1512 VETO likely does NOT apply to Option 2 (reconfiguration, not closure). [Fn 30][Fn 39]",
        "14. RSU WITHDRAWAL (§1466) is Pownal's most reliable recourse. 22-step process, 1-2 years. [Fn 32]",
        "15. Freeport attempted RSU 5 withdrawal in 2014; full agreement exists as template. [Fn 33]",
        "16. PES building would deed back to Pownal upon withdrawal (§12 of Freeport agreement). [Fn 33][Fn 34]",
        "17. EQUITY: Pownal pays the highest per-student local tax burden in RSU 5. [Fn 3][Fn 43]",
        "18. Pownal has a modest consumption deficit (~$267K) but overpays by ~$382K under the ALM formula. [Fn 3][Fn 8]",
        "19. Pownal's assessment grew ~30% over 3 years (FY24-FY27), outpacing RSU budget growth. [Fn 43][Fn 3]",
        "20. The 85/15 cost-sharing formula penalizes high-valuation towns disproportionately. [Fn 41][Fn 42]",
        "21. DOE DATA: PES suffered the largest 9-year staffing decline of any school (-28.4%). [Fn 45]",
        "22. DOE DATA: District-wide (central office) FTE grew +13.3% over the same period. [Fn 45]",
        "23. DOE DATA: PES share of district FTE fell from 6.5% to 3.9%, vs ~10.5% enrollment share. [Fn 45]",
        "24. DOE DATA: PES per-student cost premium is NOT from overstaffing; PES runs leaner than all peers. [Fn 45]",
    ]
    for f in findings:
        r = note(ws, r, f)


def build_sources(wb):
    ws = wb.create_sheet("Sources")
    ws.sheet_properties.tabColor = "808080"
    col_widths(ws, [10, 120])

    r = ttl(ws, 1, "Source Citations")
    r = note(ws, r + 1, "Every input cell in this workbook references one of these footnoted sources.")
    r += 1

    FOOTNOTES = {
        1:  "FY27 Budget Handbook, RSU 5 Superintendent's Office, pp.3-4 (02/11/2026). \"Projected Enrollment 2026-2027\" and Teacher/Class Size table.",
        2:  "FY27 Superintendent's Proposed Budget Articles, 89-page line-item detail (02/11/2026). URL: https://resources.finalsite.net/images/v1770852540/rsu5org/spzuwjskm8vh3t9bjck8/2026-2027SuperintendentsRecommendedBudgetArticlesforWebsite02112026.pdf",
        3:  "FY27 Budget Handbook, RSU 5, Budget Impact Summary pp.9-10 (02/11/2026). URL: https://resources.finalsite.net/images/v1770852522/rsu5org/ytm49nnwpbrzi4oiwg4i/FY27BUDGETHANDBOOK02112026.pdf",
        4:  "Pownal FY26 Real Estate Tax Commitment Book, committed 07/29/2025, rate $15.300/thousand. Source: Town of Pownal Assessor, https://www.pownalmaine.org/",
        5:  "Freeport FY26 tax rate $13.85/thousand at 100% assessment ratio, committed 09/15/2025. Source: Freeport Assessor, https://www.freeportmaine.com/158/Assessor",
        6:  "Freeport FY25 Budget Presentation, Proposed Tax Changes table. URL: https://www.freeportmaine.com/DocumentCenter/View/2711/FY-25-Budget-Presentation",
        7:  "Durham FY26 tax rate $33.58/thousand at ~53% assessment ratio, committed 08/12/2025. Source: Durham Assessor, https://durhammaine.gov/pages/assessing",
        8:  "Maine DOE FY25-26 Warrant Article F: Education Subsidy Information for Property Tax Bill (09/09/2025). URL: https://www.maine.gov/doe/sites/maine.gov.doe/files/inline-files/School%20Finance%20-%20FY25-26%20Warrant%20Article%20F%20Education%20Subsidy%20Information%20for%20Property%20Tax%20Bill%20-%209.9.2025.pdf",
        9:  "FY26 Adopted Budget, RSU 5, $44,455,929 total. Approved by voters 06/10/2025 (817 Yes, 494 No). Source: https://www.rsu5.org/budget/fy26",
        10: "RSU 5 Planning for the Future presentation (02/11/2026). URL: https://resources.finalsite.net/images/v1770852531/rsu5org/vfrsu75oj4ebaxvsbcwc/PlanningfortheFutureofRSU5.pdf",
        12: "Pownal tax breakdown (RSU 58.4%, County 3.3%, Town 38.3%) from Pownal municipal records, FY26.",
        13: "Census/ACS population estimates: Pownal 1,590; Durham 4,339; Freeport 8,771. Sources: censusreporter.org, maine-demographics.com (2023 est.)",
        14: "Maine CDC vital statistics / CDC WONDER. Cumberland County birth rate ~9.5/1,000; Androscoggin County similar.",
        15: "Maine DOE Public Pre-K Guidebook (01/21/2025). Chapter 124: max 16 students/classroom, 1:8 staff ratio. URL: https://www.maine.gov/doe/sites/maine.gov.doe/files/inline-files/Early%20Learning%20-%20Public%20Pre%20K%20Guidebook%20-%201.21.2025.pdf",
        16: "Pownal is in Cumberland County; Durham is in Androscoggin County; Freeport is in Cumberland County.",
        17: "Portland Press Herald, 10/02/2025, 'Freeport-area schools begin planning for new early childhood responsibilities.' By July 1, 2028, districts assume CDS responsibilities for FAPE ages 3-5. URL: https://www.pressherald.com/2025/10/02/freeport-area-schools-begin-planning-for-new-early-childhood-responsibilities/",
        18: "RSU 5 Early Childhood Transition Task Force membership and meeting schedule. URL: https://www.rsu5.org/quick-links/early-childhood-planning-cds",
        19: "RSU 5 Service Model Options PK3, presented at Dec 18, 2025 Task Force meeting. URL: https://www.rsu5.org/fs/resource-manager/view/da1426f4-8cc1-4b09-bd89-477da9c623ed",
        20: "Research on school consolidation and property values: Duncombe & Yinger (2010), Syracuse University CPR; EdWorkingPapers #22-530 (Arkansas, 2022). Findings: 5-15% property value decline in high-income areas post-closure.",
        21: "Maine Chapter 124 Section 9 - School Facilities for Public Preschool Programs. 35 sq ft/child, toilets within 40 feet, water source in classroom, natural light required. URL: https://regulations.justia.com/states/maine/05/071/chapter-124/section-071-124-9/",
        22: "Driving distances estimated from Google Maps between PES (587 Elmwood Rd, Pownal ME), DCS (654 Hallowell Rd, Durham ME), MSS (17 West St, Freeport ME).",
        24: "Patriquin Architects, 'Checklist for Early Childhood Conversion Projects.' URL: https://www.patriquinarchitects.com/checklist-for-early-childhood-conversion-projects/",
        25: "FY27 Superintendent's Proposed Budget Handbook (Revised 02/11/2026). Total operating budget $47,357,441. URL: https://www.rsu5.org/fs/resource-manager/view/78dce71e-a8e6-4435-b33f-9746e8541a3a",
        26: "Superintendent Tom Gray, 'A message to the Freeport-area schools community about the FY 2027 budget,' Portland Press Herald, 02/04/2026. Confirms three budget approaches under Board consideration. URL: https://www.pressherald.com/2026/02/04/a-message-to-the-freeport-area-schools-community-about-the-fy-2027-budget-column/",
        27: "U.S. Census Bureau, American Community Survey 5-Year Estimates (2019-2023), Table DP04. Pownal CDP, Cumberland County, ME: ~636 housing units, median home value ~$340,000. Source: data.census.gov",
        28: "Maine Revised Statutes, Title 20-A, Chapter 103: School Administrative Units. Sections 1461-1466 govern school closure procedures, including required public hearings, Board vote, and community notification. URL: https://legislature.maine.gov/statutes/20-A/title20-Ach103.pdf",
        29: "Title 20-A §4102: Closing of a school building. Specifies three conditions (replacement, condemnation, lack of need) and voter approval requirements. URL: https://www.mainelegislature.org/legis/statutes/20-A/title20-Asec4102.html",
        30: "Title 20-A §1512: Closing school in RSU. Subsection 4 defines 'school closing' as 'any action...that has the effect of providing no instruction for any students at that school.' Municipal referendum required. URL: https://legislature.maine.gov/legis/statutes/20-A/title20-Asec1512.html",
        31: "Title 20-A §1511: Supermajority vote to close school in RSU. Requires 2/3 of elected Board membership. URL: https://www.mainelegislature.org/legis/statutes/20-a/title20-Asec1511.html",
        32: "Title 20-A §1466: Withdrawal of a single municipality from a regional school unit. 22-step process; 30-month eligibility; 10% petition; Nov 30 referendum deadline. URL: https://www.mainelegislature.org/legis/statutes/20-a/title20-Asec1466.html",
        33: "RSU 5 Freeport Withdrawal Agreement (April 4, 2014). Full withdrawal agreement between RSU 5 and Town of Freeport Withdrawal Committee. Section 12: 'All real property and fixtures that the Town of Freeport deeded to RSU 5 upon creation of the RSU will be deeded back.' Vote failed 2,152-2,228. URL: https://resources.finalsite.net/images/v1657657653/rsu5org/a3niwxub5ppkhdhbiflk/fwcrevisedplan4-4-14.pdf",
        34: "Portland Press Herald, 06/24/2009: 'Freeport school properties turned over to RSU 5.' All three towns (Freeport, Pownal, Durham) deeded school properties to RSU 5 upon formation. URL: https://www.pressherald.com/2009/06/24/freeport-school-properties-turned-over-to-rsu-5/",
        35: "Maine DOE Withdrawal List: 15+ municipalities have successfully withdrawn from RSUs/SADs since FY2013, including Andover (SAD 44, 2015), Dayton (RSU 23, 2014), Athens (SAD 59, 2013), Belfast/Belmont (RSU 20, 2015). URL: https://www.maine.gov/doe/schools/structure/withdrawal/list",
        36: "Maine DOE Public School Secondary Tuition Rates (12/16/2024). State average secondary tuition: $14,080.88. URL: https://www.maine.gov/doe/funding/reports/tuition",
        37: "The Maine Monitor, 05/31/2025: 'What does it mean to withdraw from a Maine school district?' 22-step process explainer; Embden authorized $50K; Phillips authorized $50K. URL: https://themainemonitor.org/school-district-withdrawal-explainer/",
        38: "Portland Press Herald, 02/25/2026: 'SAD 75 abandons option to close Harpswell's only public school.' Community opposition led district to drop all closure options. URL: https://www.pressherald.com/2026/02/25/sad-75-abandons-option-to-close-harpswells-only-public-school/",
        39: "WAGM, 02/09/2026: 'MSAD #1 Plans Grade Reconfiguration for Next Year.' Pine Street Elementary converted to PreK-Gr1 as 'grade reconfiguration' (administrative action); Zippel Elementary closure handled separately via referendum. URL: https://www.wagmtv.com/2026/02/09/msad-1-plans-grade-reconfiguration-next-year/",
        40: "Title 20-A §4104: Proceeds from sale of school building. Properties transferred by member towns to RSU remain subject to specific ownership/disposition rules. URL: https://www.mainelegislature.org/legis/statutes/20-a/title20-Asec4104.html",
        41: "Portland Press Herald, 10/29/2019: 'RSU 5 looks at new cost-sharing formula.' Finance Committee proposed 85% state valuation / 15% enrollment for ALM. RSU 5 board review Nov 2019. URL: https://www.pressherald.com/2019/10/29/rsu-5-looks-at-new-cost-sharing-formula/",
        42: "Title 20-A §1481-A: Finances. RSU cost-sharing for additional operating costs. Subsection 3 permits alternate formulas in reorganization plans (valuation, enrollment, or combination). URL: https://www.mainelegislature.org/legis/statutes/20-a/title20-Asec1481-A.html",
        43: "Pownal 2025 Annual Report, Financial Audit for Year Ending June 30, 2024. Education expenditure: $3,427,607 (FY24 RSU 5 assessment). URL: https://www.pownalmaine.org/uploads/dm/3487/2025_Annual_Report",
        44: "RSU 5 FY26 Q&A Budget Document (revised 03/19/2025). Shows total FY26 taxation = $35,600,800; Pownal proposed RSU mil increase = $1.15/1000; FY25 mil rate $17.00 → FY26 $15.30 (revaluation). URL: https://www.rsu5.org/fs/resource-manager/view/86633cc3-a30d-4151-aaea-a35ff03d6803",
        45: "Maine DOE NEO Staff Historical Data (Dec 1, 2015-2025). Dataset: [Fed_Reporting].Reports.StaffHistoricalDec1, SAUOrgId=1449. Requested by Mark Dohle (03/04/2026), fulfilled by Maine DOE Data Team (Trevor R. Burns, 03/09/2026). DOE Helpdesk Issue #66636. File: 'MDohle RSU 5 Staff by FTE.xlsx'. Public use; FTE headcounts reported as of December 1 each year.",
    }

    for num, text in sorted(FOOTNOTES.items()):
        ws.cell(r, 1, f"[Fn {num}]").font = BOLD
        c = ws.cell(r, 2, text)
        c.alignment = Alignment(wrap_text=True)
        c.font = Font(size=10)
        r += 1


# ═══════════════════════════════════════════════════════════════
#  MAIN: BUILD ALL SHEETS, FIX CROSS-REFERENCES, REORDER, SAVE
# ═══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Build in dependency order: inputs first, then calcs
build_i_enrollment(wb)
build_i_budget(wb)
build_i_revenue(wb)
build_i_tax(wb)
build_i_ec_costs(wb)
build_i_assumptions(wb)
build_i_fy27_reductions(wb)
build_i_cost_growth(wb)
build_i_fte(wb)
build_i_doe_staffing(wb)

build_c_enrollment(wb)
build_c_budget(wb)
build_c_consumption(wb)
build_c_revenue(wb)
build_c_tax(wb)
build_c_cost_premium(wb)
build_c_ms_consolidation(wb)
build_c_ec_analysis(wb)
build_c_fy27_scenarios(wb)
build_c_fy28_paths(wb)      # must be before C-FY28Projection (provides savings refs)
build_c_fy28_projection(wb)
build_c_risk_model(wb)
build_i_withdrawal(wb)
build_i_equity(wb)
build_c_equity(wb)
build_c_independence(wb)
build_c_legal(wb)
build_c_doe_staffing(wb)
build_c_summary(wb)
build_sources(wb)

# Fix circular cross-references now that all row numbers are known
CFP = SN['cfp']
C28 = SN['c28']

# C-FY28Projection -> C-FY28Paths (savings refs)
ws28 = wb["C-FY28Projection"]
ws28.cell(R.C28_PA_SAV, 2).value = f"={CFP}!B{R.FP_PA_SAV}"
ws28.cell(R.C28_PB_SAV, 2).value = f"={CFP}!B{R.FP_PB_SAV}"

# C-FY28Paths -> C-FY28Projection (baseline and FY27 adopted refs in FINAL COMPARISON)
wfp = wb["C-FY28Paths"]
br = R.FP_FINAL_BUD
wfp.cell(br, 2).value = f"={C28}!B{R.C28_BASE}"
wfp.cell(br, 3).value = f"={C28}!B{R.C28_BASE}-B{R.FP_PA_SAV}"
wfp.cell(br, 4).value = f"={C28}!B{R.C28_BASE}-B{R.FP_PB_SAV}"
pr = R.FP_FINAL_PCT
for c in range(2, 5):
    cl = get_column_letter(c)
    wfp.cell(pr, c).value = f"=({cl}{br}-{C28}!B{R.C28_FY27_SEL})/{C28}!B{R.C28_FY27_SEL}"

# Reorder tabs: summary/results first, then detailed calcs, then inputs, then sources
desired_order = [
    "C-Summary", "C-DOEStaffing", "C-Equity", "C-Independence", "C-LegalAnalysis", "C-RiskModel",
    "C-FY28Projection", "C-FY27Scenarios", "C-FY28Paths",
    "C-Enrollment", "C-Budget", "C-Consumption", "C-Revenue",
    "C-Tax", "C-CostPremium", "C-MSConsol", "C-ECAnalysis",
    "I-Equity", "I-Withdrawal", "I-DOEStaffing", "I-Enrollment", "I-Budget", "I-Revenue", "I-Tax",
    "I-ECCosts", "I-Assumptions", "I-FY27Reductions", "I-CostGrowth", "I-FTE",
    "Sources",
]
sheet_map = {ws.title: ws for ws in wb.worksheets}
wb._sheets = [sheet_map[name] for name in desired_order]

# Save
out = "RSU 5 Financial Analysis.xlsx"
wb.save(out)
sheets = [ws.title for ws in wb.worksheets]
print(f"Saved to '{out}'")
print(f"  Summary:  {sheets[0]}")
print(f"  Calc:     {[s for s in sheets if s.startswith('C-')]}")
print(f"  Input:    {[s for s in sheets if s.startswith('I-')]}")
print(f"  Other:    {[s for s in sheets if not s.startswith(('I-','C-'))]}")
