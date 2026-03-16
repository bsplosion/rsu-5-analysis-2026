"""Shared Excel helper functions.

Extracted from ``create_excel.py`` lines 95-152.  These provide a
concise API for building formatted worksheets.
"""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from rsu5.excel.styles import (
    BOLD,
    HEADER_FILL,
    INPUT_FILL,
    NOTE,
    SECTION,
    SOURCE,
    SOURCE_FILL,
    THIN,
    TITLE,
)


def hdr(ws: Worksheet, row: int, cols: int) -> None:
    """Apply header styling to cells in *row* from column 1 to *cols*."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True)


def dat(ws: Worksheet, row: int, col: int, fill=None):
    """Style a data cell and return it."""
    cell = ws.cell(row=row, column=col)
    cell.border = THIN
    if fill:
        cell.fill = fill
    return cell


def ttl(ws: Worksheet, r: int, text: str) -> int:
    """Write a title row and return the next row number."""
    ws.cell(r, 1, text).font = TITLE
    return r + 1


def sec(ws: Worksheet, r: int, text: str) -> int:
    """Write a section heading and return the next row number."""
    ws.cell(r, 1, text).font = SECTION
    return r + 1


def source_block(ws: Worksheet, r: int, lines: list[str]) -> int:
    """Write source-citation lines and return the next row number."""
    for line in lines:
        cell = ws.cell(r, 1, line)
        cell.font = SOURCE
        cell.fill = SOURCE_FILL
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    return r


def note(ws: Worksheet, r: int, text: str) -> int:
    """Write a note row and return the next row number."""
    ws.cell(r, 1, text).font = NOTE
    return r + 1


def put(ws: Worksheet, r: int, c: int, val, fmt=None,
        fill=INPUT_FILL, font=None):
    """Write a value with optional formatting and return the cell."""
    cell = ws.cell(r, c, val)
    if fmt:
        cell.number_format = fmt
    if fill is not None:
        cell.fill = fill
    cell.border = THIN
    if font:
        cell.font = font
    return cell


def col_widths(ws: Worksheet, widths: list[int | float]) -> None:
    """Set column widths from a list (1-indexed)."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
