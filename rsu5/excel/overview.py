"""Build RSU5_Overview.xlsx -- cross-year trend workbook.

Tabs:
  - Budget History: 10-year adopted budget trend
  - Enrollment: School enrollment trends
  - Cost Per Student: Per-school cost trends
  - Staffing: DOE FTE trends
  - Article Growth: Article-level growth rates across FYs
"""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

from rsu5.config import cfg
from rsu5.excel.helpers import col_widths, hdr, note, put, sec, source_block, ttl
from rsu5.excel.styles import (
    BOLD,
    CALC_FILL,
    HEADER_FILL,
    INPUT_FILL,
    RESULT_FILL,
    RESULT_FONT,
    TAB_CALC,
    TAB_INPUT,
    TAB_SUMMARY,
    THICK_BOTTOM,
    USD,
    PCT2,
    USD2,
)
from rsu5.ingest.data_loader import BudgetData
from rsu5.model import VerifiedBaseline


def _build_budget_history_tab(wb: Workbook, data: BudgetData) -> None:
    """10-year adopted budget trend from handbook data."""
    ws = wb.create_sheet("Budget History")
    ws.sheet_properties.tabColor = TAB_INPUT
    col_widths(ws, [16, 20, 18, 14])

    r = ttl(ws, 1, "RSU 5: 10-Year Adopted Budget History")
    r = note(ws, r + 1, "Source: Superintendent's Budget Handbook data")
    r += 1

    history = data.budget_history(27) or data.budget_history(26)
    if not history:
        r = note(ws, r, "No budget history data available.")
        return

    headers = ["Fiscal Year", "Adopted Budget", "Year-over-Year Change", "% Change"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    for entry in history:
        put(ws, r, 1, f"FY{entry.fy}", fill=INPUT_FILL)
        put(ws, r, 2, entry.adopted, fmt=USD, fill=INPUT_FILL)
        put(ws, r, 3, entry.difference, fmt=USD, fill=INPUT_FILL)
        pct = entry.pct_increase / 100.0 if abs(entry.pct_increase) > 1 else entry.pct_increase
        put(ws, r, 4, pct, fmt=PCT2, fill=INPUT_FILL)
        r += 1

    # Add FY27 proposed
    scenarios = cfg.raw.get("fy27_scenarios", {})
    fy27 = scenarios.get("superintendent_proposed", 47357441)
    fy26 = scenarios.get("fy26_adopted", 44455929)
    put(ws, r, 1, "FY27*", fill=RESULT_FILL, font=BOLD)
    put(ws, r, 2, fy27, fmt=USD, fill=RESULT_FILL, font=BOLD)
    put(ws, r, 3, fy27 - fy26, fmt=USD, fill=RESULT_FILL)
    put(ws, r, 4, (fy27 - fy26) / fy26, fmt=PCT2, fill=RESULT_FILL)
    r += 1
    r = note(ws, r, "* FY27 = Superintendent's proposed (not yet adopted)")
    r += 1

    # Projections
    r = sec(ws, r, "Forward Projections (Superintendent)")
    fy28 = cfg.raw.get("fy28_projections", {})
    fy29 = cfg.raw.get("fy29_projections", {})
    for label, val in [
        ("FY28 (w/cuts)", fy28.get("with_reductions", 0)),
        ("FY28 (no cuts)", fy28.get("no_reductions", 0)),
        ("FY29 (w/cuts)", fy29.get("with_reductions", 0)),
        ("FY29 (no cuts)", fy29.get("no_reductions", 0)),
    ]:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, val, fmt=USD, fill=CALC_FILL)
        r += 1


def _build_enrollment_tab(wb: Workbook, data: BudgetData) -> None:
    """Enrollment trend by school."""
    ws = wb.create_sheet("Enrollment")
    ws.sheet_properties.tabColor = TAB_INPUT

    enrollment = data.enrollment(27)
    if not enrollment:
        col_widths(ws, [30])
        ttl(ws, 1, "Enrollment")
        note(ws, 3, "No enrollment data available.")
        return

    all_years = sorted({y for e in enrollment for y in e.years})
    col_widths(ws, [28] + [12] * len(all_years) + [14])

    r = ttl(ws, 1, "RSU 5: Enrollment by School")
    r = note(ws, r + 1, "Source: Superintendent's Budget Handbook (October 1 counts)")
    r += 1

    headers = ["School"] + [str(y) for y in all_years] + ["Change"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for entry in enrollment:
        if entry.school == "Grand Total":
            continue
        put(ws, r, 1, entry.school, fill=INPUT_FILL)
        for j, y in enumerate(all_years, 2):
            val = entry.years.get(y)
            if val:
                put(ws, r, j, val, fill=INPUT_FILL)
        if len(all_years) >= 2:
            first_cl = get_column_letter(2)
            last_cl = get_column_letter(1 + len(all_years))
            change_col = 2 + len(all_years)
            ws.cell(r, change_col).value = f"=IF({first_cl}{r}=0,0,{last_cl}{r}/{first_cl}{r}-1)"
            ws.cell(r, change_col).number_format = PCT2
        r += 1

    # Total row
    put(ws, r, 1, "TOTAL", fill=RESULT_FILL, font=BOLD)
    for j in range(2, 2 + len(all_years) + 1):
        cl = get_column_letter(j)
        start_r = r - len([e for e in enrollment if e.school != "Grand Total"])
        ws.cell(r, j).value = f"=SUM({cl}{start_r}:{cl}{r-1})"
        ws.cell(r, j).number_format = USD if j <= 1 + len(all_years) else PCT2
        ws.cell(r, j).font = BOLD
        ws.cell(r, j).fill = RESULT_FILL


def _build_staffing_tab(wb: Workbook, data: BudgetData) -> None:
    """DOE staffing FTE trends."""
    ws = wb.create_sheet("Staffing FTE")
    ws.sheet_properties.tabColor = TAB_CALC

    years = data.staffing_years()
    if not years:
        col_widths(ws, [30])
        ttl(ws, 1, "Staffing FTE")
        note(ws, 3, "No DOE staffing data available.")
        return

    schools = ["DCS", "MSS", "MLS", "PES", "FMS", "FHS", "District"]
    col_widths(ws, [20] + [12] * len(years) + [14])

    r = ttl(ws, 1, "RSU 5: DOE Staffing FTE Trends")
    r = note(ws, r + 1, "Source: Maine DOE Staff Historical Data (December 1 snapshots)")
    r += 1

    headers = ["School"] + [str(y) for y in years] + ["Change"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for school in schools:
        put(ws, r, 1, school, fill=INPUT_FILL)
        for j, y in enumerate(years, 2):
            fte = data.school_fte(y, school)
            if fte > 0:
                put(ws, r, j, fte, fmt=USD2, fill=INPUT_FILL)
        if len(years) >= 2:
            first_cl = get_column_letter(2)
            last_cl = get_column_letter(1 + len(years))
            change_col = 2 + len(years)
            ws.cell(r, change_col).value = f"=IF({first_cl}{r}=0,0,{last_cl}{r}/{first_cl}{r}-1)"
            ws.cell(r, change_col).number_format = PCT2
        r += 1

    put(ws, r, 1, "TOTAL", fill=RESULT_FILL, font=BOLD)
    for j, y in enumerate(years, 2):
        put(ws, r, j, data.total_fte(y), fmt=USD2, fill=RESULT_FILL, font=BOLD)
    if len(years) >= 2:
        first_cl = get_column_letter(2)
        last_cl = get_column_letter(1 + len(years))
        change_col = 2 + len(years)
        ws.cell(r, change_col).value = f"=IF({first_cl}{r}=0,0,{last_cl}{r}/{first_cl}{r}-1)"
        ws.cell(r, change_col).number_format = PCT2
        ws.cell(r, change_col).font = BOLD
        ws.cell(r, change_col).fill = RESULT_FILL


def _build_article_growth_tab(wb: Workbook, data: BudgetData,
                                baselines: dict[int, VerifiedBaseline]) -> None:
    """Article growth rates across all FYs."""
    ws = wb.create_sheet("Article Growth")
    ws.sheet_properties.tabColor = TAB_CALC

    fys = sorted(baselines.keys())
    if not fys:
        col_widths(ws, [30])
        ttl(ws, 1, "Article Growth")
        note(ws, 3, "No baseline data available.")
        return

    col_widths(ws, [32] + [16] * len(fys) + [14, 14])

    r = ttl(ws, 1, "RSU 5: Article Growth Rates Across All FYs")
    r += 1

    fy_col_names: dict[int, str] = {}
    for fy in fys:
        bl = baselines[fy]
        for li in bl.line_items[:5]:
            for col_name in li.amounts:
                if f"FY{fy}" in col_name:
                    fy_col_names[fy] = col_name
                    break
            if fy in fy_col_names:
                break

    headers = ["Article"] + [f"FY{fy}" for fy in fys] + ["Total Growth", "CAGR"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for art_num in sorted(cfg.articles):
        art_info = cfg.articles[art_num]
        put(ws, r, 1, f"Art {art_num} - {art_info.name}", fill=None)

        for j, fy in enumerate(fys, 2):
            bl = baselines[fy]
            col_name = fy_col_names.get(fy, "")
            total = bl.article_total(art_num, col_name) if col_name else 0
            if total:
                put(ws, r, j, total, fmt=USD, fill=None)

        n = len(fys)
        if n >= 2:
            first_cl = get_column_letter(2)
            last_cl = get_column_letter(1 + n)
            gc = 2 + n
            ws.cell(r, gc).value = f"=IF({first_cl}{r}=0,0,{last_cl}{r}/{first_cl}{r}-1)"
            ws.cell(r, gc).number_format = PCT2
            cc = gc + 1
            ws.cell(r, cc).value = f"=IF({first_cl}{r}=0,0,({last_cl}{r}/{first_cl}{r})^(1/{n-1})-1)"
            ws.cell(r, cc).number_format = PCT2
        r += 1

    # Grand total
    put(ws, r, 1, "TOTAL", fill=RESULT_FILL, font=RESULT_FONT)
    for j, fy in enumerate(fys, 2):
        bl = baselines[fy]
        col_name = fy_col_names.get(fy, "")
        total = sum(bl.article_total(a, col_name) for a in cfg.articles) if col_name else 0
        put(ws, r, j, total, fmt=USD, fill=RESULT_FILL, font=RESULT_FONT)


def build_overview_workbook(
    data: BudgetData,
    baselines: dict[int, VerifiedBaseline],
    output_path: Path | None = None,
) -> Path:
    """Build the cross-year overview workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    _build_budget_history_tab(wb, data)
    _build_enrollment_tab(wb, data)
    _build_staffing_tab(wb, data)
    _build_article_growth_tab(wb, data, baselines)

    if output_path is None:
        output_path = Path(__file__).parent.parent.parent / "RSU5_Overview.xlsx"

    wb.save(str(output_path))
    return output_path
