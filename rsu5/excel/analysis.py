"""Data-driven analytical sheets built from the verified baseline.

Sheets built here:
  - I-BudgetData: Budget by article x cost center from parsed line items
  - I-Enrollment: School enrollment from handbook data
  - I-DOEStaffing: Staffing FTE from Maine DOE data
  - I-CostGrowth: 10-year adopted budget history from handbooks
  - C-GrowthTrends: Multi-year budget growth by article
  - C-CostPerStudent: Per-school cost-per-student analysis
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

from rsu5.config import cfg
from rsu5.excel.helpers import col_widths, dat, hdr, note, put, sec, source_block, ttl
from rsu5.excel.styles import (
    BOLD,
    CALC_FILL,
    HEADER_FILL,
    INPUT_FILL,
    RESULT_FILL,
    RESULT_FONT,
    TAB_CALC,
    TAB_INPUT,
    THICK_BOTTOM,
    USD,
    USD2,
    PCT2,
    WARN_FONT,
)
from rsu5.ingest.data_loader import BudgetData
from rsu5.model import VerifiedBaseline


def build_i_budget_data(
    wb: Workbook,
    baseline: VerifiedBaseline,
) -> None:
    """Input sheet: article x cost-center totals from verified baseline.

    This is the data-driven replacement for the hardcoded I-Budget sheet
    in create_excel.py.  Downstream calc sheets reference this.
    """
    ws = wb.create_sheet("I-BudgetData")
    ws.sheet_properties.tabColor = TAB_INPUT

    # Determine the "current year" column
    fy = baseline.fy
    current_col = ""
    for li in baseline.line_items[:10]:
        for col_name in li.amounts:
            if f"FY{fy}" in col_name:
                current_col = col_name
                break
        if current_col:
            break

    if not current_col:
        current_col = f"FY{fy} Proposed"

    # Schools in RSU order
    school_codes = ["010", "020", "030", "040", "050", "300", "900", "950", "990", "000"]
    school_labels = [
        cfg.cost_centers.get(c, type("", (), {"abbrev": c})()).abbrev
        for c in school_codes
    ]

    col_widths(ws, [35] + [14] * len(school_codes) + [16])

    r = ttl(ws, 1, f"INPUT: FY{fy} Budget by Article x Cost Center")
    status = "VERIFIED" if baseline.is_clean else "UNRECONCILED"
    r = source_block(ws, r, [
        f"Source: Parsed from budget documents (doc_type={baseline.doc_type})",
        f"Baseline status: {status}",
        f"Column shown: {current_col}",
    ])
    r += 1

    # Header row
    heads = ["Article"] + school_labels + ["Total"]
    for i, h in enumerate(heads, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(heads))
    r += 1

    # Article rows
    art_start = r
    for art_num in sorted(cfg.articles):
        art_info = cfg.articles[art_num]
        ws.cell(r, 1, f"Art {art_num} - {art_info.name}").border = THICK_BOTTOM

        for j, cc in enumerate(school_codes, 2):
            total = baseline.cost_center_total(cc, current_col, article=art_num)
            if total != 0:
                put(ws, r, j, total, USD, fill=INPUT_FILL)
            else:
                ws.cell(r, j, 0).border = THICK_BOTTOM

        # Total formula
        total_col = 2 + len(school_codes)
        first_cl = get_column_letter(2)
        last_cl = get_column_letter(1 + len(school_codes))
        ws.cell(r, total_col).value = f"=SUM({first_cl}{r}:{last_cl}{r})"
        ws.cell(r, total_col).number_format = USD
        ws.cell(r, total_col).font = BOLD
        ws.cell(r, total_col).fill = CALC_FILL
        ws.cell(r, total_col).border = THICK_BOTTOM
        r += 1

    art_end = r - 1

    # Grand total row
    ws.cell(r, 1, "GRAND TOTAL").font = RESULT_FONT
    ws.cell(r, 1).fill = RESULT_FILL
    ws.cell(r, 1).border = THICK_BOTTOM
    for j in range(2, 2 + len(school_codes) + 1):
        cl = get_column_letter(j)
        ws.cell(r, j).value = f"=SUM({cl}{art_start}:{cl}{art_end})"
        ws.cell(r, j).number_format = USD
        ws.cell(r, j).font = RESULT_FONT
        ws.cell(r, j).fill = RESULT_FILL
        ws.cell(r, j).border = THICK_BOTTOM
    grand_total_row = r
    r += 2

    # ── Per-student cost section ──
    r = sec(ws, r, "Per-Student Direct Cost")
    r = note(ws, r, "Enrollment from FY27 Budget Handbook (Oct 2026 projected)")

    heads = ["School"] + school_labels[:6]  # only actual schools
    for i, h in enumerate(heads, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(heads))
    r += 1

    # Enrollment row
    ws.cell(r, 1, "Enrollment (Oct 2026P)").font = BOLD
    ws.cell(r, 1).border = THICK_BOTTOM
    enrollment_row = r
    enroll_by_abbrev = cfg.enrollment.get("by_school", {})
    for j, cc in enumerate(school_codes[:6], 2):
        abbrev = cfg.cost_centers.get(cc, type("", (), {"abbrev": ""})()).abbrev
        enroll = enroll_by_abbrev.get(abbrev, {}).get("oct_2026p", 0)
        if enroll:
            put(ws, r, j, enroll, USD, fill=INPUT_FILL)
    r += 1

    # Direct cost row (SUM of articles for this school)
    ws.cell(r, 1, "Direct Cost").font = BOLD
    ws.cell(r, 1).border = THICK_BOTTOM
    direct_row = r
    for j in range(2, 8):
        cl = get_column_letter(j)
        ws.cell(r, j).value = f"=SUM({cl}{art_start}:{cl}{art_end})"
        ws.cell(r, j).number_format = USD
        ws.cell(r, j).fill = CALC_FILL
        ws.cell(r, j).border = THICK_BOTTOM
    r += 1

    # Per-student
    ws.cell(r, 1, "Cost per Student").font = RESULT_FONT
    ws.cell(r, 1).fill = RESULT_FILL
    ws.cell(r, 1).border = THICK_BOTTOM
    for j in range(2, 8):
        cl = get_column_letter(j)
        ws.cell(r, j).value = (
            f"=IF({cl}{enrollment_row}=0,0,{cl}{direct_row}/{cl}{enrollment_row})"
        )
        ws.cell(r, j).number_format = USD
        ws.cell(r, j).font = RESULT_FONT
        ws.cell(r, j).fill = RESULT_FILL
        ws.cell(r, j).border = THICK_BOTTOM


def build_c_growth_trends(
    wb: Workbook,
    baselines: dict[int, VerifiedBaseline],
) -> None:
    """Calc sheet: multi-year budget growth by article."""
    ws = wb.create_sheet("C-GrowthTrends")
    ws.sheet_properties.tabColor = TAB_CALC

    fys = sorted(baselines.keys())
    n_fys = len(fys)
    col_widths(ws, [35] + [16] * n_fys + [14, 14])

    r = ttl(ws, 1, "CALC: Budget Growth Trends by Article")
    r = source_block(ws, r, [
        "Article totals computed independently from parsed line-item data.",
        "Growth and CAGR computed via Excel formulas.",
    ])
    r += 1

    # Determine column names per FY
    fy_col_names: dict[int, str] = {}
    for fy_val in fys:
        bl = baselines[fy_val]
        for li in bl.line_items[:5]:
            for col_name in li.amounts:
                if f"FY{fy_val}" in col_name:
                    fy_col_names[fy_val] = col_name
                    break
            if fy_val in fy_col_names:
                break

    headers = ["Article"] + [f"FY{fy_val}" for fy_val in fys] + ["Total Growth", "Avg Annual"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    art_rows: list[int] = []
    for art_num in sorted(cfg.articles):
        art_info = cfg.articles[art_num]
        ws.cell(r, 1, f"Art {art_num} - {art_info.name}").border = THICK_BOTTOM

        for j, fy_val in enumerate(fys, 2):
            bl = baselines[fy_val]
            col_name = fy_col_names.get(fy_val, "")
            total = bl.article_total(art_num, col_name) if col_name else 0
            if total:
                put(ws, r, j, total, USD, fill=None)
            else:
                ws.cell(r, j).border = THICK_BOTTOM

        # Growth formulas
        if n_fys >= 2:
            first_cl = get_column_letter(2)
            last_cl = get_column_letter(1 + n_fys)
            growth_col = 2 + n_fys
            ws.cell(r, growth_col).value = (
                f"=IF({first_cl}{r}=0,0,{last_cl}{r}/{first_cl}{r}-1)"
            )
            ws.cell(r, growth_col).number_format = PCT2
            ws.cell(r, growth_col).border = THICK_BOTTOM

            cagr_col = growth_col + 1
            ws.cell(r, cagr_col).value = (
                f"=IF({first_cl}{r}=0,0,"
                f"({last_cl}{r}/{first_cl}{r})^(1/{n_fys - 1})-1)"
            )
            ws.cell(r, cagr_col).number_format = PCT2
            ws.cell(r, cagr_col).border = THICK_BOTTOM

        art_rows.append(r)
        r += 1

    # Grand total
    ws.cell(r, 1, "TOTAL EXPENDITURES").font = RESULT_FONT
    ws.cell(r, 1).fill = RESULT_FILL
    ws.cell(r, 1).border = THICK_BOTTOM
    for j in range(2, 2 + n_fys + 2):
        cl = get_column_letter(j)
        if j < 2 + n_fys:
            refs = "+".join(f"{cl}{row}" for row in art_rows)
            ws.cell(r, j).value = f"={refs}"
            ws.cell(r, j).number_format = USD
        elif j == 2 + n_fys:
            first_cl = get_column_letter(2)
            last_cl = get_column_letter(1 + n_fys)
            ws.cell(r, j).value = (
                f"=IF({first_cl}{r}=0,0,{last_cl}{r}/{first_cl}{r}-1)"
            )
            ws.cell(r, j).number_format = PCT2
        else:
            first_cl = get_column_letter(2)
            last_cl = get_column_letter(1 + n_fys)
            ws.cell(r, j).value = (
                f"=IF({first_cl}{r}=0,0,"
                f"({last_cl}{r}/{first_cl}{r})^(1/{n_fys - 1})-1)"
            )
            ws.cell(r, j).number_format = PCT2
        ws.cell(r, j).font = RESULT_FONT
        ws.cell(r, j).fill = RESULT_FILL
        ws.cell(r, j).border = THICK_BOTTOM
    r += 2

    # Year-over-year growth rate section
    r = sec(ws, r, "Year-over-Year Growth Rate")
    yoy_headers = ["Article"] + [f"FY{fy_val}" for fy_val in fys[1:]]
    for i, h in enumerate(yoy_headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(yoy_headers))
    r += 1

    for i_art, art_num in enumerate(sorted(cfg.articles)):
        art_info = cfg.articles[art_num]
        ws.cell(r, 1, f"Art {art_num}").border = THICK_BOTTOM
        art_data_row = art_rows[i_art]

        for j_fy in range(1, n_fys):
            prev_cl = get_column_letter(1 + j_fy)
            curr_cl = get_column_letter(2 + j_fy)
            ws.cell(r, 1 + j_fy).value = (
                f"=IF({prev_cl}{art_data_row}=0,0,"
                f"{curr_cl}{art_data_row}/{prev_cl}{art_data_row}-1)"
            )
            ws.cell(r, 1 + j_fy).number_format = PCT2
            ws.cell(r, 1 + j_fy).border = THICK_BOTTOM
        r += 1


def build_i_enrollment(wb: Workbook, fy: int, data: BudgetData) -> None:
    """Input sheet: school enrollment from handbook data."""
    ws = wb.create_sheet("I-Enrollment")
    ws.sheet_properties.tabColor = TAB_INPUT
    col_widths(ws, [28] + [12] * 8)

    enrollment = data.enrollment(fy)

    r = ttl(ws, 1, f"INPUT: Enrollment Data (FY{fy} Handbook)")
    r += 1

    if not enrollment:
        r = note(ws, r, f"No enrollment data available from FY{fy} handbook.")
        return

    years = sorted({y for e in enrollment for y in e.years})

    headers = ["School"] + [str(y) for y in years]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for entry in enrollment:
        if entry.school == "Grand Total":
            continue
        put(ws, r, 1, entry.school, fill=INPUT_FILL)
        for j, y in enumerate(years, 2):
            val = entry.years.get(y)
            if val:
                put(ws, r, j, val, fill=INPUT_FILL)
        r += 1

    total_entry = next((e for e in enrollment if e.school == "Grand Total"), None)
    if total_entry:
        put(ws, r, 1, "Grand Total", fill=RESULT_FILL, font=BOLD)
        for j, y in enumerate(years, 2):
            val = total_entry.years.get(y)
            if val:
                put(ws, r, j, val, fill=RESULT_FILL, font=BOLD)
        r += 1
    r += 1

    r = source_block(ws, r, [
        "Source: RSU 5 Superintendent's Budget Handbook",
        "October 1 enrollment counts (projected for future years)",
    ])


def build_i_doe_staffing(wb: Workbook, fy: int, data: BudgetData) -> None:
    """Input sheet: DOE staffing FTE data."""
    ws = wb.create_sheet("I-DOEStaffing")
    ws.sheet_properties.tabColor = TAB_INPUT
    col_widths(ws, [32] + [12] * 8)

    r = ttl(ws, 1, f"INPUT: DOE Staffing Data (Dec 1 Snapshot)")
    r += 1

    cal_year = 2000 + fy
    years = sorted(data.staffing_years())
    if not years:
        r = note(ws, r, "No DOE staffing data available.")
        return

    show_years = [y for y in years if y <= cal_year][-5:] or years[-5:]

    schools = ["DCS", "MSS", "MLS", "PES", "FMS", "FHS", "District"]

    r = sec(ws, r, "Total FTE by School")
    headers = ["School"] + [str(y) for y in show_years]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for school in schools:
        put(ws, r, 1, school, fill=INPUT_FILL)
        for j, y in enumerate(show_years, 2):
            fte = data.school_fte(y, school)
            if fte > 0:
                put(ws, r, j, fte, fmt=USD2, fill=INPUT_FILL)
        r += 1

    put(ws, r, 1, "Total", fill=RESULT_FILL, font=BOLD)
    for j, y in enumerate(show_years, 2):
        put(ws, r, j, data.total_fte(y), fmt=USD2, fill=RESULT_FILL, font=BOLD)
    r += 2

    r = sec(ws, r, f"Detail by Position ({show_years[-1]})")
    latest_year = show_years[-1]
    staff = data.staffing_for_year(latest_year)

    by_cat: dict[str, float] = {}
    for rec in staff:
        by_cat[rec.category] = by_cat.get(rec.category, 0.0) + rec.fte

    headers2 = ["Position", "Total FTE"]
    for i, h in enumerate(headers2, 1):
        ws.cell(r, i, h)
    hdr(ws, r, 2)
    r += 1

    for cat, fte in sorted(by_cat.items(), key=lambda x: -x[1]):
        if fte > 0:
            put(ws, r, 1, cat, fill=None)
            put(ws, r, 2, fte, fmt=USD2, fill=None)
            r += 1
    r += 1

    r = source_block(ws, r, [
        "Source: Maine DOE Staff Historical Data (Dec 1 snapshots)",
        "MDohle RSU 5 Staff by FTE.xlsx",
    ])


def build_i_cost_growth(wb: Workbook, fy: int, data: BudgetData) -> None:
    """Input sheet: 10-year adopted budget history from handbooks."""
    ws = wb.create_sheet("I-CostGrowth")
    ws.sheet_properties.tabColor = TAB_INPUT
    col_widths(ws, [16, 20, 18, 14])

    r = ttl(ws, 1, f"INPUT: Budget History (from FY{fy} Handbook)")
    r += 1

    history = data.budget_history(fy)
    if not history:
        for alt_fy in range(27, 21, -1):
            history = data.budget_history(alt_fy)
            if history:
                r = note(ws, r, f"Using history from FY{alt_fy} handbook (FY{fy} had none)")
                break

    if not history:
        r = note(ws, r, "No budget history data available.")
        return

    headers = ["Fiscal Year", "Adopted Budget", "Difference", "% Increase"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for entry in history:
        put(ws, r, 1, f"FY{entry.fy}", fill=INPUT_FILL)
        put(ws, r, 2, entry.adopted, fmt=USD, fill=INPUT_FILL)
        put(ws, r, 3, entry.difference, fmt=USD, fill=INPUT_FILL)
        put(ws, r, 4, entry.pct_increase / 100.0 if abs(entry.pct_increase) > 1 else entry.pct_increase,
            fmt=PCT2, fill=INPUT_FILL)
        r += 1
    r += 1

    r = source_block(ws, r, [
        "Source: RSU 5 Superintendent's Budget Handbook",
        "10-Year Adopted Budget History table",
    ])


def build_c_cost_per_student(wb: Workbook, fy: int, data: BudgetData,
                              baseline: VerifiedBaseline | None = None) -> None:
    """Calc sheet: per-school cost-per-student analysis."""
    ws = wb.create_sheet("C-CostPerStudent")
    ws.sheet_properties.tabColor = TAB_CALC
    col_widths(ws, [28, 18, 14, 18])

    r = ttl(ws, 1, f"CALC: Cost Per Student (FY{fy})")
    r += 1

    enrollment = data.enrollment(fy)
    if not enrollment:
        r = note(ws, r, "No enrollment data available for cost-per-student analysis.")
        return

    columns = data.all_columns(fy)
    proposed_col = None
    for c in columns:
        if "proposed" in c.lower() or f"fy{fy}" in c.lower():
            proposed_col = c
            break
    if not proposed_col and columns:
        proposed_col = columns[-1]

    school_cc_map = {
        "Morse Street": "010",
        "Mast Landing": "020",
        "Pownal Elementary": "030",
        "Durham Community": "040",
        "Freeport Middle": "050",
        "Freeport High": "300",
    }

    headers = ["School", "Budget", "Enrollment", "Cost/Student"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    r += 1

    for entry in enrollment:
        if entry.school == "Grand Total":
            continue
        cc = school_cc_map.get(entry.school)
        if not cc:
            continue

        latest_year = max(entry.years.keys()) if entry.years else 0
        count = entry.years.get(latest_year, 0)

        budget = 0.0
        if proposed_col and baseline:
            budget = baseline.cost_center_total(cc, proposed_col)
        elif proposed_col:
            budget = data.cost_center_total(fy, cc, proposed_col)

        per_student = budget / count if count > 0 else 0

        put(ws, r, 1, entry.school, fill=CALC_FILL)
        put(ws, r, 2, budget, fmt=USD, fill=CALC_FILL)
        put(ws, r, 3, count, fill=CALC_FILL)
        put(ws, r, 4, per_student, fmt=USD, fill=RESULT_FILL,
            font=RESULT_FONT if per_student > 15000 else BOLD)
        r += 1

    r += 1
    r = source_block(ws, r, [
        f"Budget data from FY{fy} parsed line items ({proposed_col or 'N/A'})",
        "Enrollment from handbook October counts",
    ])


def build_analysis_sheets(
    wb: Workbook,
    fy: int,
    data: BudgetData,
    baseline: VerifiedBaseline | None = None,
    all_baselines: dict[int, VerifiedBaseline] | None = None,
) -> None:
    """Build all data-driven analysis sheets for a single FY workbook."""
    if baseline:
        build_i_budget_data(wb, baseline)

    build_i_enrollment(wb, fy, data)
    build_i_doe_staffing(wb, fy, data)
    build_i_cost_growth(wb, fy, data)

    if all_baselines:
        build_c_growth_trends(wb, all_baselines)

    build_c_cost_per_student(wb, fy, data, baseline)
