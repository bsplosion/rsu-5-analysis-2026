"""Generate article-by-article Excel tabs in RSU 5's native layout.

Each tab mirrors the format board members and administrators see in
the official budget articles documents:

    Cost Center: 010 - DURHAM COMMUNITY
      Program: 1100 - ELEMENTARY PROGRAMS
        Account | Description | FY23 Actual | ... | FY27 Proposed | $ Diff | % Diff
        1000.1100.1000.51010.010 | Teacher Salary | ... | ... | ...
        ...
        Program Total (SUM formula)
      Function Total (SUM formula)
    Cost Center Total (SUM formula)

All totals are computed independently via Excel SUM formulas so the
workbook itself is the verification artifact.
"""

from __future__ import annotations

from collections import defaultdict

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

from rsu5.config import cfg
from rsu5.excel.helpers import col_widths, hdr, put, sec, source_block, ttl
from rsu5.excel.styles import (
    BOLD,
    CALC_FILL,
    HEADER_FILL,
    RESULT_FILL,
    RESULT_FONT,
    SUMMARY_FILL,
    TAB_RSU,
    THIN,
    THICK_BOTTOM,
    USD2,
    PCT2,
)
from rsu5.model import BudgetLineItem, VerifiedBaseline


def _group_by(
    items: list[BudgetLineItem], *keys: str
) -> dict[tuple, list[BudgetLineItem]]:
    """Group items by one or more attribute names."""
    groups: dict[tuple, list[BudgetLineItem]] = defaultdict(list)
    for li in items:
        key = tuple(getattr(li, k) for k in keys)
        groups[key].append(li)
    return groups


def _amount_columns(baseline: VerifiedBaseline) -> list[str]:
    """Return the year-amount columns (excluding Diff columns)."""
    cols: set[str] = set()
    for li in baseline.line_items:
        cols.update(li.amounts.keys())
    # Filter to FY columns only, in order
    fy_cols = sorted(
        [c for c in cols if c.startswith("FY")],
        key=lambda c: c,
    )
    return fy_cols


def build_article_sheet(
    wb: Workbook,
    baseline: VerifiedBaseline,
    article_num: int,
) -> None:
    """Create one Excel tab for a budget article in RSU format."""
    article_info = cfg.articles.get(article_num)
    if article_info is None:
        return

    art_items = [
        li for li in baseline.line_items if li.article == article_num
    ]
    if not art_items:
        return

    tab_name = f"Art {article_num}"
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = TAB_RSU

    # Determine columns
    fy_cols = _amount_columns(baseline)
    all_cols = fy_cols  # We'll compute diff ourselves
    n_data_cols = len(all_cols)
    total_cols = 2 + n_data_cols + 2  # Account, Description, FY cols, $Diff, %Diff

    # Column widths
    widths = [28, 30] + [15] * n_data_cols + [15, 10]
    col_widths(ws, widths)

    # Title
    r = ttl(ws, 1, f"Article {article_num} - {article_info.name}")
    r = source_block(ws, r, [
        f"Source: FY{baseline.fy} {'Adopted' if baseline.doc_type == 'adopted' else 'Proposed'} Budget, RSU 5",
        f"Verification: {'RECONCILED' if baseline.is_clean else 'UNRECONCILED -- see Summary tab'}",
    ])
    r += 1

    # Column headers
    headers = ["Account", "Description"] + all_cols + ["$ Diff", "% Diff"]
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h)
    hdr(ws, r, len(headers))
    header_row = r
    r += 1

    # Group items: cost_center -> function -> program -> line items
    by_cc = _group_by(art_items, "cost_center")

    # Track row ranges for grand total
    grand_total_rows: list[int] = []

    for cc_code in sorted(by_cc.keys()):
        cc_items = by_cc[cc_code]
        cc_info = cfg.cost_centers.get(cc_code, None)
        cc_label = f"{cc_info.name} - {cc_code}" if cc_info else cc_code

        # Cost Center header
        r = sec(ws, r, f"Cost Center: {cc_label}")
        r += 0  # sec already increments

        by_func = _group_by(cc_items, "function")
        cc_subtotal_rows: list[int] = []

        for func_key in sorted(by_func.keys()):
            func_items = by_func[func_key]
            func_code = func_key[0]

            by_prog = _group_by(func_items, "program")
            func_subtotal_rows: list[int] = []

            for prog_key in sorted(by_prog.keys()):
                prog_items = by_prog[prog_key]
                prog_code = prog_key[0]

                # Write each line item
                item_start = r
                for li in sorted(prog_items, key=lambda x: x.account_code):
                    ws.cell(r, 1, li.account_code).border = THIN
                    ws.cell(r, 2, li.description).border = THIN
                    for j, col_name in enumerate(all_cols, 3):
                        val = li.amounts.get(col_name)
                        if val is not None:
                            put(ws, r, j, val, USD2, fill=None)
                        else:
                            ws.cell(r, j).border = THIN

                    # Dollar difference (last year col - second-to-last)
                    if n_data_cols >= 2:
                        last_cl = get_column_letter(2 + n_data_cols)
                        prev_cl = get_column_letter(1 + n_data_cols)
                        diff_col = 2 + n_data_cols + 1
                        ws.cell(r, diff_col).value = (
                            f"={last_cl}{r}-{prev_cl}{r}"
                        )
                        ws.cell(r, diff_col).number_format = USD2
                        ws.cell(r, diff_col).border = THIN

                        pct_col = diff_col + 1
                        ws.cell(r, pct_col).value = (
                            f"=IF({prev_cl}{r}=0,0,{last_cl}{r}/{prev_cl}{r}-1)"
                        )
                        ws.cell(r, pct_col).number_format = PCT2
                        ws.cell(r, pct_col).border = THIN

                    r += 1
                item_end = r - 1

                # Program subtotal (only if more than 1 line item in program)
                if item_end > item_start:
                    ws.cell(r, 1).border = THICK_BOTTOM
                    ws.cell(r, 2, f"Program {prog_code} Total").font = BOLD
                    ws.cell(r, 2).fill = SUMMARY_FILL
                    ws.cell(r, 2).border = THICK_BOTTOM
                    for j in range(3, total_cols + 1):
                        cl = get_column_letter(j)
                        ws.cell(r, j).value = (
                            f"=SUM({cl}{item_start}:{cl}{item_end})"
                        )
                        ws.cell(r, j).number_format = USD2
                        ws.cell(r, j).font = BOLD
                        ws.cell(r, j).fill = SUMMARY_FILL
                        ws.cell(r, j).border = THICK_BOTTOM
                    func_subtotal_rows.append(r)
                    r += 1

            # Function subtotal (if multiple programs)
            if len(by_prog) > 1 and func_subtotal_rows:
                ws.cell(r, 2, f"Function {func_code} Total").font = BOLD
                ws.cell(r, 2).fill = CALC_FILL
                ws.cell(r, 2).border = THICK_BOTTOM
                ws.cell(r, 1).border = THICK_BOTTOM
                for j in range(3, total_cols + 1):
                    cl = get_column_letter(j)
                    refs = "+".join(f"{cl}{row}" for row in func_subtotal_rows)
                    ws.cell(r, j).value = f"={refs}"
                    ws.cell(r, j).number_format = USD2
                    ws.cell(r, j).font = BOLD
                    ws.cell(r, j).fill = CALC_FILL
                    ws.cell(r, j).border = THICK_BOTTOM
                cc_subtotal_rows.append(r)
                r += 1
            elif func_subtotal_rows:
                cc_subtotal_rows.extend(func_subtotal_rows)

        # Cost Center total
        if cc_subtotal_rows:
            ws.cell(r, 2, f"Cost Center {cc_code} Total").font = RESULT_FONT
            ws.cell(r, 2).fill = RESULT_FILL
            ws.cell(r, 2).border = THICK_BOTTOM
            ws.cell(r, 1).border = THICK_BOTTOM
            for j in range(3, total_cols + 1):
                cl = get_column_letter(j)
                refs = "+".join(f"{cl}{row}" for row in cc_subtotal_rows)
                ws.cell(r, j).value = f"={refs}"
                ws.cell(r, j).number_format = USD2
                ws.cell(r, j).font = RESULT_FONT
                ws.cell(r, j).fill = RESULT_FILL
                ws.cell(r, j).border = THICK_BOTTOM
            grand_total_rows.append(r)
            r += 2  # extra blank row between cost centers

    # Grand total for the article
    if grand_total_rows:
        ws.cell(r, 2, f"ARTICLE {article_num} TOTAL").font = RESULT_FONT
        ws.cell(r, 2).fill = HEADER_FILL
        ws.cell(r, 2).border = THICK_BOTTOM
        ws.cell(r, 1).border = THICK_BOTTOM
        for j in range(3, total_cols + 1):
            cl = get_column_letter(j)
            refs = "+".join(f"{cl}{row}" for row in grand_total_rows)
            ws.cell(r, j).value = f"={refs}"
            ws.cell(r, j).number_format = USD2
            ws.cell(r, j).font = RESULT_FONT
            ws.cell(r, j).fill = HEADER_FILL
            ws.cell(r, j).border = THICK_BOTTOM


def build_all_article_sheets(
    wb: Workbook, baseline: VerifiedBaseline
) -> None:
    """Create RSU-format tabs for all articles with data."""
    for article_num in sorted(cfg.articles):
        build_article_sheet(wb, baseline, article_num)
