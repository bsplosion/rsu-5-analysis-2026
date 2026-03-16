"""Shared Excel styles and number formats.

Extracted from ``create_excel.py`` lines 26-46 so they can be reused
across RSU-format output, analysis sheets, and summary sheets.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Fonts ──────────────────────────────────────────────────────
BOLD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=14)
SECTION = Font(bold=True, size=12, color="1F4E79")
NOTE = Font(italic=True, size=10, color="666666")
SOURCE = Font(italic=True, size=10, color="0070C0")
RESULT_FONT = Font(bold=True, size=11, color="C00000")
WARN_FONT = Font(bold=True, size=11, color="FF6600")

# ── Fills ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")
RESULT_FILL = PatternFill("solid", fgColor="FCE4D6")
SOURCE_FILL = PatternFill("solid", fgColor="DAEEF3")
PARAM_FILL = PatternFill("solid", fgColor="FFFF00")
SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")
MISMATCH_FILL = PatternFill("solid", fgColor="FFC7CE")
VERIFIED_FILL = PatternFill("solid", fgColor="C6EFCE")

# ── Borders ────────────────────────────────────────────────────
THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
THICK_BOTTOM = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("medium"),
)

# ── Number Formats ─────────────────────────────────────────────
USD = '#,##0'
USD2 = '#,##0.00'
PCT = '0.0%'
PCT2 = '0.00%'
SIGNED = '+#,##0;-#,##0;0'

# ── Tab Colors ─────────────────────────────────────────────────
TAB_INPUT = "FFC000"
TAB_CALC = "70AD47"
TAB_RSU = "4472C4"
TAB_SUMMARY = "808080"
