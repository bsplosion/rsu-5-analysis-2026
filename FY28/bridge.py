"""Bridge FY27 analytical model sheets into per-FY workbooks.

Imports builder functions from legacy/fy27_analytical_model.py and runs them
against new workbooks, preserving the cross-sheet R namespace references
that the complex analytical logic depends on.

FY27 workbook gets:
  All I-sheets (Enrollment, Budget, Revenue, Tax, ECCosts, Assumptions,
  FY27Reductions, CostGrowth, FTE, DOEStaffing, Withdrawal, Equity)
  and C-sheets (Enrollment, Budget, Consumption, Revenue, Tax, CostPremium,
  FY27Scenarios, DOEStaffing, Sources)

FY28 workbook gets:
  C-FY28Projection, C-FY28Paths, C-MSConsol, C-ECAnalysis, C-RiskModel,
  C-Independence, C-Equity, C-LegalAnalysis
  (These build atop the FY27 I-sheets, so FY27 I-sheets are included too)
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl.workbook import Workbook

# Add project root to path so create_excel.py can be imported
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))


def _import_builders():
    """Import all builder functions from create_excel.py.

    Done as a function to avoid module-level side effects (create_excel.py
    has top-level workbook creation code at the bottom).
    """
    import importlib
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "fy27_analytical_model", _PROJECT_ROOT / "legacy" / "fy27_analytical_model.py",
        submodule_search_locations=[]
    )
    # We need to import only the functions, not execute the module's
    # top-level wb creation code. The module has a guard-less main block.
    # Read the file and exec only the function/class/constant definitions.
    source = (_PROJECT_ROOT / "legacy" / "fy27_analytical_model.py").read_text(encoding="utf-8")

    # Find where the main execution block starts (the wb = openpyxl.Workbook() line)
    lines = source.split("\n")
    cutoff = len(lines)
    for i, line in enumerate(lines):
        if line.strip() == "wb = openpyxl.Workbook()":
            cutoff = i
            break

    safe_source = "\n".join(lines[:cutoff])

    module_dict: dict = {}
    exec(compile(safe_source, str(_PROJECT_ROOT / "legacy" / "fy27_analytical_model.py"), "exec"), module_dict)

    return module_dict


_CE = None

def _get_ce():
    global _CE
    if _CE is None:
        _CE = _import_builders()
    return _CE


def bridge_fy27(wb: Workbook) -> None:
    """Add all FY27-relevant analytical sheets from create_excel.py."""
    ce = _get_ce()
    R = ce["R"]

    # Input sheets (dependency order)
    ce["build_i_enrollment"](wb)
    ce["build_i_budget"](wb)
    ce["build_i_revenue"](wb)
    ce["build_i_tax"](wb)
    ce["build_i_ec_costs"](wb)
    ce["build_i_assumptions"](wb)
    ce["build_i_fy27_reductions"](wb)
    ce["build_i_cost_growth"](wb)
    ce["build_i_fte"](wb)
    ce["build_i_doe_staffing"](wb)

    # Calc sheets for FY27 scope
    ce["build_c_enrollment"](wb)
    ce["build_c_budget"](wb)
    ce["build_c_consumption"](wb)
    ce["build_c_revenue"](wb)
    ce["build_c_tax"](wb)
    ce["build_c_cost_premium"](wb)
    ce["build_c_fy27_scenarios"](wb)
    ce["build_c_doe_staffing"](wb)

    # Sources
    ce["build_sources"](wb)


def bridge_fy28(wb: Workbook) -> None:
    """Add FY28-relevant analytical sheets from create_excel.py.

    Includes all I-sheets (since FY28 analysis depends on them) plus
    the FY28 projection and path analysis sheets.
    """
    ce = _get_ce()
    R = ce["R"]
    SN = ce["SN"]

    # Must build I-sheets first (they set R namespace values)
    ce["build_i_enrollment"](wb)
    ce["build_i_budget"](wb)
    ce["build_i_revenue"](wb)
    ce["build_i_tax"](wb)
    ce["build_i_ec_costs"](wb)
    ce["build_i_assumptions"](wb)
    ce["build_i_fy27_reductions"](wb)
    ce["build_i_cost_growth"](wb)
    ce["build_i_fte"](wb)
    ce["build_i_doe_staffing"](wb)
    ce["build_i_withdrawal"](wb)
    ce["build_i_equity"](wb)

    # C-sheets that establish baseline calculations
    ce["build_c_enrollment"](wb)
    ce["build_c_budget"](wb)
    ce["build_c_consumption"](wb)
    ce["build_c_revenue"](wb)
    ce["build_c_tax"](wb)
    ce["build_c_cost_premium"](wb)

    # FY27 scenarios (needed as inputs to FY28 projection)
    ce["build_c_fy27_scenarios"](wb)

    # FY28 analysis suite
    ce["build_c_ms_consolidation"](wb)
    ce["build_c_ec_analysis"](wb)
    ce["build_c_fy28_paths"](wb)
    ce["build_c_fy28_projection"](wb)
    ce["build_c_risk_model"](wb)

    # Equity and independence analysis
    ce["build_c_equity"](wb)
    ce["build_c_independence"](wb)
    ce["build_c_legal"](wb)

    # DOE staffing calc
    ce["build_c_doe_staffing"](wb)

    # Sources
    ce["build_sources"](wb)

    # Fix circular cross-references (same as create_excel.py main block)
    CFP = SN['cfp']
    C28 = SN['c28']

    ws28 = wb["C-FY28Projection"]
    ws28.cell(R.C28_PA_SAV, 2).value = f"={CFP}!B{R.FP_PA_SAV}"
    ws28.cell(R.C28_PB_SAV, 2).value = f"={CFP}!B{R.FP_PB_SAV}"

    wfp = wb["C-FY28Paths"]
    from openpyxl.utils import get_column_letter
    br = R.FP_FINAL_BUD
    wfp.cell(br, 2).value = f"={C28}!B{R.C28_BASE}"
    wfp.cell(br, 3).value = f"={C28}!B{R.C28_BASE}-B{R.FP_PA_SAV}"
    wfp.cell(br, 4).value = f"={C28}!B{R.C28_BASE}-B{R.FP_PB_SAV}"
    pr = R.FP_FINAL_PCT
    for c in range(2, 5):
        cl = get_column_letter(c)
        wfp.cell(pr, c).value = f"=({cl}{br}-{C28}!B{R.C28_FY27_SEL})/{C28}!B{R.C28_FY27_SEL}"
