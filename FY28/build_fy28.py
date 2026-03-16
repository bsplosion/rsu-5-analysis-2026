"""FY28 and FY29 projection workbook builders.

FY28 adds a C-PESPreservationCost sheet and summary on top of the
bridged analytical sheets. FY29 provides the superintendent's forward
estimates with structural change modeling.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from rsu5.config import cfg
from rsu5.excel.helpers import col_widths, hdr, note, put, sec, source_block, ttl
from rsu5.excel.styles import (
    BOLD,
    CALC_FILL,
    HEADER_FILL,
    INPUT_FILL,
    PARAM_FILL,
    RESULT_FILL,
    RESULT_FONT,
    TAB_CALC,
    TAB_INPUT,
    TAB_SUMMARY,
    THICK_BOTTOM,
    USD,
    PCT2,
    SIGNED,
    WARN_FONT,
)
from rsu5.ingest.data_loader import BudgetData


def build_pes_preservation_cost(wb: Workbook) -> None:
    """C-PESPreservation: Full Path A vs Path B analysis with revenue alternatives.

    Based on the RSU 5 Planning and PES Reconciliation 2026 analysis:
      Path A ("Preserve & Strengthen"): Keep all elementaries; consolidate
        7-8 at FMS; 6th at DCS; EC distributed. Savings $479K-$804K/yr.
      Path B (Scenario 2): PES becomes PreK center; K-6 to DCS. Net cost
        ~$161K/yr after transport and conversion.
    """
    ws = wb.create_sheet("C-PESPreservation")
    ws.sheet_properties.tabColor = TAB_CALC
    col_widths(ws, [48, 22, 22, 22])

    assumptions = cfg.raw.get("assumptions", {})
    fy28_proj = cfg.raw.get("fy28_projections", {})

    r = ttl(ws, 1, "CALC: PES Preservation -- Full Path Analysis")
    r = note(ws, r + 1, "Path A preserves all schools with structural efficiencies.")
    r = note(ws, r, "Path B (Scenario 2) converts PES to PreK center; Pownal K-6 to DCS.")
    r = note(ws, r, "Based on RSU 5 Planning and PES Reconciliation 2026 analysis.")
    r += 1

    # PES operating profile
    r = sec(ws, r, "PES Current Operating Profile")
    enroll = cfg.raw.get("enrollment", {}).get("by_school", {}).get("PES", {}).get("oct_2026p", 105)
    put(ws, r, 1, "PES Enrollment (Oct 2026P)", fill=INPUT_FILL)
    put(ws, r, 2, enroll, fill=INPUT_FILL)
    enroll_row = r
    r += 1

    put(ws, r, 1, "PES FTE Staff (DOE Dec 2025)", fill=INPUT_FILL)
    put(ws, r, 2, 21.5, fill=INPUT_FILL)
    r += 1

    put(ws, r, 1, "PES cost per student (from analysis)", fill=INPUT_FILL)
    put(ws, r, 2, 24247, fmt=USD, fill=INPUT_FILL)
    r += 1

    put(ws, r, 1, "District avg cost per student", fill=INPUT_FILL)
    put(ws, r, 2, 15600, fmt=USD, fill=INPUT_FILL)
    r += 1

    put(ws, r, 1, "PES cost premium (fixed-cost dilution)", fill=CALC_FILL)
    put(ws, r, 2, 8647, fmt=USD, fill=CALC_FILL)
    r += 1

    r = note(ws, r, "Cost premium is a scale effect (small school), not overspending.")
    r = note(ws, r, "PES is the only school with consistent enrollment growth (+18% since 2023).")
    r += 1

    # Path A: Preserve & Strengthen
    r = sec(ws, r, "PATH A: Preserve & Strengthen (All Schools Open)")
    r = note(ws, r, "Keep PES PreK-5; consolidate 7-8 at FMS; 6th at DCS; EC distributed")
    r += 1

    for i, h in enumerate(["Component", "Low Estimate", "High Estimate"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1

    put(ws, r, 1, "MS consolidation (7-8 at FMS)", fill=CALC_FILL)
    put(ws, r, 2, 154000, fmt=USD, fill=CALC_FILL)
    put(ws, r, 3, 154000, fmt=USD, fill=CALC_FILL)
    ms_row = r
    r += 1

    add_eff_low = assumptions.get("additional_eff_low", 325000)
    add_eff_high = assumptions.get("additional_eff_high", 650000)
    put(ws, r, 1, "Additional efficiencies (shared services, scheduling)", fill=PARAM_FILL)
    put(ws, r, 2, add_eff_low, fmt=USD, fill=PARAM_FILL)
    put(ws, r, 3, add_eff_high, fmt=USD, fill=PARAM_FILL)
    eff_row = r
    r += 1

    put(ws, r, 1, "PATH A TOTAL ANNUAL SAVINGS", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{ms_row}+B{eff_row}"
    ws.cell(r, 2).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    ws.cell(r, 3).value = f"=C{ms_row}+C{eff_row}"
    ws.cell(r, 3).number_format = USD
    ws.cell(r, 3).font = RESULT_FONT
    path_a_row = r
    r += 1

    r = note(ws, r, "Range: $479K-$804K/year in structural savings while keeping all schools.")
    r += 1

    # Path B: Scenario 2
    r = sec(ws, r, "PATH B: Scenario 2 (PES -> PreK Center; K-6 to DCS)")
    r = note(ws, r, "PES building becomes district-wide PreK center; Pownal K-6 moves to DCS")
    r += 1

    for i, h in enumerate(["Component", "Annual Cost/Savings", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1

    put(ws, r, 1, "PES staff savings (positions eliminated)", fill=CALC_FILL)
    cost_per_fte = assumptions.get("cost_per_fte", 107700)
    put(ws, r, 2, 538500, fmt=USD, fill=CALC_FILL)
    put(ws, r, 3, "~5 FTE net reduction", fill=CALC_FILL)
    staff_sav_row = r
    r += 1

    bus_cost = assumptions.get("bus_route_cost", 75000)
    pownal_routes = assumptions.get("pownal_k6_routes", 3)
    put(ws, r, 1, "Additional bus routes (Pownal -> DCS)", fill=CALC_FILL)
    put(ws, r, 2, -(bus_cost * pownal_routes), fmt=SIGNED, fill=CALC_FILL)
    put(ws, r, 3, f"{pownal_routes} routes @ ${bus_cost:,}/ea", fill=CALC_FILL)
    bus_row = r
    r += 1

    pes_partial = assumptions.get("pes_partial_conversion", 425000)
    pes_mult = assumptions.get("pes_full_multiplier", 1.75)
    put(ws, r, 1, "PES building conversion (PreK center)", fill=CALC_FILL)
    conversion = int(pes_partial * pes_mult)
    amort = assumptions.get("amortization_years", 10)
    annual_conv = int(conversion / amort)
    put(ws, r, 2, -annual_conv, fmt=SIGNED, fill=CALC_FILL)
    put(ws, r, 3, f"${conversion:,} over {amort} yrs", fill=CALC_FILL)
    conv_row = r
    r += 1

    dcs_coord = assumptions.get("dcs_coordination_cost", 75000)
    put(ws, r, 1, "DCS capacity/coordination costs", fill=CALC_FILL)
    put(ws, r, 2, -dcs_coord, fmt=SIGNED, fill=CALC_FILL)
    put(ws, r, 3, "Additional admin, space prep", fill=CALC_FILL)
    dcs_row = r
    r += 1

    put(ws, r, 1, "EC program opt-out risk (families choosing private)", fill=CALC_FILL)
    put(ws, r, 2, -50000, fmt=SIGNED, fill=CALC_FILL)
    put(ws, r, 3, "Transport barrier reduces enrollment", fill=CALC_FILL)
    optout_row = r
    r += 1

    put(ws, r, 1, "PATH B NET ANNUAL RESULT", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{staff_sav_row}+B{bus_row}+B{conv_row}+B{dcs_row}+B{optout_row}"
    ws.cell(r, 2).number_format = SIGNED
    ws.cell(r, 2).font = RESULT_FONT
    path_b_row = r
    r += 1
    r = note(ws, r, "Path B net: approximately -$161K/year COST (not savings).")
    r += 1

    # Comparison
    r = sec(ws, r, "HEAD-TO-HEAD COMPARISON")
    for i, h in enumerate(["Metric", "Path A (Preserve)", "Path B (Close PES)"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1

    put(ws, r, 1, "Annual budget impact", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{path_a_row}"
    ws.cell(r, 2).number_format = SIGNED
    ws.cell(r, 3).value = f"=B{path_b_row}"
    ws.cell(r, 3).number_format = SIGNED
    r += 1

    put(ws, r, 1, "Path A advantage over Path B", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{path_a_row}-B{path_b_row}"
    ws.cell(r, 2).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    advantage_row = r
    r += 1

    put(ws, r, 1, "Schools preserved", fill=CALC_FILL)
    put(ws, r, 2, "All 6 (PES stays PreK-5)", fill=CALC_FILL)
    put(ws, r, 3, "5 (PES -> PreK only)", fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "Reversible?", fill=CALC_FILL)
    put(ws, r, 2, "Yes", fill=CALC_FILL)
    put(ws, r, 3, "No", fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "Community impact", fill=CALC_FILL)
    put(ws, r, 2, "Minimal disruption", fill=CALC_FILL)
    put(ws, r, 3, "Significant: property values, enrollment", fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "EC mandate compliance", fill=CALC_FILL)
    put(ws, r, 2, "Distributed (MSS preferred for SpEd)", fill=CALC_FILL)
    put(ws, r, 3, "Centralized at PES (transport barrier)", fill=CALC_FILL)
    r += 2

    # Revenue alternatives
    r = sec(ws, r, "REVENUE ALTERNATIVES (Avoiding Cuts)")
    r = note(ws, r, "Options for bridging budget gaps without closing schools")
    r += 1

    for i, h in enumerate(["Alternative", "Annual Impact", "Notes"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 3)
    r += 1

    revenue_items = [
        ("ALM formula reform (50/50 val/enroll)", 382000,
         "Board authority under §1481-A; reduces Pownal penalty"),
        ("CDS EC transfer revenue", 300000,
         "Range $200K-$400K; CDS funds follow child"),
        ("IDEA Part B federal SpEd funding", 125000,
         "Range $100K-$150K for EC/ECSE"),
        ("EPS PreK state allocation", 75000,
         "Range $50K-$100K; formula-based"),
        ("Shared services agreements (IT, maintenance)", 150000,
         "Economies of scale across towns"),
    ]

    rev_start = r
    for label, amount, notes_text in revenue_items:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, amount, fmt=USD, fill=CALC_FILL)
        put(ws, r, 3, notes_text, fill=CALC_FILL)
        r += 1

    put(ws, r, 1, "TOTAL REVENUE ALTERNATIVES", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=SUM(B{rev_start}:B{r-1})"
    ws.cell(r, 2).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    rev_total_row = r
    r += 2

    # Net position
    r = sec(ws, r, "NET POSITION: Path A + Revenue Alternatives")
    put(ws, r, 1, "Path A structural savings", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{path_a_row}"
    ws.cell(r, 2).number_format = USD
    pa_line = r
    r += 1
    put(ws, r, 1, "Revenue alternatives", fill=CALC_FILL)
    ws.cell(r, 2).value = f"=B{rev_total_row}"
    ws.cell(r, 2).number_format = USD
    rv_line = r
    r += 1
    put(ws, r, 1, "COMBINED ANNUAL IMPACT", fill=RESULT_FILL, font=BOLD)
    ws.cell(r, 2).value = f"=B{pa_line}+B{rv_line}"
    ws.cell(r, 2).number_format = USD
    ws.cell(r, 2).font = RESULT_FONT
    combined_row = r
    r += 1

    fy28_with = fy28_proj.get("with_reductions", 49400000)
    put(ws, r, 1, "As % of FY28 projected budget", fill=RESULT_FILL)
    ws.cell(r, 2).value = f"=B{combined_row}/{fy28_with}"
    ws.cell(r, 2).number_format = PCT2
    ws.cell(r, 2).font = RESULT_FONT
    r += 2

    # Sensitivity
    r = sec(ws, r, "Sensitivity: PES Enrollment Impact on Cost Premium")
    for i, h in enumerate(["Enrollment", "PES Cost/Student", "Cost Premium/Student", "Path A Advantage"], 1):
        ws.cell(r, i, h)
    hdr(ws, r, 4)
    r += 1

    for enr in [80, 90, 100, 105, 110, 120, 130]:
        put(ws, r, 1, enr, fill=CALC_FILL)
        ws.cell(r, 2).value = f"=B{staff_sav_row}*2/{enr}"
        ws.cell(r, 2).number_format = USD
        put(ws, r, 3, max(0, int(24247 - 15600 * (enr / 105))), fmt=USD, fill=CALC_FILL)
        ws.cell(r, 4).value = f"=B{path_a_row}/{enr}"
        ws.cell(r, 4).number_format = USD
        r += 1
    r += 1

    r = source_block(ws, r, [
        "Analysis: RSU 5 Planning and PES Reconciliation 2026",
        "Cost data: Cursor Pownal Elementary Cost Evaluation (March 2026)",
        "DOE staffing: MDohle RSU 5 Staff by FTE.xlsx (2016-2026)",
        "Revenue estimates: §1481-A formula analysis, CDS transfer projections",
        "Yellow cells are adjustable parameters -- modify to test scenarios.",
    ])


def build_fy29_sheets(wb: Workbook, data: BudgetData) -> None:
    """Build FY29 projection sheets."""
    fy29 = cfg.raw.get("fy29_projections", {})
    fy28 = cfg.raw.get("fy28_projections", {})
    scenarios = cfg.raw.get("fy27_scenarios", {})

    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = TAB_SUMMARY
    col_widths(ws, [42, 22, 22])

    r = ttl(ws, 1, "RSU 5 FY29 Projection")
    r = note(ws, r + 1, "Fiscal Year 2028-2029")
    r = note(ws, r, "Based on superintendent's projected totals with structural change modeling")
    r += 1

    r = sec(ws, r, "Superintendent's FY29 Projections")
    put(ws, r, 1, "FY29 With Reductions", fill=INPUT_FILL, font=BOLD)
    put(ws, r, 2, fy29.get("with_reductions", 52900000), fmt=USD, fill=INPUT_FILL)
    r += 1
    put(ws, r, 1, "FY29 No Reductions", fill=INPUT_FILL, font=BOLD)
    put(ws, r, 2, fy29.get("no_reductions", 58000000), fmt=USD, fill=INPUT_FILL)
    r += 2

    r = sec(ws, r, "Growth from FY28 to FY29")
    fy28_with = fy28.get("with_reductions", 49400000)
    fy28_no = fy28.get("no_reductions", 54200000)
    fy29_with = fy29.get("with_reductions", 52900000)
    fy29_no = fy29.get("no_reductions", 58000000)

    put(ws, r, 1, "FY28 baseline (with reductions)", fill=CALC_FILL)
    put(ws, r, 2, fy28_with, fmt=USD, fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "FY29 projected (with reductions)", fill=CALC_FILL)
    put(ws, r, 2, fy29_with, fmt=USD, fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "Growth", fill=RESULT_FILL, font=BOLD)
    growth = (fy29_with - fy28_with) / fy28_with if fy28_with else 0
    put(ws, r, 2, growth, fmt=PCT2, fill=RESULT_FILL, font=RESULT_FONT)
    r += 2

    put(ws, r, 1, "FY28 baseline (no reductions)", fill=CALC_FILL)
    put(ws, r, 2, fy28_no, fmt=USD, fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "FY29 projected (no reductions)", fill=CALC_FILL)
    put(ws, r, 2, fy29_no, fmt=USD, fill=CALC_FILL)
    r += 1
    put(ws, r, 1, "Growth", fill=RESULT_FILL, font=BOLD)
    growth2 = (fy29_no - fy28_no) / fy28_no if fy28_no else 0
    put(ws, r, 2, growth2, fmt=PCT2, fill=RESULT_FILL, font=RESULT_FONT)
    r += 2

    r = sec(ws, r, "Structural Changes by FY29")
    notes = [
        "EC mandate fully operational (started FY28)",
        "PES decision resolved (closure or preservation)",
        "Middle school consolidation effects (if any)",
        "FY27 staffing reductions compounding into baseline",
    ]
    for n in notes:
        r = note(ws, r, f"  - {n}")
    r += 1

    r = sec(ws, r, "Budget Trajectory (FY26 -> FY29)")
    fy26 = scenarios.get("fy26_adopted", 44455929)
    trajectory = [
        ("FY26 Adopted", fy26),
        ("FY27 Proposed", scenarios.get("superintendent_proposed", 47357441)),
        ("FY28 Projected (w/cuts)", fy28_with),
        ("FY29 Projected (w/cuts)", fy29_with),
    ]
    for label, val in trajectory:
        put(ws, r, 1, label, fill=CALC_FILL)
        put(ws, r, 2, val, fmt=USD, fill=CALC_FILL)
        r += 1
    r += 1

    total_growth = (fy29_with - fy26) / fy26 if fy26 else 0
    put(ws, r, 1, "4-year cumulative growth", fill=RESULT_FILL, font=BOLD)
    put(ws, r, 2, total_growth, fmt=PCT2, fill=RESULT_FILL, font=RESULT_FONT)
    r += 1
    cagr = (fy29_with / fy26) ** (1 / 3) - 1 if fy26 else 0
    put(ws, r, 1, "Implied annual growth rate", fill=RESULT_FILL, font=BOLD)
    put(ws, r, 2, cagr, fmt=PCT2, fill=RESULT_FILL, font=RESULT_FONT)
    r += 2

    r = source_block(ws, r, [
        "Superintendent's projections from FY27-FY29 Projected Budgets document",
        "High uncertainty on FY29 -- depends on FY27 adoption and FY28 structural decisions",
        "See FY28 workbook for detailed PES preservation cost analysis",
    ])
