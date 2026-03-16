"""Build per-FY RSU 5 Budget Analysis workbooks.

Emits one workbook per fiscal year into /FYxx/ folders, each containing:
  - Summary tab (verification status, article breakdown)
  - P-BudgetData (parsed article x cost center matrix)
  - P-Enrollment, P-DOEStaffing, P-CostGrowth (parsed from handbook/DOE)
  - Art 1-11 (RSU-format line-item detail)
  - C-CostPerStudent, C-Verification

FY27 additionally bridges the analytical model from legacy/create_excel.py.
FY28/FY29 projections use FY28-specific builders in FY28/.

Usage::

    python build_workbook.py                 # all FYs, FY22-FY29
    python build_workbook.py --fy 27         # single FY
    python build_workbook.py --fy 26 27      # specific FYs
    python build_workbook.py --all-docs      # ingest all docs per FY
    python build_workbook.py --dry-run       # ingest + reconcile only
"""

from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

import openpyxl

# Ensure project root is importable (for FY28/ package)
_ROOT = Path(__file__).resolve().parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from rsu5.ingest.data_loader import BudgetData
from rsu5.reconcile import reconcile
from rsu5.excel.rsu_format import build_all_article_sheets
from rsu5.excel.analysis import build_analysis_sheets
from rsu5.excel.summary import build_summary_sheet
from rsu5.excel.verification import build_verification


def _build_fy27_workbook(
    data: BudgetData,
    baseline,
    all_baselines: dict,
    output_dir: Path,
) -> Path:
    """Build FY27 workbook: bridge first (clean names), then parsed data sheets.

    The bridged analytical model from legacy/create_excel.py gets clean sheet
    names so its cross-sheet formulas work. The data-pipeline sheets (parsed
    from source documents) use "P-" prefix for clarity.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    from FY28.bridge import bridge_fy27
    bridge_fy27(wb)

    from rsu5.excel.analysis import (
        build_i_enrollment, build_i_doe_staffing, build_i_cost_growth,
        build_i_budget_data, build_c_cost_per_student,
    )

    if baseline:
        build_i_budget_data(wb, baseline)
        _ws = wb["I-BudgetData"]
        _ws.title = "P-BudgetData"
        _ws.sheet_properties.tabColor = "4472C4"

    build_i_enrollment(wb, 27, data)
    wb["I-Enrollment1"].title = "P-Enrollment"
    wb["P-Enrollment"].sheet_properties.tabColor = "4472C4"

    build_i_doe_staffing(wb, 27, data)
    wb["I-DOEStaffing1"].title = "P-DOEStaffing"
    wb["P-DOEStaffing"].sheet_properties.tabColor = "4472C4"

    build_i_cost_growth(wb, 27, data)
    wb["I-CostGrowth1"].title = "P-CostGrowth"
    wb["P-CostGrowth"].sheet_properties.tabColor = "4472C4"

    build_c_cost_per_student(wb, 27, data, baseline)

    if baseline and baseline.line_items:
        build_all_article_sheets(wb, baseline)

    build_verification(wb, 27, data, baseline)
    build_summary_sheet(wb, 27, data, baseline)

    sheet_map = {ws.title: ws for ws in wb.worksheets}
    priority = ["Summary", "P-BudgetData", "P-Enrollment", "P-DOEStaffing",
                "P-CostGrowth", "C-CostPerStudent", "C-Verification"]
    art_sheets = [f"Art {i}" for i in range(1, 12)]
    first = [sheet_map[n] for n in priority if n in sheet_map]
    arts = [sheet_map[n] for n in art_sheets if n in sheet_map]
    rest = [ws for ws in wb.worksheets if ws not in first and ws not in arts]
    wb._sheets = first + arts + rest

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "RSU5_FY27.xlsx"
    wb.save(str(out_path))
    return out_path


def _build_fy_workbook(
    fy: int,
    data: BudgetData,
    baseline,
    all_baselines: dict,
    output_dir: Path,
) -> Path:
    """Build a single historical FY workbook (FY22-FY26)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, fy, data, baseline)
    build_analysis_sheets(wb, fy, data, baseline, all_baselines)

    if baseline and baseline.line_items:
        build_all_article_sheets(wb, baseline)

    build_verification(wb, fy, data, baseline)

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"RSU5_FY{fy}.xlsx"
    wb.save(str(out_path))
    return out_path


def _build_fy28_workbook(data: BudgetData, output_dir: Path) -> Path:
    """Build the FY28 projection workbook with PES preservation analysis."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    from FY28.bridge import bridge_fy28
    bridge_fy28(wb)

    from FY28.build_fy28 import build_pes_preservation_cost
    build_pes_preservation_cost(wb)

    sheet_map = {ws.title: ws for ws in wb.worksheets}
    priority_front = ["C-PESPreservation", "C-FY28Projection", "C-FY28Paths"]
    front = [sheet_map[n] for n in priority_front if n in sheet_map]
    rest = [ws for ws in wb.worksheets if ws not in front]
    wb._sheets = front + rest

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "RSU5_FY28_Projection.xlsx"
    wb.save(str(out_path))
    return out_path


def _build_fy29_workbook(data: BudgetData, output_dir: Path) -> Path:
    """Build the FY29 projection workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    from FY28.build_fy28 import build_fy29_sheets
    build_fy29_sheets(wb, data)

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "RSU5_FY29_Projection.xlsx"
    wb.save(str(out_path))
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build per-FY RSU 5 budget analysis workbooks."
    )
    parser.add_argument(
        "--fy", type=int, nargs="*",
        help="Fiscal years to build (2-digit, e.g. 27). Default: 22-27.",
    )
    parser.add_argument(
        "--all-docs", action="store_true",
        help="Ingest all available documents per FY (not just preferred).",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Run ingestion and reconciliation only; skip Excel generation.",
    )
    args = parser.parse_args()

    t0 = time.time()
    root = Path(__file__).parent

    default_fys = list(range(22, 28))
    fys_to_build = args.fy if args.fy else default_fys

    print("=" * 60)
    print("STEP 1: Ingesting all data sources")
    print("=" * 60)
    data = BudgetData.load(
        fys=fys_to_build,
        preferred_only=not args.all_docs,
    )
    print(
        f"\nIngested {len(data.line_items)} line items, "
        f"{len(data.summary_rows)} summary rows "
        f"across FYs {data.fiscal_years()}"
    )

    print()
    print("=" * 60)
    print("STEP 2: Reconciliation (independent verification)")
    print("=" * 60)
    baselines = reconcile(data)

    clean_count = sum(1 for bl in baselines.values() if bl.is_clean)
    total_count = len(baselines)
    print(
        f"\nReconciliation: {clean_count}/{total_count} FYs fully verified"
    )

    if args.dry_run:
        print("\n[dry-run] Skipping Excel generation.")
        return

    print()
    print("=" * 60)
    print("STEP 3: Generating per-FY workbooks")
    print("=" * 60)

    built_count = 0
    for fy in sorted(fys_to_build):
        baseline = baselines.get(fy)
        output_dir = root / f"FY{fy}"

        print(f"\n  Building FY{fy}...")

        if fy == 27:
            try:
                out_path = _build_fy27_workbook(
                    data, baseline, baselines, output_dir,
                )
                wb_tmp = openpyxl.load_workbook(str(out_path))
                print(f"    + Bridged analytical model + parsed data sheets")
                print(f"    -> {out_path} ({len(wb_tmp.sheetnames)} sheets)")
                wb_tmp.close()
            except Exception as e:
                print(f"    WARN: FY27 special build failed: {e}")
                import traceback; traceback.print_exc()
                out_path = _build_fy_workbook(
                    fy, data, baseline, baselines, output_dir,
                )
                wb_tmp = openpyxl.load_workbook(str(out_path))
                print(f"    -> {out_path} ({len(wb_tmp.sheetnames)} sheets) [fallback]")
                wb_tmp.close()
        else:
            out_path = _build_fy_workbook(
                fy, data, baseline, baselines, output_dir,
            )
            wb_tmp = openpyxl.load_workbook(str(out_path))
            print(f"    -> {out_path} ({len(wb_tmp.sheetnames)} sheets)")
            wb_tmp.close()

        built_count += 1

    # FY28 projection
    print(f"\n  Building FY28 projection...")
    try:
        out28 = _build_fy28_workbook(data, root / "FY28")
        wb28 = openpyxl.load_workbook(str(out28))
        print(f"    -> {out28} ({len(wb28.sheetnames)} sheets)")
        wb28.close()
        built_count += 1
    except Exception as e:
        print(f"    WARN: FY28 build failed: {e}")
        import traceback; traceback.print_exc()

    # FY29 projection
    print(f"\n  Building FY29 projection...")
    try:
        out29 = _build_fy29_workbook(data, root / "FY29")
        wb29 = openpyxl.load_workbook(str(out29))
        print(f"    -> {out29} ({len(wb29.sheetnames)} sheets)")
        wb29.close()
        built_count += 1
    except Exception as e:
        print(f"    WARN: FY29 build failed: {e}")

    # Cross-year overview (FY17-FY29)
    print(f"\n  Building cross-year overview (FY17-FY29)...")
    try:
        from rsu5.excel.overview import build_overview_workbook
        out_ov = build_overview_workbook(
            data, baselines,
            root / "RSU5 Overview FY17-FY29.xlsx",
        )
        wb_ov = openpyxl.load_workbook(str(out_ov))
        print(f"    -> {out_ov} ({len(wb_ov.sheetnames)} sheets)")
        wb_ov.close()
        built_count += 1
    except Exception as e:
        print(f"    WARN: Overview build failed: {e}")

    elapsed = time.time() - t0
    print(f"\nDone in {elapsed:.1f}s. Built {built_count} workbooks.")


if __name__ == "__main__":
    main()
